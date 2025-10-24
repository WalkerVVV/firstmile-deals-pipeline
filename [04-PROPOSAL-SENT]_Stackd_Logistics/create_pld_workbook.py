import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def calculate_billable_weight_oz(weight_lb):
    """
    Calculate billable weight in ounces following carrier rounding rules:
    - Under 1 lb: Round UP to next whole oz, MAX 15.99 oz
    - 16 oz exactly: Bills as 1 lb (16 oz)
    - Over 1 lb: Round UP to next whole pound (32, 48, 64 oz, etc.)
    """
    if pd.isna(weight_lb) or weight_lb <= 0:
        return 0

    weight_oz = weight_lb * 16

    if weight_oz < 16:
        # Under 1 lb: round up to next whole oz, max 15.99
        return min(int(np.ceil(weight_oz)), 15)
    elif weight_oz == 16:
        # Exactly 16 oz bills as 16 oz (1 lb)
        return 16
    else:
        # Over 1 lb: round up to next whole pound
        billable_lbs = int(np.ceil(weight_lb))
        return billable_lbs * 16

def format_dimensions(length, width, height):
    """Format dimensions as LxWxH string"""
    try:
        l = float(length) if pd.notna(length) else 0
        w = float(width) if pd.notna(width) else 0
        h = float(height) if pd.notna(height) else 0

        if l == 0 and w == 0 and h == 0:
            return "(blank)"

        # Sort dimensions largest to smallest
        dims = sorted([l, w, h], reverse=True)
        return f"{dims[0]:.0f}x{dims[1]:.0f}x{dims[2]:.0f}"
    except:
        return "(blank)"

print("Loading Stackd Logistics data...")
df = pd.read_csv("20250918193042_221aaf59f30469602caf8f7f7485b114.csv")

df["Order date"] = pd.to_datetime(df["Order date"], errors="coerce")
df["Weight (lb)"] = pd.to_numeric(df["Weight (lb)"], errors="coerce")
df["Length (in)"] = pd.to_numeric(df["Length (in)"], errors="coerce")
df["Width (in)"] = pd.to_numeric(df["Width (in)"], errors="coerce")
df["Height (in)"] = pd.to_numeric(df["Height (in)"], errors="coerce")

date_range_days = (df["Order date"].max() - df["Order date"].min()).days + 1
weeks_in_period = 4  # User-specified period for analysis

def map_service(method):
    method_lower = str(method).lower()
    if "2nd day" in method_lower:
        return "Expedited"
    elif "ground" in method_lower or "parcel expedited" in method_lower:
        return "Ground"
    elif "priority" in method_lower or "express" in method_lower:
        return "Priority"
    else:
        return "Ground"

df["Service Category"] = df["Shipping Method"].apply(map_service)
df["Bill Weight (oz)"] = df["Weight (lb)"].apply(calculate_billable_weight_oz)
df["DIMS"] = df.apply(lambda x: format_dimensions(x["Length (in)"], x["Width (in)"], x["Height (in)"]), axis=1)

# Country categorization
df["Country Category"] = df["Country"].apply(lambda x: "US" if str(x).upper() in ["US", "USA", "UNITED STATES"] else "(blank)" if pd.isna(x) else "International")

print("Creating pivot tables...")

# PRIMARY PIVOT: Weight × Service Level Matrix (for tier tool)
# Create pivot with actual data
pivot_weight_service = pd.pivot_table(df, values="Shipping Label ID",
                                      index="Bill Weight (oz)",
                                      columns="Service Category",
                                      aggfunc="count",
                                      fill_value=0)

# Ensure all service columns exist
for svc in ["Priority", "Expedited", "Ground"]:
    if svc not in pivot_weight_service.columns:
        pivot_weight_service[svc] = 0

pivot_weight_service = pivot_weight_service[["Priority", "Expedited", "Ground"]]

# CREATE COMPLETE WEIGHT BREAK TEMPLATE
# 1-15 oz (under 1 lb), then 15.99, then 16, 32, 48, 64... up to 400 oz
weight_breaks = list(range(1, 16))  # 1-15 oz
weight_breaks.append(15.99)  # 15.99 oz
weight_breaks.extend([16 * i for i in range(1, 26)])  # 16-400 oz (1-25 lbs in 16 oz increments)

# Create DataFrame with all weight breaks
template_df = pd.DataFrame(index=weight_breaks, columns=["Priority", "Expedited", "Ground"])
template_df = template_df.fillna(0)

# Merge actual data into template
for weight in pivot_weight_service.index:
    if weight in template_df.index:
        template_df.loc[weight] = pivot_weight_service.loc[weight]

# Calculate totals
template_df["Total Volume"] = template_df["Priority"] + template_df["Expedited"] + template_df["Ground"]
total_shipments = template_df["Total Volume"].sum()
template_df["% of Total"] = (template_df["Total Volume"] / total_shipments * 100).round(0).astype(int)
template_df["Avg per Week"] = (template_df["Total Volume"] / weeks_in_period).round(0).astype(int)

template_df = template_df.reset_index()
template_df.columns = ["Weight", "Priority", "Expedited", "Ground", "Total Volume", "% of Total", "Avg per Week"]

# Pivot 2: Dimensions
pivot_dims = pd.pivot_table(df, values="Shipping Label ID", index="DIMS",
                            aggfunc="count", fill_value=0)
pivot_dims.columns = ["COUNT"]
pivot_dims["%"] = (pivot_dims["COUNT"] / pivot_dims["COUNT"].sum() * 100).round(2)
pivot_dims = pivot_dims.reset_index()
pivot_dims.columns = ["DIMS", "COUNT", "%"]
pivot_dims = pivot_dims.sort_values("COUNT", ascending=False).reset_index(drop=True)

# Add Grand Total row
dims_total = pd.DataFrame([["Grand Total", pivot_dims["COUNT"].sum(), 100.00]],
                          columns=["DIMS", "COUNT", "%"])
pivot_dims = pd.concat([pivot_dims, dims_total], ignore_index=True)

# Pivot 3: Country (US vs International)
pivot_country = pd.pivot_table(df, values="Shipping Label ID", index="Country Category",
                               aggfunc="count", fill_value=0)
pivot_country.columns = ["COUNT"]
pivot_country["%"] = (pivot_country["COUNT"] / pivot_country["COUNT"].sum() * 100).round(2)
pivot_country = pivot_country.reset_index()
pivot_country.columns = ["Country", "COUNT", "%"]

# Add Grand Total row
country_total = pd.DataFrame([["Grand Total", pivot_country["COUNT"].sum(), 100.00]],
                             columns=["Country", "COUNT", "%"])
pivot_country = pd.concat([pivot_country, country_total], ignore_index=True)

# Pivot 4: State Distribution
pivot_state = pd.pivot_table(df, values="Shipping Label ID", index="State",
                             aggfunc="count", fill_value=0)
pivot_state.columns = ["COUNT"]
pivot_state["%"] = (pivot_state["COUNT"] / pivot_state["COUNT"].sum() * 100).round(2)
pivot_state = pivot_state.reset_index()
pivot_state.columns = ["State", "COUNT", "%"]
pivot_state = pivot_state.sort_values("COUNT", ascending=False).reset_index(drop=True)

total_volume = df.shape[0]
service_mix = {
    "Priority": df[df["Service Category"] == "Priority"].shape[0] / total_volume,
    "Expedited": df[df["Service Category"] == "Expedited"].shape[0] / total_volume,
    "Ground": df[df["Service Category"] == "Ground"].shape[0] / total_volume
}

# Calculate weight band distribution
def categorize_weight_band(weight_oz):
    if weight_oz <= 5:
        return "1-5 oz"
    elif weight_oz <= 10:
        return "6-10 oz"
    elif weight_oz <= 15:
        return "11-15 oz"
    elif weight_oz <= 80:  # 1-5 lbs
        return "1-5 lbs"
    else:
        return "5+ lbs"

df["Weight Band"] = df["Bill Weight (oz)"].apply(categorize_weight_band)
weight_bands = df["Weight Band"].value_counts(normalize=True).sort_index()

print("Creating Excel workbook...")
wb = Workbook()

# SHEET 1: Paste Data - Weight × Service Matrix (PRIMARY OUTPUT FOR TIER TOOL)
ws1 = wb.active
ws1.title = "Paste Data"

# Styling
header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
ground_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
border = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD')
)

# Write headers for Weight × Service matrix
headers = ["Weight", "Priority", "Expedited", "Ground", "Total Volume", "% of Total", "Avg per Week"]
for col_idx, header in enumerate(headers, start=1):
    cell = ws1.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

# Write data rows - ALL weight breaks
for row_idx, row_data in enumerate(template_df.values, start=2):
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws1.cell(row=row_idx, column=col_idx)

        # Format weight column to show 15.99 correctly
        if col_idx == 1:
            if val == 15.99:
                cell.value = 15.99
            else:
                cell.value = int(val)
        else:
            cell.value = int(val) if isinstance(val, (int, float)) else val

        cell.alignment = Alignment(horizontal="center")
        cell.border = border

        # Highlight Ground column (column D)
        if col_idx == 4:
            cell.fill = ground_fill

        # Format % of Total column
        if col_idx == 6:
            cell.value = f"{int(val)}%"

# Add Summary panel in columns H-I
ws1["H1"] = "Summary"
ws1["H1"].fill = header_fill
ws1["H1"].font = Font(color="FFFFFF", size=14, bold=True)
ws1["H1"].alignment = Alignment(horizontal="center")

ws1["H2"] = "Total Volume (All XP Methods)"
ws1["I2"] = total_volume
ws1["H3"] = "Weeks in period"
ws1["I3"] = round(weeks_in_period, 1)
ws1["H4"] = "Avg Volume / Week"
ws1["I4"] = round(total_volume / weeks_in_period, 0)

ws1["H5"] = "Assumption if mix missing"
ws1["I5"] = "Default = 100% Expedited"

ws1["H7"] = "Priority %"
ws1["I7"] = f"{service_mix['Priority']:.0%}"
ws1["I7"].fill = yellow_fill
ws1["H8"] = "Expedited %"
ws1["I8"] = f"{service_mix['Expedited']:.0%}"
ws1["I8"].fill = yellow_fill
ws1["H9"] = "Ground %"
ws1["I9"] = f"{service_mix['Ground']:.0%}"
ws1["I9"].fill = yellow_fill

# Weight Band distribution
ws1["H12"] = "Weight Band"
ws1["H12"].fill = header_fill
ws1["H12"].font = header_font
ws1["I12"] = "%"
ws1["I12"].fill = header_fill
ws1["I12"].font = header_font

row_num = 13
for band in ["1-5 oz", "6-10 oz", "11-15 oz", "1-5 lbs", "5+ lbs"]:
    ws1[f"H{row_num}"] = band
    pct = weight_bands.get(band, 0)
    ws1[f"I{row_num}"] = f"{pct:.0%}"
    row_num += 1

# Auto-size columns A-G
for col in range(1, 8):
    ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = 15

ws1.column_dimensions['H'].width = 25
ws1.column_dimensions['I'].width = 25

# SHEET 2: PLD Volume Analysis - All 4 pivots side-by-side
ws2 = wb.create_sheet("PLD Volume Analysis")

# Helper function to write pivot table
def write_pivot(ws, pivot_df, start_col, title):
    col_idx = start_col

    # Write headers
    for col_name in pivot_df.columns:
        cell = ws.cell(row=4, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        col_idx += 1

    # Write data
    for row_idx, row_data in enumerate(pivot_df.values, start=5):
        col_idx = start_col
        for val in row_data:
            cell = ws.cell(row=row_idx, column=col_idx)

            # Format based on column type
            if isinstance(val, (int, float)) and col_idx != start_col:
                if pivot_df.columns[col_idx - start_col].endswith("%"):
                    cell.value = f"{val:.2f}%"
                else:
                    cell.value = val
            else:
                cell.value = val

            cell.alignment = Alignment(horizontal="center")
            cell.border = border

            # Highlight Grand Total row
            if row_idx == 5 + len(pivot_df) - 1:
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.font = Font(bold=True)

            col_idx += 1

# Write all 4 pivot tables
pivot_weight_for_analysis = template_df.copy()
pivot_weight_for_analysis.columns = ["Bill Weight", "Priority", "Expedited", "Ground", "COUNT", "%", "Avg"]
pivot_weight_for_analysis["COUNT"] = pivot_weight_for_analysis["Priority"] + pivot_weight_for_analysis["Expedited"] + pivot_weight_for_analysis["Ground"]
pivot_weight_for_analysis["%"] = (pivot_weight_for_analysis["COUNT"] / pivot_weight_for_analysis["COUNT"].sum() * 100).round(2)
pivot_weight_for_analysis = pivot_weight_for_analysis[["Bill Weight", "COUNT", "%"]]

write_pivot(ws2, pivot_weight_for_analysis, 1, "Bill Weight")  # Columns A-C
write_pivot(ws2, pivot_dims, 5, "Dimensions")     # Columns E-G
write_pivot(ws2, pivot_country, 9, "Country")     # Columns I-K
write_pivot(ws2, pivot_state, 13, "State")        # Columns M-O

# Auto-size columns
for col in range(1, 20):
    ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = 15

# SHEET 3: Pivot Paste Configuration
ws3 = wb.create_sheet("Pivot Paste")

ws3["A1"] = "Weeks in period (enter number)"
ws3["D1"] = round(weeks_in_period, 1)
ws3["D1"].fill = yellow_fill

ws3["A3"] = "Service Mix (decimals sum to 1)"
ws3["A4"] = "Priority %"
ws3["D4"] = f"{service_mix['Priority']:.0%}"
ws3["D4"].fill = yellow_fill
ws3["A5"] = "Expedited %"
ws3["D5"] = f"{service_mix['Expedited']:.0%}"
ws3["D5"].fill = yellow_fill
ws3["A6"] = "Ground %"
ws3["D6"] = f"{service_mix['Ground']:.0%}"
ws3["D6"].fill = yellow_fill

ws3["A8"] = "Check: mix sums to 1"
ws3["D8"] = round(sum(service_mix.values()), 1)

ws3["A10"] = "Paste instructions"
ws3["D10"] = "Copy columns A-G from 'Paste Data' sheet into FirstMile tier tool"

ws3["A12"] = "Common period to weeks"
ws3["A13"] = "1 week"
ws3["D13"] = 1
ws3["A14"] = "2 weeks"
ws3["D14"] = 2
ws3["A15"] = "1 month"
ws3["D15"] = 4.345
ws3["A16"] = "3 months"
ws3["D16"] = 13.035

output_file = "Stackd_Logistics_PLD_4WEEKS.xlsx"
wb.save(output_file)

print(f"\nSUCCESS: {output_file} created")
print(f"  - Total volume: {total_volume:,} shipments")
print(f"  - Period: {weeks_in_period:.1f} weeks ({date_range_days} days)")
print(f"  - Service mix: Priority {service_mix['Priority']:.0%}, Expedited {service_mix['Expedited']:.0%}, Ground {service_mix['Ground']:.0%}")
print(f"\n  Weight Band Distribution:")
for band in ["1-5 oz", "6-10 oz", "11-15 oz", "1-5 lbs", "5+ lbs"]:
    pct = weight_bands.get(band, 0)
    print(f"    {band}: {pct:.1%}")
print(f"\n  Top 10 States:")
for _, row in pivot_state.head(10).iterrows():
    print(f"    {row['State']}: {row['COUNT']:,} ({row['%']:.2f}%)")
print(f"\n[OK] Sheet 1 'Paste Data' contains ALL weight breaks 1-15.99, 16-400 oz")
print(f"[OK] Ready to copy/paste columns A-G into FirstMile tier tool!")
