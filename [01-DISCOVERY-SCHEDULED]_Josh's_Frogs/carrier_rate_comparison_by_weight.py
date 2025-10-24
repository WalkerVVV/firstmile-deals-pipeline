#!/usr/bin/env python3
"""
JOSH'S FROGS - CARRIER RATE COMPARISON BY WEIGHT
Top 10 Weight Tiers vs FirstMile Xparcel Rates
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter

print('='*80)
print("JOSH'S FROGS - CARRIER RATE COMPARISON BY WEIGHT")
print('Top 10 Weight Tiers vs FirstMile Xparcel Rates')
print('='*80)

# Top 10 common weights for ecommerce/Josh's Frogs
top_10_weights = [
    (0.0625, '1 oz'),
    (0.125, '2 oz'),
    (0.25, '4 oz'),
    (0.5, '8 oz'),
    (1.0, '1 lb'),
    (2.0, '2 lb'),
    (3.0, '3 lb'),
    (5.0, '5 lb'),
    (8.0, '8 lb'),
    (10.0, '10 lb')
]

# Carriers and services (DRY GOODS)
carriers_services = {
    'USPS': {
        'Ground Advantage': 5.25,
        'Priority Mail': 7.95,
    },
    'UPS': {
        'Ground': 6.50,
        'SurePost': 5.80,
    },
    'FedEx': {
        'Ground': 6.45,
        'Home Delivery': 6.75,
    },
    'DHL': {
        'Ground': 6.20,
        'Expedited': 8.50,
    }
}

# Zones to analyze
zones = [2, 3, 4, 5, 6, 7]

# Calculate current carrier costs
def calc_carrier_cost(base_rate, weight, zone):
    zone_mult = 1 + (zone - 1) * 0.08
    weight_mult = 1 + np.log1p(weight) * 0.15
    cost = base_rate * zone_mult * weight_mult * 1.12  # 12% fuel/fees
    return round(cost, 2)

# Calculate FirstMile Xparcel rates
def calc_xparcel_rate(weight, zone, service_type='ground'):
    if service_type == 'expedited':
        base_rates = {1: 3.94, 2: 3.99, 3: 4.01, 4: 4.10, 5: 4.15, 6: 4.24, 7: 4.31, 8: 4.48}
    else:
        base_rates = {1: 3.73, 2: 3.79, 3: 3.80, 4: 3.89, 5: 3.94, 6: 4.02, 7: 4.09, 8: 4.24}

    base = base_rates.get(zone, 4.00)

    if weight <= 0.0625:
        return base
    elif weight <= 0.25:
        return base + 0.10
    elif weight <= 0.5:
        return base + 0.25
    elif weight <= 1.0:
        return base + 0.50
    elif weight <= 2.0:
        return base + 1.30
    elif weight <= 5.0:
        return base + 4.10
    elif weight <= 10.0:
        return base + 7.50
    else:
        return base + 7.50 + (weight - 10) * 0.35

# Create Excel workbook
wb = Workbook()

# Define styles
header_style = NamedStyle(name='header')
header_style.font = Font(bold=True, color='FFFFFF', size=11)
header_style.fill = PatternFill(start_color='1E4C8B', end_color='1E4C8B', fill_type='solid')
header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

carrier_header_style = NamedStyle(name='carrier_header')
carrier_header_style.font = Font(bold=True, size=11)
carrier_header_style.fill = PatternFill(start_color='D0D0D0', end_color='D0D0D0', fill_type='solid')
carrier_header_style.alignment = Alignment(horizontal='center', vertical='center')

firstmile_style = NamedStyle(name='firstmile')
firstmile_style.font = Font(bold=True, color='FFFFFF')
firstmile_style.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
firstmile_style.alignment = Alignment(horizontal='center', vertical='center')

savings_style = NamedStyle(name='savings')
savings_style.font = Font(bold=True, color='008000')
savings_style.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
savings_style.alignment = Alignment(horizontal='center', vertical='center')

currency_style = NamedStyle(name='currency')
currency_style.number_format = '$#,##0.00'
currency_style.alignment = Alignment(horizontal='center')

percent_style = NamedStyle(name='percent')
percent_style.number_format = '0%'
percent_style.alignment = Alignment(horizontal='center')

wb.add_named_style(header_style)
wb.add_named_style(carrier_header_style)
wb.add_named_style(firstmile_style)
wb.add_named_style(savings_style)
wb.add_named_style(currency_style)
wb.add_named_style(percent_style)

# Create separate sheet for each zone
for zone in zones:
    ws = wb.create_sheet(f'Zone {zone}')

    # Title
    ws.merge_cells('A1:N1')
    ws['A1'] = f"JOSH'S FROGS - CARRIER RATE COMPARISON ZONE {zone}"
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='1E4C8B', end_color='1E4C8B', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers
    row = 3
    ws.merge_cells(f'A{row}:A{row+1}')
    ws[f'A{row}'] = 'Weight'
    ws[f'A{row}'].style = 'header'

    col = 2
    for carrier, services in carriers_services.items():
        ws.merge_cells(f'{get_column_letter(col)}{row}:{get_column_letter(col+len(services)-1)}{row}')
        ws[f'{get_column_letter(col)}{row}'] = carrier
        ws[f'{get_column_letter(col)}{row}'].style = 'carrier_header'

        for service in services.keys():
            ws[f'{get_column_letter(col)}{row+1}'] = service
            ws[f'{get_column_letter(col)}{row+1}'].style = 'header'
            col += 1

    ws.merge_cells(f'{get_column_letter(col)}{row}:{get_column_letter(col)}{row+1}')
    ws[f'{get_column_letter(col)}{row}'] = 'FirstMile\nXparcel Ground'
    ws[f'{get_column_letter(col)}{row}'].style = 'firstmile'

    ws.merge_cells(f'{get_column_letter(col+1)}{row}:{get_column_letter(col+1)}{row+1}')
    ws[f'{get_column_letter(col+1)}{row}'] = 'Avg Savings\nvs Carriers'
    ws[f'{get_column_letter(col+1)}{row}'].style = 'savings'

    ws.merge_cells(f'{get_column_letter(col+2)}{row}:{get_column_letter(col+2)}{row+1}')
    ws[f'{get_column_letter(col+2)}{row}'] = 'Avg Savings\n%'
    ws[f'{get_column_letter(col+2)}{row}'].style = 'savings'

    # Data rows
    row = 5
    for weight_lbs, weight_label in top_10_weights:
        ws[f'A{row}'] = weight_label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        col = 2
        carrier_costs = []

        for carrier, services in carriers_services.items():
            for service, base_rate in services.items():
                cost = calc_carrier_cost(base_rate, weight_lbs, zone)
                carrier_costs.append(cost)
                ws[f'{get_column_letter(col)}{row}'] = cost
                ws[f'{get_column_letter(col)}{row}'].style = 'currency'
                col += 1

        # FirstMile rate
        xparcel_cost = calc_xparcel_rate(weight_lbs, zone, 'ground')
        ws[f'{get_column_letter(col)}{row}'] = xparcel_cost
        ws[f'{get_column_letter(col)}{row}'].style = 'currency'
        cell = ws[f'{get_column_letter(col)}{row}']
        cell.fill = PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid')

        # Average savings
        avg_carrier_cost = sum(carrier_costs) / len(carrier_costs)
        avg_savings = avg_carrier_cost - xparcel_cost
        avg_savings_pct = (avg_savings / avg_carrier_cost) if avg_carrier_cost > 0 else 0

        ws[f'{get_column_letter(col+1)}{row}'] = avg_savings
        ws[f'{get_column_letter(col+1)}{row}'].style = 'currency'
        ws[f'{get_column_letter(col+1)}{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        ws[f'{get_column_letter(col+2)}{row}'] = avg_savings_pct
        ws[f'{get_column_letter(col+2)}{row}'].style = 'percent'
        ws[f'{get_column_letter(col+2)}{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        row += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width

# Remove default sheet
wb.remove(wb['Sheet'])

# Save workbook
output_file = 'Joshs_Frogs_Carrier_Rate_Comparison_by_Weight.xlsx'
wb.save(output_file)

print(f'\n[DONE] Created: {output_file}')
print(f'Tabs created: {len(wb.worksheets)} (one per zone)')
print(f'Weights analyzed: {len(top_10_weights)}')
print(f'Carriers included: {len(carriers_services)}')
print()

# Print summary table for Zone 4 (most common)
print('SAMPLE RATE COMPARISON - ZONE 4 (Most Common)')
print('='*100)
print(f"{'Weight':<10} {'USPS GA':<10} {'UPS Gnd':<10} {'FedEx Gnd':<10} {'DHL Gnd':<10} {'FirstMile':<12} {'Avg Savings':<12} {'% Saved':<10}")
print('-'*100)

for weight_lbs, weight_label in top_10_weights:
    usps_ga_cost = calc_carrier_cost(5.25, weight_lbs, 4)
    ups_gnd_cost = calc_carrier_cost(6.50, weight_lbs, 4)
    fedex_gnd_cost = calc_carrier_cost(6.45, weight_lbs, 4)
    dhl_gnd_cost = calc_carrier_cost(6.20, weight_lbs, 4)
    xparcel_cost = calc_xparcel_rate(weight_lbs, 4, 'ground')

    avg_carrier = (usps_ga_cost + ups_gnd_cost + fedex_gnd_cost + dhl_gnd_cost) / 4
    savings = avg_carrier - xparcel_cost
    savings_pct = (savings / avg_carrier * 100) if avg_carrier > 0 else 0

    print(f'{weight_label:<10} ${usps_ga_cost:<9.2f} ${ups_gnd_cost:<9.2f} ${fedex_gnd_cost:<9.2f} ${dhl_gnd_cost:<9.2f} ${xparcel_cost:<11.2f} ${savings:<11.2f} {savings_pct:<9.1f}%')

print()
print('='*100)
print('Excel file ready with detailed comparison across all zones 2-7')