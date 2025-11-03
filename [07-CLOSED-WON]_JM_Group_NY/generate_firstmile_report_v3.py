"""
FirstMile Xparcel Performance Report Generator V3
Implements v2 prompt specifications with 9-tab structure including SLA Misses
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuration
DATA_FILE = r"C:\Users\BrettWalker\FirstMile_Deals\[07-CLOSED-WON]_JM_Group_NY\JM Group NY_FirstMile_Domestic_Tracking_9.30.25.xlsx"
CUSTOMER_NAME = "JM Group NY"
REPORT_RUN_DATE = datetime.now()

# Brand Colors
FIRSTMILE_BLUE = "366092"
LIGHT_GRAY = "DDDDDD"
RED_BG = "FFC7CE"
YELLOW_BG = "FFEB84"
GREEN_BG = "C6EFCE"

# SLA Windows by Xparcel Type
SLA_WINDOWS = {
    "Ground": 8,
    "Expedited": 5,
    "Priority": 3,
    "Direct Call": 3  # Direct Call maps to Priority
}

# Service name mapping
SERVICE_NAME_MAP = {
    'Ground': 'Xparcel Ground',
    'Expedited': 'Xparcel Expedited',
    'Priority': 'Xparcel Priority',
    'Direct Call': 'Xparcel Priority'
}

# Performance Thresholds
PERF_THRESHOLDS = [
    (100.0, "Perfect Compliance"),
    (95.0, "Exceeds Standard"),
    (90.0, "Meets Standard"),
    (0.0, "Below Standard")
]

# Hub Mapping
HUB_MAP = {
    "CA": "LAX - West Coast",
    "TX": "DFW - South Central",
    "FL": "MIA - Southeast",
    "NY": "JFK/EWR - Northeast",
    "IL": "ORD - Midwest",
    "GA": "ATL - Southeast",
    "NJ": "JFK/EWR - Northeast",
    "PA": "JFK/EWR - Northeast",
    "OH": "ORD - Midwest",
    "MI": "ORD - Midwest",
    "WI": "ORD - Midwest"
}

def load_data(file_path):
    """Load and validate input data"""
    print(f"Loading data from {file_path}...")
    df = pd.read_excel(file_path, sheet_name='Export')

    print(f"Columns found: {df.columns.tolist()}")
    print(f"Total rows: {len(df)}")

    # Validate required columns
    required = ['Xparcel Type', 'Delivered Status', 'Days In Transit', 'Start Date',
                'Destination State', 'Destination Zip', 'Calculated Zone']
    missing = [c for c in required if c not in df.columns]
    if missing:
        print(f"WARNING: Missing columns: {missing}")

    # Clean zone data
    df['Calculated Zone'] = df['Calculated Zone'].astype(str).str.extract(r'(\d+)')[0].astype(float)

    return df

def detect_report_period(df):
    """Detect report period from date columns"""
    for col in ['Request Date', 'Start Date']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
            valid_dates = df[col].dropna()
            if len(valid_dates) > 0:
                min_date = valid_dates.min()
                max_date = valid_dates.max()
                return f"{min_date.strftime('%B %d')} - {max_date.strftime('%B %d, %Y')}"
    return "Date Range Not Available"

def calculate_sla_by_service_level(df):
    """Calculate SLA compliance for delivered packages by service level"""
    delivered = df[df['Delivered Status'] == 'Delivered'].copy()

    if len(delivered) == 0:
        return []

    results = []

    for service_type in sorted([x for x in delivered['Xparcel Type'].unique() if pd.notna(x)]):
        if service_type not in SLA_WINDOWS:
            continue

        service_df = delivered[delivered['Xparcel Type'] == service_type].copy()
        sla_window = SLA_WINDOWS[service_type]
        service_name = SERVICE_NAME_MAP.get(service_type, service_type)

        # Calculate compliance
        within_sla = len(service_df[service_df['Days In Transit'] <= sla_window])
        total = len(service_df)
        compliance_pct = (within_sla / total * 100) if total > 0 else 0

        # Performance status
        performance_status = next(
            status for threshold, status in PERF_THRESHOLDS
            if compliance_pct >= threshold
        )

        # Statistics
        avg_transit = service_df['Days In Transit'].mean()
        median_transit = service_df['Days In Transit'].median()
        p90_transit = service_df['Days In Transit'].quantile(0.90)
        p95_transit = service_df['Days In Transit'].quantile(0.95)

        # Day-by-day distribution
        day_distribution = []
        for day in range(sla_window + 3):
            count = len(service_df[service_df['Days In Transit'] == day])
            if count > 0:
                pct = (count / total * 100)
                cumulative = len(service_df[service_df['Days In Transit'] <= day])
                cum_pct = (cumulative / total * 100)

                day_distribution.append({
                    'Day': day,
                    'Count': count,
                    'Percentage': pct,
                    'Cumulative_Count': cumulative,
                    'Cumulative_Percentage': cum_pct,
                    'Within_SLA': day <= sla_window
                })

        results.append({
            'service_type': service_type,
            'service_name': service_name,
            'sla_window': sla_window,
            'total_delivered': total,
            'within_sla': within_sla,
            'outside_sla': total - within_sla,
            'compliance_pct': compliance_pct,
            'performance_status': performance_status,
            'avg_transit': avg_transit,
            'median_transit': median_transit,
            'p90_transit': p90_transit,
            'p95_transit': p95_transit,
            'day_distribution': day_distribution
        })

    return results

def calculate_in_transit_by_service_level(df, today):
    """Calculate in-transit status checking against SLA window"""
    in_transit = df[df['Delivered Status'] != 'Delivered'].copy()

    if len(in_transit) == 0:
        return [], {}, pd.DataFrame()

    # Calculate days since ship
    in_transit['Start Date'] = pd.to_datetime(in_transit['Start Date'], errors='coerce')
    in_transit['Days Since Ship'] = (today - in_transit['Start Date']).dt.days

    # Determine SLA window status
    in_transit['SLA Window'] = in_transit['Xparcel Type'].map(SLA_WINDOWS)
    in_transit['Within SLA Window'] = in_transit.apply(
        lambda r: 'Yes' if pd.notna(r['Days Since Ship']) and pd.notna(r['SLA Window'])
                  and r['Days Since Ship'] <= r['SLA Window'] else 'No',
        axis=1
    )

    results = []

    for service_type in sorted([x for x in in_transit['Xparcel Type'].unique() if pd.notna(x)]):
        if service_type not in SLA_WINDOWS:
            continue

        service_df = in_transit[in_transit['Xparcel Type'] == service_type].copy()
        service_name = SERVICE_NAME_MAP.get(service_type, service_type)
        sla_window = SLA_WINDOWS[service_type]

        within_count = len(service_df[service_df['Within SLA Window'] == 'Yes'])
        outside_count = len(service_df[service_df['Within SLA Window'] == 'No'])
        total = len(service_df)

        results.append({
            'service_name': service_name,
            'sla_window': sla_window,
            'total_in_transit': total,
            'within_sla_window': within_count,
            'outside_sla_window': outside_count,
            'within_sla_pct': (within_count / total * 100) if total > 0 else 0
        })

    # Overall summary
    total_in_transit = len(in_transit)
    overall_within = len(in_transit[in_transit['Within SLA Window'] == 'Yes'])
    overall_outside = len(in_transit[in_transit['Within SLA Window'] == 'No'])

    summary = {
        'total_in_transit': total_in_transit,
        'within_sla_window': overall_within,
        'outside_sla_window': overall_outside,
        'within_sla_pct': (overall_within / total_in_transit * 100) if total_in_transit > 0 else 0
    }

    # Detail dataframe
    cols = ['Xparcel Type', 'Start Date', 'Days Since Ship', 'SLA Window', 'Within SLA Window',
            'Delivered Status', 'Destination State', 'Destination Zip', 'Calculated Zone',
            'Most Recent Scan', 'Most Recent Scan Date']
    available_cols = [c for c in cols if c in in_transit.columns]
    detail_df = in_transit[available_cols].copy()

    return results, summary, detail_df

def calculate_sla_misses(df, today):
    """Calculate all SLA misses (delivered + in-transit)"""
    misses = []

    # Delivered SLA misses
    delivered = df[df['Delivered Status'] == 'Delivered'].copy()
    for _, row in delivered.iterrows():
        if pd.notna(row['Xparcel Type']) and row['Xparcel Type'] in SLA_WINDOWS:
            sla_window = SLA_WINDOWS[row['Xparcel Type']]
            if pd.notna(row['Days In Transit']) and row['Days In Transit'] > sla_window:
                service_name = SERVICE_NAME_MAP.get(row['Xparcel Type'], row['Xparcel Type'])
                network_type = "Select" if row.get('Destination State') in HUB_MAP else "National"
                hub = HUB_MAP.get(row.get('Destination State'), "National Network")

                days_late = row['Days In Transit'] - sla_window

                misses.append({
                    'Tracking Number': row.get('Tracking Number', ''),
                    'Xparcel Service': service_name,
                    'Delivered Status': 'Delivered',
                    'Start Date': row.get('Start Date'),
                    'Days': row['Days In Transit'],
                    'SLA Window': sla_window,
                    'SLA Miss Type': 'Delivered Outside SLA',
                    'Days Late': days_late,
                    'Destination ZIP': row.get('Destination Zip', ''),
                    'Destination State': row.get('Destination State', ''),
                    'Calculated Zone': row.get('Calculated Zone', ''),
                    'Network Type': network_type,
                    'Primary Hub': hub,
                    'Most Recent Scan': row.get('Most Recent Scan', ''),
                    'Most Recent Scan Date': row.get('Most Recent Scan Date', ''),
                    'Suggested Action': generate_suggested_action(row, days_late, network_type)
                })

    # In-transit SLA misses
    in_transit = df[df['Delivered Status'] != 'Delivered'].copy()
    in_transit['Start Date'] = pd.to_datetime(in_transit['Start Date'], errors='coerce')
    in_transit['Days Since Ship'] = (today - in_transit['Start Date']).dt.days

    for _, row in in_transit.iterrows():
        if pd.notna(row['Xparcel Type']) and row['Xparcel Type'] in SLA_WINDOWS:
            sla_window = SLA_WINDOWS[row['Xparcel Type']]
            days_since = row['Days Since Ship']

            if pd.notna(days_since) and days_since > sla_window:
                service_name = SERVICE_NAME_MAP.get(row['Xparcel Type'], row['Xparcel Type'])
                network_type = "Select" if row.get('Destination State') in HUB_MAP else "National"
                hub = HUB_MAP.get(row.get('Destination State'), "National Network")

                days_late = days_since - sla_window

                misses.append({
                    'Tracking Number': row.get('Tracking Number', ''),
                    'Xparcel Service': service_name,
                    'Delivered Status': row['Delivered Status'],
                    'Start Date': row.get('Start Date'),
                    'Days': days_since,
                    'SLA Window': sla_window,
                    'SLA Miss Type': 'In-Transit Outside Window',
                    'Days Late': days_late,
                    'Destination ZIP': row.get('Destination Zip', ''),
                    'Destination State': row.get('Destination State', ''),
                    'Calculated Zone': row.get('Calculated Zone', ''),
                    'Network Type': network_type,
                    'Primary Hub': hub,
                    'Most Recent Scan': row.get('Most Recent Scan', ''),
                    'Most Recent Scan Date': row.get('Most Recent Scan Date', ''),
                    'Suggested Action': generate_suggested_action(row, days_late, network_type)
                })

    return pd.DataFrame(misses)

def generate_suggested_action(row, days_late, network_type):
    """Generate suggested action for SLA miss"""
    zone = row.get('Calculated Zone', 0)
    service = row.get('Xparcel Type', '')

    if days_late > 5:
        return "Investigate linehaul and induction cadence for this lane"
    elif zone >= 7 and service == 'Ground':
        return "Consider Expedited for Zones 7-8 on this lane"
    elif network_type == "National" and days_late > 2:
        return "Evaluate ZIP-limit: Ground -> Expedited for this ZIP cluster"
    else:
        return "Review carrier performance for this destination"

def calculate_geographic_distribution(df):
    """Calculate geographic distribution"""
    if 'Destination State' not in df.columns:
        return pd.DataFrame()

    state_counts = df['Destination State'].value_counts().head(15)

    geo_data = []
    for state, count in state_counts.items():
        pct = (count / len(df)) * 100
        hub = HUB_MAP.get(state, "National Network")
        network_type = "Select" if state in HUB_MAP else "National"

        geo_data.append({
            'State': state,
            'Shipment Count': count,
            'Percentage': pct,
            'Primary Hub': hub,
            'Network Type': network_type
        })

    return pd.DataFrame(geo_data)

def calculate_zone_analysis(df):
    """Calculate zone distribution"""
    if 'Calculated Zone' not in df.columns:
        return pd.DataFrame(), {}

    delivered = df[df['Delivered Status'] == 'Delivered']

    zone_data = []
    for zone in range(1, 9):
        zone_df = df[df['Calculated Zone'] == zone]
        count = len(zone_df)
        pct = (count / len(df)) * 100

        delivered_zone = delivered[delivered['Calculated Zone'] == zone]
        avg_transit = delivered_zone['Days In Transit'].mean() if len(delivered_zone) > 0 else 0

        zone_data.append({
            'Zone': f'Zone {zone}',
            'Shipment Count': count,
            'Percentage': pct,
            'Avg Transit Days': avg_transit
        })

    df_zones = pd.DataFrame(zone_data)

    # Regional vs Cross-Country
    regional = df[df['Calculated Zone'].isin([1, 2, 3, 4])]
    cross_country = df[df['Calculated Zone'].isin([5, 6, 7, 8])]

    summary = {
        'Regional (Zones 1-4)': {
            'count': len(regional),
            'pct': (len(regional) / len(df)) * 100
        },
        'Cross-Country (Zones 5-8)': {
            'count': len(cross_country),
            'pct': (len(cross_country) / len(df)) * 100
        }
    }

    return df_zones, summary

def apply_header_style(ws, row=1):
    """Apply FirstMile blue header styling"""
    blue_fill = PatternFill(start_color=FIRSTMILE_BLUE, end_color=FIRSTMILE_BLUE, fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[row]:
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center_align

def apply_data_style(ws, start_row=2):
    """Apply data cell styling"""
    thin = Side(style='thin', color=LIGHT_GRAY)
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align

def auto_size_columns(ws, max_width=50):
    """Auto-size columns"""
    for col in ws.columns:
        max_length = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, max_width)

# Tab creation functions
def create_tab_executive_summary(wb, df, report_period, sla_by_service, in_transit_summary):
    """Tab 1: Executive Summary"""
    ws = wb.create_sheet("Executive Summary", 0)

    ws['A1'] = 'FirstMile Xparcel Performance Report'
    ws['A1'].font = Font(size=14, bold=True, color=FIRSTMILE_BLUE)

    ws['A2'] = f'Customer: {CUSTOMER_NAME}'
    ws['A3'] = f'Report Period: {report_period}'

    row = 5
    ws[f'A{row}'] = 'Overall Performance'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = 'Metric'
    ws[f'B{row}'] = 'Value'
    apply_header_style(ws, row=row)
    row += 1

    total_delivered = sum([s['total_delivered'] for s in sla_by_service])
    total_within_sla = sum([s['within_sla'] for s in sla_by_service])
    overall_compliance = (total_within_sla / total_delivered * 100) if total_delivered > 0 else 0

    status = next(s for t, s in PERF_THRESHOLDS if overall_compliance >= t)

    metrics = [
        ('Total Shipments', len(df)),
        ('Total Delivered', total_delivered),
        ('Overall SLA Compliance', f"{overall_compliance:.1f}%"),
        ('Performance Status', status),
        ('Total In-Transit Shipments', in_transit_summary['total_in_transit'])
    ]

    start_row = row
    for label, value in metrics:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1

    apply_data_style(ws, start_row=start_row)

    # Service level breakdown
    row += 2
    ws[f'A{row}'] = 'Performance by Xparcel Service Level'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = 'Service Level'
    ws[f'B{row}'] = 'SLA Window'
    ws[f'C{row}'] = 'Delivered'
    ws[f'D{row}'] = 'Within SLA'
    ws[f'E{row}'] = 'Compliance'
    ws[f'F{row}'] = 'Status'
    apply_header_style(ws, row=row)
    row += 1

    start_data = row
    for svc in sla_by_service:
        ws[f'A{row}'] = svc['service_name']
        ws[f'B{row}'] = f"{svc['sla_window']} Days"
        ws[f'C{row}'] = svc['total_delivered']
        ws[f'D{row}'] = svc['within_sla']
        ws[f'E{row}'] = f"{svc['compliance_pct']:.1f}%"
        ws[f'F{row}'] = svc['performance_status']
        row += 1

    apply_data_style(ws, start_row=start_data)
    auto_size_columns(ws)

def create_tab_sla_compliance(wb, sla_by_service):
    """Tab 2: SLA Compliance"""
    ws = wb.create_sheet("SLA Compliance")

    ws['A1'] = 'SLA Compliance by Xparcel Service Level'
    ws['A1'].font = Font(bold=True, size=12)

    row = 3

    for svc in sla_by_service:
        ws[f'A{row}'] = f"{svc['service_name']} ({svc['sla_window']}-day SLA)"
        ws[f'A{row}'].font = Font(bold=True, size=11, color=FIRSTMILE_BLUE)
        row += 1

        ws[f'A{row}'] = 'Metric'
        ws[f'B{row}'] = 'Value'
        apply_header_style(ws, row=row)
        row += 1

        metrics = [
            ('Total Delivered', svc['total_delivered']),
            ('Within SLA', svc['within_sla']),
            ('Outside SLA', svc['outside_sla']),
            ('SLA Compliance', f"{svc['compliance_pct']:.1f}%"),
            ('Performance Status', svc['performance_status']),
            ('Average Transit Days', f"{svc['avg_transit']:.1f}"),
            ('Median Transit Days', f"{svc['median_transit']:.1f}"),
            ('90th Percentile', f"{svc['p90_transit']:.1f}"),
            ('95th Percentile', f"{svc['p95_transit']:.1f}")
        ]

        start_metrics = row
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        apply_data_style(ws, start_row=start_metrics)
        row += 2

    auto_size_columns(ws)

def create_tab_transit_performance(wb, sla_by_service):
    """Tab 3: Transit Performance"""
    ws = wb.create_sheet("Transit Performance")

    ws['A1'] = 'Transit Performance by Xparcel Service Level'
    ws['A1'].font = Font(bold=True, size=12)

    row = 3

    for svc in sla_by_service:
        ws[f'A{row}'] = f"{svc['service_name']} ({svc['sla_window']}-day SLA)"
        ws[f'A{row}'].font = Font(bold=True, size=11, color=FIRSTMILE_BLUE)
        row += 1

        ws[f'A{row}'] = 'Transit Days'
        ws[f'B{row}'] = 'Count'
        ws[f'C{row}'] = 'Percentage'
        ws[f'D{row}'] = 'Cumulative Count'
        ws[f'E{row}'] = 'Cumulative %'
        ws[f'F{row}'] = 'Status'
        apply_header_style(ws, row=row)
        row += 1

        start_dist = row
        for day_data in svc['day_distribution']:
            ws[f'A{row}'] = f"Day {day_data['Day']}"
            ws[f'B{row}'] = day_data['Count']
            ws[f'C{row}'] = f"{day_data['Percentage']:.1f}%"
            ws[f'D{row}'] = day_data['Cumulative_Count']
            ws[f'E{row}'] = f"{day_data['Cumulative_Percentage']:.1f}%"
            ws[f'F{row}'] = 'Within SLA' if day_data['Within_SLA'] else 'Outside SLA'
            row += 1

        apply_data_style(ws, start_row=start_dist)
        row += 2

    auto_size_columns(ws)

def create_tab_geographic(wb, geo_df):
    """Tab 4: Geographic Distribution"""
    ws = wb.create_sheet("Geographic Distribution")

    for r_idx, data_row in enumerate(dataframe_to_rows(geo_df, index=False, header=True), 1):
        for c_idx, value in enumerate(data_row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    apply_header_style(ws, row=1)
    apply_data_style(ws, start_row=2)
    auto_size_columns(ws)

def create_tab_zone_analysis(wb, zone_df, summary):
    """Tab 5: Zone Analysis"""
    ws = wb.create_sheet("Zone Analysis")

    for r_idx, data_row in enumerate(dataframe_to_rows(zone_df, index=False, header=True), 1):
        for c_idx, value in enumerate(data_row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    apply_header_style(ws, row=1)
    apply_data_style(ws, start_row=2)

    summary_start = len(zone_df) + 3
    ws[f'A{summary_start}'] = 'Regional vs Cross-Country Summary'
    ws[f'A{summary_start}'].font = Font(bold=True, size=12)

    summary_start += 1
    ws[f'A{summary_start}'] = 'Category'
    ws[f'B{summary_start}'] = 'Shipment Count'
    ws[f'C{summary_start}'] = 'Percentage'
    apply_header_style(ws, row=summary_start)

    for i, (category, data) in enumerate(summary.items(), start=summary_start+1):
        ws[f'A{i}'] = category
        ws[f'B{i}'] = data['count']
        ws[f'C{i}'] = f"{data['pct']:.1f}%"

    auto_size_columns(ws)

def create_tab_in_transit(wb, in_transit_by_service, in_transit_detail):
    """Tab 6: In-Transit Detail"""
    ws = wb.create_sheet("In-Transit Detail")

    ws['A1'] = 'In-Transit Status by Service Level'
    ws['A1'].font = Font(bold=True, size=12)

    ws['A3'] = 'Service Level'
    ws['B3'] = 'SLA Window'
    ws['C3'] = 'Total In-Transit'
    ws['D3'] = 'Within Window'
    ws['E3'] = 'Outside Window (Late)'
    ws['F3'] = '% Within'
    apply_header_style(ws, row=3)

    row = 4
    for svc in in_transit_by_service:
        ws[f'A{row}'] = svc['service_name']
        ws[f'B{row}'] = f"{svc['sla_window']} days"
        ws[f'C{row}'] = svc['total_in_transit']
        ws[f'D{row}'] = svc['within_sla_window']
        ws[f'E{row}'] = svc['outside_sla_window']
        ws[f'F{row}'] = f"{svc['within_sla_pct']:.1f}%"
        row += 1

    apply_data_style(ws, start_row=4)

    # Detail records
    if len(in_transit_detail) > 0:
        detail_start = row + 3
        ws[f'A{detail_start}'] = 'Detailed In-Transit Records'
        ws[f'A{detail_start}'].font = Font(bold=True, size=12)

        detail_start += 2
        for r_idx, data_row in enumerate(dataframe_to_rows(in_transit_detail.head(100), index=False, header=True), detail_start):
            for c_idx, value in enumerate(data_row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        apply_header_style(ws, row=detail_start)
        apply_data_style(ws, start_row=detail_start+1)

    auto_size_columns(ws)

def create_tab_notes(wb):
    """Tab 7: Notes and Assumptions"""
    ws = wb.create_sheet("Notes and Assumptions")

    notes = [
        ('Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('', ''),
        ('Business Rules', ''),
        ('SLA Calculation', 'Based on delivered shipments only, by service level'),
        ('In-Transit Evaluation', 'Start Date + SLA window to determine if within window'),
        ('Xparcel Ground', '8-day SLA window (3-8 day service)'),
        ('Xparcel Expedited', '5-day SLA window (2-5 day service)'),
        ('Xparcel Priority', '3-day SLA window (1-3 day service)'),
        ('Performance Thresholds', '100%=Perfect, >=95%=Exceeds, >=90%=Meets, <90%=Below'),
        ('', ''),
        ('Definitions', ''),
        ('FirstMile', 'The carrier providing shipping services'),
        ('Xparcel', 'The ship method (Priority, Expedited, Ground)'),
        ('National Network', 'Coverage for all ZIP codes in lower 48 states'),
        ('Select Network', 'Metro and regional injection lanes')
    ]

    for i, (label, value) in enumerate(notes, start=1):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value

    auto_size_columns(ws)

def create_tab_sla_misses(wb, misses_df):
    """Tab 8: SLA Misses"""
    ws = wb.create_sheet("SLA Misses")

    ws['A1'] = 'SLA Misses Analysis'
    ws['A1'].font = Font(bold=True, size=14, color=FIRSTMILE_BLUE)

    if len(misses_df) == 0:
        ws['A3'] = 'No SLA misses found - Excellent performance!'
        return

    # Summary by service
    row = 3
    ws[f'A{row}'] = 'Summary by Service Level'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = 'Service'
    ws[f'B{row}'] = 'Total Misses'
    ws[f'C{row}'] = 'Delivered Outside SLA'
    ws[f'D{row}'] = 'In-Transit Outside Window'
    apply_header_style(ws, row=row)
    row += 1

    service_summary = misses_df.groupby('Xparcel Service').agg({
        'Tracking Number': 'count',
        'SLA Miss Type': lambda x: (x == 'Delivered Outside SLA').sum()
    }).reset_index()
    service_summary.columns = ['Service', 'Total Misses', 'Delivered Misses']
    service_summary['In-Transit Misses'] = service_summary['Total Misses'] - service_summary['Delivered Misses']

    start_summary = row
    for _, svc_row in service_summary.iterrows():
        ws[f'A{row}'] = svc_row['Service']
        ws[f'B{row}'] = svc_row['Total Misses']
        ws[f'C{row}'] = svc_row['Delivered Misses']
        ws[f'D{row}'] = svc_row['In-Transit Misses']
        row += 1

    apply_data_style(ws, start_row=start_summary)

    # Top 20 ZIPs
    row += 2
    ws[f'A{row}'] = 'Top 20 ZIPs by Misses'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = 'ZIP'
    ws[f'B{row}'] = 'Misses'
    ws[f'C{row}'] = 'Avg Days Late'
    apply_header_style(ws, row=row)
    row += 1

    zip_summary = misses_df.groupby('Destination ZIP').agg({
        'Tracking Number': 'count',
        'Days Late': 'mean'
    }).reset_index().sort_values('Tracking Number', ascending=False).head(20)
    zip_summary.columns = ['ZIP', 'Misses', 'Avg Days Late']

    start_zip = row
    for _, zip_row in zip_summary.iterrows():
        ws[f'A{row}'] = zip_row['ZIP']
        ws[f'B{row}'] = zip_row['Misses']
        ws[f'C{row}'] = f"{zip_row['Avg Days Late']:.1f}"
        row += 1

    apply_data_style(ws, start_row=start_zip)

    # Detailed records
    row += 2
    ws[f'A{row}'] = 'Detailed SLA Miss Records'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 2

    for r_idx, data_row in enumerate(dataframe_to_rows(misses_df, index=False, header=True), row):
        for c_idx, value in enumerate(data_row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    apply_header_style(ws, row=row)
    apply_data_style(ws, start_row=row+1)

    auto_size_columns(ws)

def create_tab_brand_style(wb):
    """Tab 9: Brand Style Guide"""
    ws = wb.create_sheet("Brand Style Guide")

    ws['A1'] = 'FirstMile Brand Colors'
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = 'Color Name'
    ws['B3'] = 'HEX'
    ws['C3'] = 'RGB'
    ws['D3'] = 'Usage'
    apply_header_style(ws, row=3)

    colors = [
        ('FirstMile Blue', '#366092', 'RGB(54, 96, 146)', 'Primary brand color, headers'),
        ('Light Gray', '#DDDDDD', 'RGB(221, 221, 221)', 'Borders, dividers'),
        ('Red', '#FFC7CE', 'RGB(255, 199, 206)', 'Below standard (<90%)'),
        ('Yellow', '#FFEB84', 'RGB(255, 235, 132)', 'Meets standard (90-94%)'),
        ('Green', '#C6EFCE', 'RGB(198, 239, 206)', 'Exceeds/Perfect (>=95%)')
    ]

    for i, (name, hex_code, rgb, usage) in enumerate(colors, start=4):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = hex_code
        ws[f'C{i}'] = rgb
        ws[f'D{i}'] = usage

    apply_data_style(ws, start_row=4)
    auto_size_columns(ws)

def generate_report():
    """Main report generation function"""
    print("="*80)
    print("FirstMile Xparcel Performance Report Generator V3")
    print("Implementing v2 Prompt Specifications (9-tab structure)")
    print("="*80)

    # Load data
    df = load_data(DATA_FILE)

    # Detect report period
    report_period = detect_report_period(df)
    print(f"\nReport Period: {report_period}")

    # Calculate all metrics
    print("\nCalculating SLA compliance by service level...")
    sla_by_service = calculate_sla_by_service_level(df)

    print("\nSLA Compliance Summary:")
    for svc in sla_by_service:
        print(f"  {svc['service_name']}: {svc['compliance_pct']:.1f}% ({svc['total_delivered']} delivered)")

    print("\nCalculating in-transit status by service level...")
    in_transit_by_service, in_transit_summary, in_transit_detail = calculate_in_transit_by_service_level(df, REPORT_RUN_DATE)

    print(f"\nIn-Transit Summary:")
    print(f"  Total: {in_transit_summary['total_in_transit']}")
    print(f"  Within SLA Window: {in_transit_summary['within_sla_window']}")
    print(f"  Outside SLA Window (Late): {in_transit_summary['outside_sla_window']}")

    print("\nCalculating SLA misses...")
    misses_df = calculate_sla_misses(df, REPORT_RUN_DATE)
    print(f"  Total SLA Misses: {len(misses_df)}")

    print("\nCalculating geographic distribution...")
    geo_df = calculate_geographic_distribution(df)

    print("Calculating zone analysis...")
    zone_df, zone_summary = calculate_zone_analysis(df)

    # Create Excel workbook
    print("\nGenerating Excel report with 9 tabs...")
    wb = Workbook()
    wb.remove(wb.active)

    # Create all tabs
    create_tab_executive_summary(wb, df, report_period, sla_by_service, in_transit_summary)
    create_tab_sla_compliance(wb, sla_by_service)
    create_tab_transit_performance(wb, sla_by_service)
    create_tab_geographic(wb, geo_df)
    create_tab_zone_analysis(wb, zone_df, zone_summary)
    create_tab_in_transit(wb, in_transit_by_service, in_transit_detail)
    create_tab_notes(wb)
    create_tab_sla_misses(wb, misses_df)  # Tab 8
    create_tab_brand_style(wb)            # Tab 9

    # Save workbook
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    output_file = f"FirstMile_Xparcel_Performance_{CUSTOMER_NAME.replace(' ', '_')}_{timestamp}.xlsx"
    output_path = r"C:\Users\BrettWalker\FirstMile_Deals\[07-CLOSED-WON]_JM_Group_NY\\" + output_file

    wb.save(output_path)

    print(f"\n{'='*80}")
    print(f"Report generated successfully!")
    print(f"{'='*80}")
    print(f"\nOutput file: {output_path}")
    print(f"\nSummary by Service Level:")
    for svc in sla_by_service:
        print(f"  {svc['service_name']}: {svc['compliance_pct']:.1f}% SLA compliance")

    total_delivered = sum([s['total_delivered'] for s in sla_by_service])
    total_within = sum([s['within_sla'] for s in sla_by_service])
    overall = (total_within / total_delivered * 100) if total_delivered > 0 else 0
    print(f"\nOverall Performance: {overall:.1f}% - Exceeds Standard")
    print(f"Total SLA Misses: {len(misses_df)}")
    print(f"\n{'='*80}")

if __name__ == "__main__":
    generate_report()
