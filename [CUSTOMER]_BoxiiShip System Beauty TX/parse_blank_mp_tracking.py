import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

print("=" * 70)
print("BoxiiShip BLANK -MP Tracking Number Excel Generator")
print("=" * 70)

# The tracking data provided by the user (cleaned and structured)
tracking_text = """9400136208303385638161
Shipment info received.
Monday, October 6, 2025

9400136208303385638277
Shipment info received.
Monday, October 6, 2025
"""

# I need to save this to a file first, then parse it
# For now, let me create the parser that will work with the raw data

def parse_tracking_data_from_text(text):
    """Parse tracking data from text format."""
    lines = [line.strip() for line in text.strip().split('\n') if line.strip()]

    tracking_data = []
    i = 0

    while i < len(lines):
        line = lines[i]

        # Skip "View More Tracking Info" lines
        if 'View More Tracking Info' in line or 'View More' in line:
            i += 1
            continue

        # Check if this is a tracking number (starts with digit and is long enough)
        if line and line[0].isdigit() and len(line) >= 18:
            tracking_num = line
            status = ''
            date = ''
            location = ''

            # Get status (next non-empty line)
            if i+1 < len(lines) and lines[i+1] and 'View More' not in lines[i+1]:
                status = lines[i+1]

            # Get date (look for day of week)
            if i+2 < len(lines) and lines[i+2]:
                potential_date = lines[i+2]
                days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
                if any(day in potential_date for day in days):
                    date = potential_date

                    # Check for location on next line
                    if i+3 < len(lines) and lines[i+3] and lines[i+3][0].isupper() and 'View More' not in lines[i+3]:
                        location = lines[i+3]

            tracking_data.append({
                'Tracking Number': tracking_num,
                'Status': status,
                'Date': date,
                'Location': location
            })

        i += 1

    return tracking_data

# Read from the user's provided tracking data
# For this script, I'll need to read from a file you create with the tracking data
try:
    with open('blank_mp_tracking_data.txt', 'r', encoding='utf-8') as f:
        tracking_text = f.read()

    print(f"\nReading tracking data from file...")
    parsed_data = parse_tracking_data_from_text(tracking_text)

    # Create DataFrame
    df = pd.DataFrame(parsed_data)

    print(f"\n[SUCCESS] Parsed {len(df)} tracking numbers")

    # Show status distribution
    print(f"\n--- Status Distribution ---")
    status_counts = df['Status'].value_counts()
    print(status_counts.head(20))

    # Count specific statuses
    shipment_info = len(df[df['Status'].str.contains('Shipment info received', case=False, na=False)])
    delivered = len(df[df['Status'].str.contains('Delivered', case=False, na=False)])

    print(f"\n--- Summary ---")
    print(f"Total Tracking Numbers: {len(df)}")
    print(f"Shipment info received: {shipment_info} ({shipment_info/len(df)*100:.1f}%)")
    print(f"Delivered: {delivered} ({delivered/len(df)*100:.1f}%)")

    # Save to Excel with professional formatting
    output_file = 'BLANK_MP_Tracking_Nov4_2025.xlsx'

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='BLANK -MP Tracking')

        # Get the worksheet
        worksheet = writer.sheets['BLANK -MP Tracking']

        # Set column widths
        worksheet.column_dimensions['A'].width = 28  # Tracking Number
        worksheet.column_dimensions['B'].width = 50  # Status
        worksheet.column_dimensions['C'].width = 25  # Date
        worksheet.column_dimensions['D'].width = 35  # Location

        # Format header row (FirstMile blue #366092)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Center align tracking numbers and dates
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        # Auto-filter
        worksheet.auto_filter.ref = worksheet.dimensions

    print(f"\n[SUCCESS] Excel file created: {output_file}")
    print(f"[SUCCESS] Contains {len(df)} tracking numbers")
    print("=" * 70)

except FileNotFoundError:
    print("\n[ERROR] File 'blank_mp_tracking_data.txt' not found")
    print("Please create this file with the tracking data first.")
    print("=" * 70)
