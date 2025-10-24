"""
Update Implementation Plan tab in Stackd Logistics Excel report
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Open existing workbook
filename = "Stackd_Logistics_FirstMile_Xparcel_Savings_Analysis_20251014_1108.xlsx"
wb = openpyxl.load_workbook(filename)

# Get Implementation Plan sheet
ws = wb["Implementation Plan"]

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
title_font = Font(bold=True, size=14, color="366092")
bold_font = Font(bold=True)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD')
)

def style_header_row(ws, row_num, start_col=1, end_col=10):
    """Apply header styling to a row"""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

def auto_adjust_column_width(ws, max_width=50):
    """Auto-adjust column widths"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column_letter].width = adjusted_width

# Clear existing content (keep header)
for row in ws.iter_rows(min_row=5, max_row=30):
    for cell in row:
        cell.value = None
        cell.fill = PatternFill()
        cell.font = Font()
        cell.alignment = Alignment()
        cell.border = Border()

# Updated timeline
timeline = [
    ["Phase", "Activity", "Timeline", "Owner"],
    ["1. Rate Acceptance", "Review and accept FirstMile Xparcel rates", "Week 1", "Stackd Logistics"],
    ["2. Integration Setup", "ShipHero API already connected. FirstMile credentials created during account setup (7-10 days from FM Move Update Form and HubSpot New Customer Form submission)", "Weeks 1-2", "FirstMile + Stackd IT"],
    ["3. Performance Review", "TouchBase meetings to review SLA compliance and cost savings", "Weekly (Weeks 1-4), then Bi-Monthly", "Both teams"],
    ["4. Optimization", "Ongoing - Xparcel can be adjusted to improve SLA performance to underperforming destination shipping zones", "Ongoing", "Both teams"],
]

for i, row in enumerate(timeline, start=5):
    for j, val in enumerate(row, start=1):
        cell = ws.cell(row=i, column=j, value=val)
        cell.border = thin_border
        cell.alignment = left_align if j == 2 else center_align  # Left align Activity column

style_header_row(ws, 5, 1, 4)

# Update Success Metrics section position
ws['A11'] = "SUCCESS METRICS"
ws['A11'].font = Font(bold=True, size=12, color="366092")

metrics = [
    ["Metric", "Target", "Measurement"],
    ["Cost Savings", "9.5% reduction vs DHL", "Monthly invoice comparison"],
    ["SLA Compliance", ">90% within 3-8 day window", "Tracking data analysis"],
    ["Claims Rate", "<1% of total volume", "Claims ticket tracking"],
    ["Integration Uptime", ">99.5% API availability", "System monitoring"],
]

for i, row in enumerate(metrics, start=13):
    for j, val in enumerate(row, start=1):
        cell = ws.cell(row=i, column=j, value=val)
        cell.border = thin_border
        cell.alignment = center_align

style_header_row(ws, 13, 1, 3)

# Update Contact Information section position
ws['A19'] = "CONTACT INFORMATION"
ws['A19'].font = Font(bold=True, size=12, color="366092")

contacts = [
    ["Role", "Name", "Contact"],
    ["Sales Executive", "Brett Walker", "brett@firstmile.com"],
    ["Implementation Manager", "[To be assigned]", "ops@firstmile.com"],
    ["Technical Support", "FirstMile Support", "support@firstmile.com"],
]

for i, row in enumerate(contacts, start=21):
    for j, val in enumerate(row, start=1):
        cell = ws.cell(row=i, column=j, value=val)
        cell.border = thin_border
        cell.alignment = center_align

style_header_row(ws, 21, 1, 3)

# Auto-adjust column widths
auto_adjust_column_width(ws)

# Make Activity column wider for the longer text
ws.column_dimensions['B'].width = 60

# Save updated workbook
wb.save(filename)

print(f"Successfully updated Implementation Plan in {filename}")
print("\nUpdated Timeline:")
print("  1. Rate Acceptance: Week 1")
print("  2. Integration Setup: Weeks 1-2 (ShipHero already connected, FM credentials during setup)")
print("  3. Performance Review: Weekly (Weeks 1-4), then Bi-Monthly")
print("  4. Optimization: Ongoing (Xparcel adjustments for underperforming zones)")
print("\nRemoved phases:")
print("  - Pilot Testing")
print("  - Full Rollout")
