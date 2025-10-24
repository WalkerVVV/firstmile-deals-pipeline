"""
Build Josh's Frogs Current Rate Table from Actual Ship Data
Format matches Xparcel Expedited - Select template
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Load the detailed shipping data
print("Loading Josh's Frogs detailed shipping data...")
df = pd.read_csv('Joshs_Frogs_Detailed_Data.csv')

print(f"Total records: {len(df):,}")
print(f"Date range: {df['ship_date'].min()} to {df['ship_date'].max()}")

# Define weight tiers matching the template format
weight_tiers = [
    # Ounces (under 1 lb)
    1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15.99,
    # Pounds
    1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25
]

# Create weight tier labels
weight_labels = []
for tier in weight_tiers[:15]:  # Ounces
    if tier == int(tier):
        weight_labels.append(f"{int(tier)} oz")
    else:
        weight_labels.append(f"{tier} oz")

for tier in weight_tiers[15:]:  # Pounds
    weight_labels.append(f"{int(tier)} lb")

# Function to assign weight to tier
def assign_weight_tier(weight_lbs):
    """Assign actual weight to billing tier"""
    if pd.isna(weight_lbs):
        return None

    # Convert to ounces for under 1 lb
    if weight_lbs < 1:
        weight_oz = weight_lbs * 16
        # Round up to next oz tier
        for tier in weight_tiers[:15]:
            if weight_oz <= tier:
                return f"{tier} oz" if tier != int(tier) else f"{int(tier)} oz"
        return "15.99 oz"
    else:
        # Round up to next pound tier
        weight_ceiling = int(np.ceil(weight_lbs))
        if weight_ceiling > 25:
            weight_ceiling = 25
        return f"{weight_ceiling} lb"

# Assign weight tiers
df['weight_tier'] = df['weight'].apply(assign_weight_tier)

# Calculate average current cost by weight tier and zone
print("\nCalculating average rates by weight tier and zone...")
rate_matrix = df.groupby(['weight_tier', 'zone'])['current_cost'].agg(['mean', 'count']).reset_index()

# Pivot to create zone columns
rate_pivot = rate_matrix.pivot(index='weight_tier', columns='zone', values='mean')

# Ensure all zones 1-8 are present
for zone in range(1, 9):
    if zone not in rate_pivot.columns:
        rate_pivot[zone] = np.nan

# Sort columns by zone
rate_pivot = rate_pivot[[1, 2, 3, 4, 5, 6, 7, 8]]

# Reindex to match weight tier order
rate_pivot = rate_pivot.reindex(weight_labels)

# Fill missing values with interpolation or nearby values
rate_pivot = rate_pivot.interpolate(method='linear', axis=1, limit_direction='both')
rate_pivot = rate_pivot.interpolate(method='linear', axis=0, limit_direction='both')

# Display the rate table
print("\n" + "="*100)
print("JOSH'S FROGS - CURRENT RATE TABLE (from actual ship data)")
print("="*100)
print("\nWeight    Zone 1    Zone 2    Zone 3    Zone 4    Zone 5    Zone 6    Zone 7    Zone 8")
print("-" * 100)

for weight in weight_labels:
    if weight in rate_pivot.index:
        row_data = rate_pivot.loc[weight]
        rates = [f"${val:.2f}" if not pd.isna(val) else "N/A" for val in row_data]
        print(f"{weight:8s}  {rates[0]:>8s}  {rates[1]:>8s}  {rates[2]:>8s}  {rates[3]:>8s}  {rates[4]:>8s}  {rates[5]:>8s}  {rates[6]:>8s}  {rates[7]:>8s}")

# Create Excel workbook
print("\nCreating Excel rate table...")
wb = Workbook()
ws = wb.active
ws.title = "Josh's Frogs Current Rates"

# Header styling
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
header_alignment = Alignment(horizontal="center", vertical="center")

# Border style
thin_border = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD')
)

# Title
ws.merge_cells('A1:I1')
ws['A1'] = "Josh's Frogs - Current Rate Table (from actual shipping data)"
ws['A1'].font = Font(bold=True, size=14, color="366092")
ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 25

# Subtitle
ws.merge_cells('A2:I2')
ws['A2'] = f"Analysis Period: {df['ship_date'].min()} to {df['ship_date'].max()}"
ws['A2'].font = Font(size=10, italic=True)
ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

# Column headers
headers = ['Weight', 'Zone 1', 'Zone 2', 'Zone 3', 'Zone 4', 'Zone 5', 'Zone 6', 'Zone 7', 'Zone 8']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Data rows
row_num = 5
for weight in weight_labels:
    ws.cell(row=row_num, column=1, value=weight).alignment = Alignment(horizontal="left")
    ws.cell(row=row_num, column=1).border = thin_border

    if weight in rate_pivot.index:
        for col_num, zone in enumerate(range(1, 9), 2):
            value = rate_pivot.loc[weight, zone]
            if not pd.isna(value):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="center")
            else:
                cell = ws.cell(row=row_num, column=col_num, value="")
            cell.border = thin_border
    row_num += 1

# Auto-size columns
for col in range(1, 10):
    ws.column_dimensions[get_column_letter(col)].width = 12

# Save workbook
output_file = "Joshs_Frogs_Current_Rate_Table.xlsx"
wb.save(output_file)

print(f"\nExcel rate table saved: {output_file}")

# Summary statistics
print("\n" + "="*100)
print("RATE TABLE STATISTICS")
print("="*100)

# Volume distribution
print("\nVolume Distribution by Weight Tier (Top 15):")
weight_vol = df['weight_tier'].value_counts().head(15)
for weight, count in weight_vol.items():
    pct = (count / len(df)) * 100
    print(f"  {weight:10s}: {count:>6,} packages ({pct:>5.1f}%)")

# Zone distribution
print("\nVolume Distribution by Zone:")
zone_vol = df['zone'].value_counts().sort_index()
for zone, count in zone_vol.items():
    pct = (count / len(df)) * 100
    print(f"  Zone {int(zone)}: {count:>6,} packages ({pct:>5.1f}%)")

# Average costs
print("\nAverage Current Costs:")
print(f"  Overall average: ${df['current_cost'].mean():.2f}")
print(f"  Median cost: ${df['current_cost'].median():.2f}")
print(f"  Range: ${df['current_cost'].min():.2f} - ${df['current_cost'].max():.2f}")

print("\n" + "="*100)
print("ANALYSIS COMPLETE")
print("="*100)
