"""
Stackd Logistics - October 14 Excel Report Generator
Creates Excel file matching October 6 format with corrected October 14 data
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Load the comparison data
print("Loading comparison data...")
df = pd.read_csv('stackd_firstmile_vs_dhl_comparison.csv')

# Calculate key metrics
total_packages = len(df)
total_dhl_cost = df['Label Cost'].sum()
total_fm_cost = df['FirstMile_Cost'].sum()
total_savings = total_dhl_cost - total_fm_cost
savings_pct = (total_savings / total_dhl_cost * 100)

monthly_dhl = total_dhl_cost
monthly_fm = total_fm_cost
monthly_savings = total_savings
annual_savings = monthly_savings * 12

avg_dhl = df['Label Cost'].mean()
avg_fm = df['FirstMile_Cost'].mean()
avg_savings = avg_dhl - avg_fm

print(f"\nKey Metrics:")
print(f"  Packages: {total_packages:,}")
print(f"  Monthly Savings: ${monthly_savings:,.2f} ({savings_pct:.1f}%)")
print(f"  Annual Savings: ${annual_savings:,.2f}")

# Select Network states
select_states = ['CA', 'TX', 'FL', 'NY', 'WA', 'AZ', 'NJ', 'IL', 'MA', 'GA']
df['Network'] = df['State'].apply(lambda x: 'Select' if x in select_states else 'National')

select_df = df[df['Network'] == 'Select']
national_df = df[df['Network'] == 'National']

select_savings = select_df['Label Cost'].sum() - select_df['FirstMile_Cost'].sum()
select_pct = (select_savings / select_df['Label Cost'].sum() * 100)
national_savings = national_df['Label Cost'].sum() - national_df['FirstMile_Cost'].sum()
national_pct = (national_savings / national_df['Label Cost'].sum() * 100)

# Weight categories
def categorize_weight(weight):
    if weight < 0.25:
        return '<4 oz'
    elif weight < 0.5:
        return '4-8 oz'
    elif weight < 0.75:
        return '8-12 oz'
    elif weight < 1:
        return '12-16 oz'
    elif weight <= 2:
        return '1-2 lbs'
    else:
        return '>2 lbs'

df['Weight_Category'] = df['Weight (lb)'].apply(categorize_weight)

# Zone mapping (simplified)
zone_map = {
    'CA': 3, 'TX': 4, 'FL': 6, 'NY': 7, 'WA': 3,
    'AZ': 3, 'NJ': 7, 'IL': 5, 'MA': 7, 'GA': 6,
    'NC': 6, 'CO': 4, 'PA': 6, 'OH': 5, 'MI': 5,
    'VA': 6, 'MD': 7, 'OR': 3, 'WI': 5, 'MO': 5
}
df['Zone'] = df['State'].map(zone_map).fillna(5).astype(int)

# Create Excel workbook
print("\nCreating Excel workbook...")
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
title_font = Font(bold=True, size=14, color="366092")
bold_font = Font(bold=True)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center')
thin_border = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD')
)

def style_header_row(ws, row_num, start_col=1, end_col=10):
    """Apply header styling to a row"""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

def auto_adjust_column_width(ws, max_width=50):
    """Auto-adjust column widths"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column_letter].width = adjusted_width

# Sheet 1: Executive Summary
print("Creating Executive Summary sheet...")
ws1 = wb.create_sheet("Executive Summary")

ws1['A1'] = "Stackd Logistics - FirstMile Xparcel Savings Analysis"
ws1['A1'].font = title_font
ws1['A2'] = f"Analysis Date: October 14, 2025 at 10:19 AM MST"
ws1['A2'].font = Font(italic=True, size=10)

ws1['A4'] = "EXECUTIVE SUMMARY"
ws1['A4'].font = Font(bold=True, size=12, color="366092")

summary_data = [
    ["Metric", "Value"],
    ["Total Monthly Volume", f"{total_packages:,} packages"],
    ["Current Monthly Cost (DHL)", f"${monthly_dhl:,.2f}"],
    ["FirstMile Monthly Cost", f"${monthly_fm:,.2f}"],
    ["Monthly Savings", f"${monthly_savings:,.2f}"],
    ["Savings Percentage", f"{savings_pct:.1f}%"],
    ["ANNUAL SAVINGS", f"${annual_savings:,.2f}"],
    ["", ""],
    ["Average Cost per Package", ""],
    ["  DHL eCommerce Current", f"${avg_dhl:.2f}"],
    ["  FirstMile Xparcel", f"${avg_fm:.2f}"],
    ["  Savings per Package", f"${avg_savings:.2f}"],
]

for i, row in enumerate(summary_data, start=6):
    ws1[f'A{i}'] = row[0]
    ws1[f'B{i}'] = row[1]
    if "ANNUAL" in row[0]:
        ws1[f'A{i}'].font = bold_font
        ws1[f'B{i}'].font = bold_font

style_header_row(ws1, 6, 1, 2)

ws1['A19'] = "NETWORK BREAKDOWN"
ws1['A19'].font = Font(bold=True, size=12, color="366092")

network_data = [
    ["Network", "Volume", "% of Total", "DHL Cost", "FirstMile Cost", "Savings", "% Savings"],
    ["Select Network", f"{len(select_df):,}", f"{len(select_df)/total_packages*100:.1f}%",
     f"${select_df['Label Cost'].sum():,.2f}", f"${select_df['FirstMile_Cost'].sum():,.2f}",
     f"${select_savings:,.2f}", f"{select_pct:.1f}%"],
    ["National Network", f"{len(national_df):,}", f"{len(national_df)/total_packages*100:.1f}%",
     f"${national_df['Label Cost'].sum():,.2f}", f"${national_df['FirstMile_Cost'].sum():,.2f}",
     f"${national_savings:,.2f}", f"{national_pct:.1f}%"],
]

for i, row in enumerate(network_data, start=21):
    for j, val in enumerate(row, start=1):
        ws1.cell(row=i, column=j, value=val)

style_header_row(ws1, 21, 1, 7)

ws1['A25'] = "KEY INSIGHTS"
ws1['A25'].font = Font(bold=True, size=12, color="366092")
insights = [
    "• FirstMile achieves 9.5% overall savings vs DHL eCommerce baseline",
    "• Select Network metros (CA, TX, FL, NY, WA) deliver 15.7% savings on 53% of volume",
    "• National Network delivers 2.7% savings on remaining 47% of volume",
    "• 92.5% of packages are under 1 lb (FirstMile's sweet spot)",
    "• DHL eCommerce ground is the addressable volume (non-guaranteed service)",
    "• Excluded UPS 2nd Day Air (guaranteed service, out of scope for Xparcel)",
]
for i, insight in enumerate(insights, start=26):
    ws1[f'A{i}'] = insight

auto_adjust_column_width(ws1)

# Sheet 2: FirstMile Advantages
print("Creating FirstMile Advantages sheet...")
ws2 = wb.create_sheet("FirstMile Advantages")

ws2['A1'] = "FirstMile Xparcel Advantages"
ws2['A1'].font = title_font

ws2['A3'] = "OPERATIONAL BENEFITS"
ws2['A3'].font = Font(bold=True, size=12, color="366092")

advantages = [
    ["Advantage", "Description", "Business Impact"],
    ["Dynamic Routing", "Automatically selects best carrier (DHL, ACI, USPS) based on destination",
     "Optimal delivery performance without manual carrier selection"],
    ["Audit Queue", "Catches mis-rated labels before invoicing",
     "Prevents overcharges and billing disputes"],
    ["Unified Platform", "Single integration for all carriers and networks",
     "Reduced IT complexity, one API instead of multiple"],
    ["Claims & Returns", "Single support contact for all carriers",
     "Faster resolution, no carrier finger-pointing"],
    ["Select Network", "Major metro injection points for zone-skipping",
     "Faster transit and lower costs to major markets"],
    ["National Network", "100% ZIP coverage through nationwide partners",
     "Reach every destination with competitive pricing"],
]

for i, row in enumerate(advantages, start=5):
    for j, val in enumerate(row, start=1):
        ws2.cell(row=i, column=j, value=val)

style_header_row(ws2, 5, 1, 3)

ws2['A12'] = "SERVICE COMPARISON"
ws2['A12'].font = Font(bold=True, size=12, color="366092")

service_comp = [
    ["Service", "Transit Time", "Type", "Use Case"],
    ["Xparcel Ground", "3-8 days", "Non-Guaranteed", "Economy shipping, cost-sensitive customers"],
    ["Xparcel Expedited", "2-5 days", "Non-Guaranteed", "Faster ground, 1-20 lb packages"],
    ["Xparcel Priority", "1-3 days", "Guaranteed", "Premium service with money-back guarantee"],
]

for i, row in enumerate(service_comp, start=14):
    for j, val in enumerate(row, start=1):
        ws2.cell(row=i, column=j, value=val)

style_header_row(ws2, 14, 1, 4)

auto_adjust_column_width(ws2)

# Sheet 3: Weight Analysis
print("Creating Weight Analysis sheet...")
ws3 = wb.create_sheet("Weight Analysis")

ws3['A1'] = "Weight Distribution Analysis"
ws3['A1'].font = title_font

weight_analysis = df.groupby('Weight_Category').agg({
    'Shipping Label ID': 'count',
    'Label Cost': ['sum', 'mean'],
    'FirstMile_Cost': ['sum', 'mean'],
    'Savings': 'sum'
}).round(2)

weight_analysis.columns = ['Volume', 'DHL Total', 'DHL Avg', 'FM Total', 'FM Avg', 'Total Savings']
weight_analysis['Savings/Pkg'] = (weight_analysis['DHL Avg'] - weight_analysis['FM Avg']).round(2)
weight_analysis['% Savings'] = ((weight_analysis['Total Savings'] / weight_analysis['DHL Total']) * 100).round(1)
weight_analysis['% of Volume'] = (weight_analysis['Volume'] / total_packages * 100).round(1)

# Reorder for logical presentation
weight_order = ['<4 oz', '4-8 oz', '8-12 oz', '12-16 oz', '1-2 lbs', '>2 lbs']
weight_analysis = weight_analysis.reindex([w for w in weight_order if w in weight_analysis.index])

ws3['A3'] = "WEIGHT TIER PERFORMANCE"
ws3['A3'].font = Font(bold=True, size=12, color="366092")

# Write headers
headers = ['Weight Tier', 'Volume', '% of Volume', 'DHL Total', 'DHL Avg', 'FM Total', 'FM Avg', 'Total Savings', 'Savings/Pkg', '% Savings']
for j, header in enumerate(headers, start=1):
    ws3.cell(row=5, column=j, value=header)

style_header_row(ws3, 5, 1, len(headers))

# Write data
for i, (idx, row) in enumerate(weight_analysis.iterrows(), start=6):
    ws3.cell(row=i, column=1, value=idx)
    ws3.cell(row=i, column=2, value=int(row['Volume']))
    ws3.cell(row=i, column=3, value=f"{row['% of Volume']:.1f}%")
    ws3.cell(row=i, column=4, value=f"${row['DHL Total']:,.2f}")
    ws3.cell(row=i, column=5, value=f"${row['DHL Avg']:.2f}")
    ws3.cell(row=i, column=6, value=f"${row['FM Total']:,.2f}")
    ws3.cell(row=i, column=7, value=f"${row['FM Avg']:.2f}")
    ws3.cell(row=i, column=8, value=f"${row['Total Savings']:,.2f}")
    ws3.cell(row=i, column=9, value=f"${row['Savings/Pkg']:.2f}")
    ws3.cell(row=i, column=10, value=f"{row['% Savings']:.1f}%")

ws3['A13'] = "KEY FINDINGS"
ws3['A13'].font = Font(bold=True, size=12, color="366092")

findings = [
    f"• {weight_analysis.loc['<4 oz', 'Volume']:,.0f} packages (<4 oz) = {weight_analysis.loc['<4 oz', '% of Volume']:.1f}% of volume",
    f"• Lightweight packages (<1 lb) represent {(df['Weight (lb)'] < 1).sum()/total_packages*100:.1f}% of total volume",
    f"• Best savings in <4 oz tier: {weight_analysis.loc['<4 oz', '% Savings']:.1f}% savings",
    f"• Heavier packages (>2 lbs) show {weight_analysis.loc['>2 lbs', '% Savings']:.1f}% savings on small volume",
]

for i, finding in enumerate(findings, start=14):
    ws3[f'A{i}'] = finding

auto_adjust_column_width(ws3)

# Sheet 4: Zone Analysis
print("Creating Zone Analysis sheet...")
ws4 = wb.create_sheet("Zone Analysis")

ws4['A1'] = "Zone Distribution Analysis"
ws4['A1'].font = title_font

zone_analysis = df.groupby('Zone').agg({
    'Shipping Label ID': 'count',
    'Label Cost': ['sum', 'mean'],
    'FirstMile_Cost': ['sum', 'mean'],
    'Savings': 'sum'
}).round(2)

zone_analysis.columns = ['Volume', 'DHL Total', 'DHL Avg', 'FM Total', 'FM Avg', 'Total Savings']
zone_analysis['% of Volume'] = (zone_analysis['Volume'] / total_packages * 100).round(1)
zone_analysis['% Savings'] = ((zone_analysis['Total Savings'] / zone_analysis['DHL Total']) * 100).round(1)

ws4['A3'] = "ZONE PERFORMANCE"
ws4['A3'].font = Font(bold=True, size=12, color="366092")

# Write headers
headers = ['Zone', 'Volume', '% of Volume', 'DHL Total', 'DHL Avg', 'FM Total', 'FM Avg', 'Total Savings', '% Savings']
for j, header in enumerate(headers, start=1):
    ws4.cell(row=5, column=j, value=header)

style_header_row(ws4, 5, 1, len(headers))

# Write data
for i, (zone, row) in enumerate(zone_analysis.iterrows(), start=6):
    ws4.cell(row=i, column=1, value=f"Zone {int(zone)}")
    ws4.cell(row=i, column=2, value=int(row['Volume']))
    ws4.cell(row=i, column=3, value=f"{row['% of Volume']:.1f}%")
    ws4.cell(row=i, column=4, value=f"${row['DHL Total']:,.2f}")
    ws4.cell(row=i, column=5, value=f"${row['DHL Avg']:.2f}")
    ws4.cell(row=i, column=6, value=f"${row['FM Total']:,.2f}")
    ws4.cell(row=i, column=7, value=f"${row['FM Avg']:.2f}")
    ws4.cell(row=i, column=8, value=f"${row['Total Savings']:,.2f}")
    ws4.cell(row=i, column=9, value=f"{row['% Savings']:.1f}%")

# Regional breakdown
ws4['A15'] = "REGIONAL BREAKDOWN"
ws4['A15'].font = Font(bold=True, size=12, color="366092")

regional_zones = df[df['Zone'] <= 4]
crosscountry_zones = df[df['Zone'] >= 5]

regional_data = [
    ["Region", "Zones", "Volume", "% of Total", "DHL Cost", "FirstMile Cost", "Savings", "% Savings"],
    ["Regional", "1-4", f"{len(regional_zones):,}", f"{len(regional_zones)/total_packages*100:.1f}%",
     f"${regional_zones['Label Cost'].sum():,.2f}", f"${regional_zones['FirstMile_Cost'].sum():,.2f}",
     f"${(regional_zones['Label Cost'].sum() - regional_zones['FirstMile_Cost'].sum()):,.2f}",
     f"{((regional_zones['Label Cost'].sum() - regional_zones['FirstMile_Cost'].sum())/regional_zones['Label Cost'].sum()*100):.1f}%"],
    ["Cross-Country", "5-8", f"{len(crosscountry_zones):,}", f"{len(crosscountry_zones)/total_packages*100:.1f}%",
     f"${crosscountry_zones['Label Cost'].sum():,.2f}", f"${crosscountry_zones['FirstMile_Cost'].sum():,.2f}",
     f"${(crosscountry_zones['Label Cost'].sum() - crosscountry_zones['FirstMile_Cost'].sum()):,.2f}",
     f"{((crosscountry_zones['Label Cost'].sum() - crosscountry_zones['FirstMile_Cost'].sum())/crosscountry_zones['Label Cost'].sum()*100):.1f}%"],
]

for i, row in enumerate(regional_data, start=17):
    for j, val in enumerate(row, start=1):
        ws4.cell(row=i, column=j, value=val)

style_header_row(ws4, 17, 1, 8)

auto_adjust_column_width(ws4)

# Sheet 5: Rate Comparison
print("Creating Rate Comparison sheet...")
ws5 = wb.create_sheet("Rate Comparison")

ws5['A1'] = "DHL vs FirstMile Rate Comparison"
ws5['A1'].font = title_font

ws5['A3'] = "TOP 10 STATES BY VOLUME"
ws5['A3'].font = Font(bold=True, size=12, color="366092")

state_analysis = df.groupby('State').agg({
    'Shipping Label ID': 'count',
    'Label Cost': ['sum', 'mean'],
    'FirstMile_Cost': ['sum', 'mean'],
    'Savings': 'sum'
}).round(2)

state_analysis.columns = ['Volume', 'DHL Total', 'DHL Avg', 'FM Total', 'FM Avg', 'Total Savings']
state_analysis['Savings/Pkg'] = (state_analysis['DHL Avg'] - state_analysis['FM Avg']).round(2)
state_analysis['Network'] = state_analysis.index.map(lambda x: 'Select' if x in select_states else 'National')
state_analysis = state_analysis.sort_values('Volume', ascending=False).head(10)

headers = ['State', 'Network', 'Volume', 'DHL Total', 'DHL Avg', 'FM Total', 'FM Avg', 'Total Savings', 'Savings/Pkg']
for j, header in enumerate(headers, start=1):
    ws5.cell(row=5, column=j, value=header)

style_header_row(ws5, 5, 1, len(headers))

for i, (state, row) in enumerate(state_analysis.iterrows(), start=6):
    ws5.cell(row=i, column=1, value=state)
    ws5.cell(row=i, column=2, value=row['Network'])
    ws5.cell(row=i, column=3, value=int(row['Volume']))
    ws5.cell(row=i, column=4, value=f"${row['DHL Total']:,.2f}")
    ws5.cell(row=i, column=5, value=f"${row['DHL Avg']:.2f}")
    ws5.cell(row=i, column=6, value=f"${row['FM Total']:,.2f}")
    ws5.cell(row=i, column=7, value=f"${row['FM Avg']:.2f}")
    ws5.cell(row=i, column=8, value=f"${row['Total Savings']:,.2f}")
    ws5.cell(row=i, column=9, value=f"${row['Savings/Pkg']:.2f}")

ws5['A17'] = "COMPETITIVE POSITIONING"
ws5['A17'].font = Font(bold=True, size=12, color="366092")

positioning = [
    ["Aspect", "DHL eCommerce (Current)", "FirstMile Xparcel (Proposed)"],
    ["Average Cost per Package", f"${avg_dhl:.2f}", f"${avg_fm:.2f}"],
    ["Monthly Volume", f"{total_packages:,} packages", f"{total_packages:,} packages"],
    ["Monthly Cost", f"${monthly_dhl:,.2f}", f"${monthly_fm:,.2f}"],
    ["Service Type", "Non-Guaranteed Ground", "Non-Guaranteed Ground (3-8 days)"],
    ["Network Coverage", "Nationwide", "Select + National (100% coverage)"],
    ["Dynamic Routing", "Single carrier", "Multi-carrier optimization"],
    ["Support Model", "Direct carrier support", "Unified FirstMile support"],
]

for i, row in enumerate(positioning, start=19):
    for j, val in enumerate(row, start=1):
        ws5.cell(row=i, column=j, value=val)

style_header_row(ws5, 19, 1, 3)

auto_adjust_column_width(ws5)

# Sheet 6: Implementation Plan
print("Creating Implementation Plan sheet...")
ws6 = wb.create_sheet("Implementation Plan")

ws6['A1'] = "Implementation Plan & Next Steps"
ws6['A1'].font = title_font

ws6['A3'] = "PROPOSED TIMELINE"
ws6['A3'].font = Font(bold=True, size=12, color="366092")

timeline = [
    ["Phase", "Activity", "Timeline", "Owner"],
    ["1. Rate Acceptance", "Review and accept FirstMile Xparcel rates", "Week 1", "Stackd Logistics"],
    ["2. Integration Setup", "ShipHero → FirstMile API integration", "Weeks 2-3", "FirstMile + Stackd IT"],
    ["3. Pilot Testing", "Test 500-1,000 packages through FirstMile", "Week 4", "Both teams"],
    ["4. Performance Review", "Review SLA compliance and cost savings", "Week 5", "Both teams"],
    ["5. Full Rollout", "Transition all DHL ground volume to FirstMile", "Week 6", "Stackd Logistics"],
    ["6. Optimization", "Fine-tune routing and performance", "Ongoing", "Both teams"],
]

for i, row in enumerate(timeline, start=5):
    for j, val in enumerate(row, start=1):
        ws6.cell(row=i, column=j, value=val)

style_header_row(ws6, 5, 1, 4)

ws6['A13'] = "SUCCESS METRICS"
ws6['A13'].font = Font(bold=True, size=12, color="366092")

metrics = [
    ["Metric", "Target", "Measurement"],
    ["Cost Savings", "9.5% reduction vs DHL", "Monthly invoice comparison"],
    ["SLA Compliance", ">90% within 3-8 day window", "Tracking data analysis"],
    ["Claims Rate", "<1% of total volume", "Claims ticket tracking"],
    ["Integration Uptime", ">99.5% API availability", "System monitoring"],
]

for i, row in enumerate(metrics, start=15):
    for j, val in enumerate(row, start=1):
        ws6.cell(row=i, column=j, value=val)

style_header_row(ws6, 15, 1, 3)

ws6['A20'] = "CONTACT INFORMATION"
ws6['A20'].font = Font(bold=True, size=12, color="366092")

contacts = [
    ["Role", "Name", "Contact"],
    ["Sales Executive", "Brett Walker", "brett@firstmile.com"],
    ["Implementation Manager", "[To be assigned]", "ops@firstmile.com"],
    ["Technical Support", "FirstMile Support", "support@firstmile.com"],
]

for i, row in enumerate(contacts, start=22):
    for j, val in enumerate(row, start=1):
        ws6.cell(row=i, column=j, value=val)

style_header_row(ws6, 22, 1, 3)

auto_adjust_column_width(ws6)

# Save workbook
timestamp = datetime.now().strftime("%Y%m%d_%H%M")
filename = f"Stackd_Logistics_FirstMile_Xparcel_Savings_Analysis_{timestamp}.xlsx"
wb.save(filename)

print(f"\nExcel report created successfully: {filename}")
print("\nSummary:")
print(f"  6 sheets created (Executive Summary, FirstMile Advantages, Weight Analysis, Zone Analysis, Rate Comparison, Implementation Plan)")
print(f"  Total packages analyzed: {total_packages:,}")
print(f"  Annual savings: ${annual_savings:,.2f} ({savings_pct:.1f}%)")
print(f"  Select Network: {len(select_df):,} packages ({len(select_df)/total_packages*100:.1f}%) at {select_pct:.1f}% savings")
print(f"  National Network: {len(national_df):,} packages ({len(national_df)/total_packages*100:.1f}%) at {national_pct:.1f}% savings")
