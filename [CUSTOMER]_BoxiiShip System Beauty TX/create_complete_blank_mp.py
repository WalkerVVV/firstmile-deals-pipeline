import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

print("=" * 70)
print("BoxiiShip BLANK -MP COMPLETE Tracking Parser (ALL 1,436)")
print("=" * 70)

# Read the user's complete tracking list
print("\nReading complete tracking list...")
with open('blank_mp_tracking_list.txt', 'r') as f:
    all_tracking_numbers = [line.strip() for line in f.readlines() if line.strip()]

print(f"Total tracking numbers in list: {len(all_tracking_numbers)}")

# Parse the tracking data from the .txt file
print("\nParsing tracking data from file...")
with open('BLANK -MP BoxiiShip SB Logistics LLC TX Oct 6.txt', 'r', encoding='utf-8') as f:
    lines = [line.strip() for line in f.readlines()]

# Create a dictionary to store tracking data
tracking_dict = {}
i = 0

while i < len(lines):
    line = lines[i]

    # Skip empty lines and header
    if not line or 'BoxiiShip' in line or 'Tracking IDs' in line:
        i += 1
        continue

    # Check if this is a tracking number
    if line and line[0].isdigit() and len(line) >= 18:
        tracking_num = line
        status = ''
        date = ''
        location = ''

        # Get status (next non-empty line)
        j = i + 1
        while j < len(lines) and not lines[j]:
            j += 1

        if j < len(lines) and 'View More' not in lines[j]:
            status = lines[j]
            j += 1

        # Get date (look for day of week)
        while j < len(lines) and not lines[j]:
            j += 1

        if j < len(lines):
            potential_date = lines[j]
            days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            if any(day in potential_date for day in days):
                date = potential_date
                j += 1

                # Check for location
                while j < len(lines) and not lines[j]:
                    j += 1

                if j < len(lines) and 'View More' not in lines[j]:
                    location = lines[j]

        tracking_dict[tracking_num] = {
            'Status': status,
            'Date': date,
            'Location': location
        }

    i += 1

print(f"Parsed {len(tracking_dict)} tracking numbers with data")

# Create complete list with ALL tracking numbers
tracking_data = []
for tracking_num in all_tracking_numbers:
    if tracking_num in tracking_dict:
        tracking_data.append({
            'Tracking Number': tracking_num,
            'Status': tracking_dict[tracking_num]['Status'],
            'Date': tracking_dict[tracking_num]['Date'],
            'Location': tracking_dict[tracking_num]['Location']
        })
    else:
        # Tracking number not found in .txt file - include with blank data
        tracking_data.append({
            'Tracking Number': tracking_num,
            'Status': '',
            'Date': '',
            'Location': ''
        })

# Create DataFrame
df = pd.DataFrame(tracking_data)

print(f"\n[SUCCESS] Created complete dataset with {len(df)} tracking numbers")

# Show summary
with_data = len(df[df['Status'] != ''])
without_data = len(df[df['Status'] == ''])

print(f"\nWith tracking data: {with_data}")
print(f"Without tracking data: {without_data}")

# Save to Excel with professional formatting
output_file = 'BLANK_MP_Tracking_Nov4_2025.xlsx'

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='BLANK -MP Tracking')

    # Get the worksheet
    worksheet = writer.sheets['BLANK -MP Tracking']

    # Set column widths
    worksheet.column_dimensions['A'].width = 28  # Tracking Number
    worksheet.column_dimensions['B'].width = 50  # Status
    worksheet.column_dimensions['C'].width = 30  # Date
    worksheet.column_dimensions['D'].width = 35  # Location

    # Format header row (FirstMile blue #366092)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Center align tracking numbers and dates
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    # Auto-filter
    worksheet.auto_filter.ref = worksheet.dimensions

print(f"\n[SUCCESS] Excel file created: {output_file}")
print(f"[SUCCESS] Contains ALL {len(df)} BLANK -MP tracking numbers")
print("=" * 70)
