import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import DateAxis

# Load the data
df = pd.read_csv('247bef97-8663-431e-b2f5-dd2ca243633d.csv')

# Prepare data
df['date'] = pd.to_datetime(df['Number'].str[:10], format='%Y-%m-%d', errors='coerce')
df['year_month'] = df['date'].dt.to_period('M')
df['weight_oz'] = df['Weight'] * 16

# Calculate billable weight
def calculate_billable_weight(weight_oz):
    if weight_oz <= 16:
        return min(np.ceil(weight_oz), 15.99)
    else:
        return np.ceil(weight_oz / 16) * 16

df['billable_oz'] = df['weight_oz'].apply(calculate_billable_weight)
df['billable_lbs'] = df['billable_oz'] / 16

# Create weight categories
def categorize_weight(weight_oz):
    if weight_oz <= 4:
        return '01. 1-4 oz'
    elif weight_oz <= 8:
        return '02. 5-8 oz'
    elif weight_oz <= 12:
        return '03. 9-12 oz'
    elif weight_oz <= 15:
        return '04. 13-15 oz'
    elif weight_oz <= 15.99:
        return '05. 15.01-15.99 oz'
    elif weight_oz == 16:
        return '06. 16 oz exactly'
    elif weight_oz <= 32:
        return '07. 1-2 lbs'
    elif weight_oz <= 48:
        return '08. 2-3 lbs'
    elif weight_oz <= 64:
        return '09. 3-4 lbs'
    elif weight_oz <= 80:
        return '10. 4-5 lbs'
    else:
        return '11. Over 5 lbs'

df['weight_category'] = df['weight_oz'].apply(categorize_weight)

# Zone estimation
df['origin_zip3'] = df['Origin'].str[:3]
df['dest_zip3'] = df['Destination'].astype(str).str[:3]

def estimate_zone(origin_zip3, dest_zip3):
    try:
        diff = abs(int(origin_zip3) - int(dest_zip3))
        if diff <= 100:
            return 'Zone 1-2 (Local)'
        elif diff <= 300:
            return 'Zone 3-4 (Regional)'
        elif diff <= 500:
            return 'Zone 5-6 (National)'
        else:
            return 'Zone 7-8 (Cross-Country)'
    except:
        return 'Unknown'

df['zone_estimate'] = df.apply(lambda x: estimate_zone(x['origin_zip3'], x['dest_zip3']), axis=1)

# State mapping
zip_to_state = {
    range(10, 28): 'MA', range(28, 30): 'RI', range(30, 39): 'NH',
    range(39, 50): 'ME', range(50, 60): 'VT', range(60, 70): 'CT',
    range(100, 150): 'NY', range(70, 90): 'NJ', range(150, 197): 'PA',
    range(197, 200): 'DE', range(200, 213): 'DC', range(206, 221): 'MD',
    range(220, 247): 'VA', range(247, 269): 'WV', range(270, 290): 'NC',
    range(290, 300): 'SC', range(300, 320): 'GA', range(320, 340): 'FL',
    range(350, 370): 'AL', range(370, 380): 'TN', range(380, 400): 'MS',
    range(400, 428): 'KY', range(430, 460): 'OH', range(460, 480): 'IN',
    range(480, 500): 'MI', range(500, 529): 'IA', range(530, 550): 'WI',
    range(550, 568): 'MN', range(570, 580): 'SD', range(580, 590): 'ND',
    range(600, 630): 'IL', range(630, 660): 'MO', range(660, 680): 'KS',
    range(680, 694): 'NE', range(700, 715): 'LA', range(716, 730): 'AR',
    range(730, 750): 'OK', range(750, 800): 'TX', range(800, 817): 'CO',
    range(820, 832): 'WY', range(832, 839): 'ID', range(840, 848): 'UT',
    range(850, 866): 'AZ', range(870, 885): 'NM', range(889, 899): 'NV',
    range(900, 962): 'CA', range(970, 979): 'OR', range(980, 995): 'WA',
    range(995, 1000): 'AK', range(967, 969): 'HI'
}

def get_state(zip_code):
    try:
        zip3 = int(str(zip_code)[:3])
        for range_obj, state in zip_to_state.items():
            if zip3 in range_obj:
                return state
    except:
        pass
    return 'Unknown'

df['dest_state'] = df['Destination'].apply(get_state)

# Create Excel workbook
wb = Workbook()

# Define styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 1. Summary Sheet
ws_summary = wb.active
ws_summary.title = "Executive Summary"

summary_data = [
    ["JOSH'S FROGS - SHIPPING PROFILE ANALYSIS", "", "", ""],
    ["", "", "", ""],
    ["KEY METRICS", "", "", ""],
    ["Total Shipments", f"{len(df):,}", "Date Range", f"{df['date'].min().strftime('%m/%d/%Y')} - {df['date'].max().strftime('%m/%d/%Y')}"],
    ["Daily Average", f"{len(df) / ((df['date'].max() - df['date'].min()).days + 1):.0f}", "Total Spend", f"${df['Cost'].sum():,.2f}"],
    ["Avg Cost per Package", f"${df['Cost'].mean():.2f}", "Median Cost", f"${df['Cost'].median():.2f}"],
    ["", "", "", ""],
    ["CARRIER MIX", "", "", ""],
    ["Carrier", "Volume", "Spend", "Avg Cost"],
]

carrier_summary = df.groupby('Carrier').agg({'Number': 'count', 'Cost': 'sum'})
carrier_summary['avg_cost'] = carrier_summary['Cost'] / carrier_summary['Number']
for carrier, row in carrier_summary.iterrows():
    summary_data.append([carrier, f"{row['Number']:,}", f"${row['Cost']:,.2f}", f"${row['avg_cost']:.2f}"])

summary_data.extend([
    ["", "", "", ""],
    ["TOP OPTIMIZATION OPPORTUNITIES", "", "", ""],
    ["1. Weight Threshold Management", f"{len(df[(df['weight_oz'] > 14) & (df['weight_oz'] < 16)]):,} packages near 16 oz threshold", "", ""],
    ["2. Lightweight Focus", f"{len(df[df['Weight'] < 1]):,} packages under 1 lb ({len(df[df['Weight'] < 1])/len(df)*100:.1f}%)", "", ""],
    ["3. Billable Weight Impact", f"{len(df[df['billable_oz'] - df['weight_oz'] > df['weight_oz']*0.25]):,} packages with >25% increase", "", ""],
    ["4. Regional Concentration", f"{len(df[df['zone_estimate'].isin(['Zone 1-2 (Local)', 'Zone 3-4 (Regional)'])]):,} packages in Zones 1-4 ({len(df[df['zone_estimate'].isin(['Zone 1-2 (Local)', 'Zone 3-4 (Regional)'])])/len(df)*100:.1f}%)", "", ""],
])

for row in summary_data:
    ws_summary.append(row)

# Format summary sheet
ws_summary.merge_cells('A1:D1')
ws_summary['A1'].font = Font(bold=True, size=14)
ws_summary['A1'].alignment = Alignment(horizontal="center")

for row in [3, 8, 12]:
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)

# 2. Monthly Volume Sheet
ws_monthly = wb.create_sheet("Monthly Volumes")

monthly_stats = df.groupby('year_month').agg({
    'Number': 'count',
    'Cost': 'sum'
}).rename(columns={'Number': 'Volume'})
monthly_stats['Avg Cost'] = monthly_stats['Cost'] / monthly_stats['Volume']
monthly_stats['MoM Growth %'] = monthly_stats['Volume'].pct_change() * 100
monthly_stats.index = monthly_stats.index.to_timestamp()

ws_monthly.append(["Month", "Volume", "Total Cost", "Avg Cost", "MoM Growth %"])
for idx, row in monthly_stats.iterrows():
    ws_monthly.append([
        idx.strftime('%Y-%m'),
        row['Volume'],
        row['Cost'],
        row['Avg Cost'],
        row['MoM Growth %'] if pd.notna(row['MoM Growth %']) else 0
    ])

# 3. Carrier Mix Sheet
ws_carrier = wb.create_sheet("Carrier Analysis")

carrier_detail = df.groupby(['Carrier', 'Service']).agg({
    'Number': 'count',
    'Cost': 'sum'
}).rename(columns={'Number': 'Volume'})
carrier_detail['Avg Cost'] = carrier_detail['Cost'] / carrier_detail['Volume']
carrier_detail['Volume %'] = carrier_detail['Volume'] / len(df) * 100
carrier_detail['Spend %'] = carrier_detail['Cost'] / df['Cost'].sum() * 100

ws_carrier.append(["Carrier", "Service", "Volume", "Volume %", "Total Cost", "Spend %", "Avg Cost"])
for (carrier, service), row in carrier_detail.iterrows():
    ws_carrier.append([
        carrier,
        service,
        row['Volume'],
        row['Volume %'],
        row['Cost'],
        row['Spend %'],
        row['Avg Cost']
    ])

# 4. Weight Distribution Sheet
ws_weight = wb.create_sheet("Weight Analysis")

weight_dist = df.groupby('weight_category').agg({
    'Number': 'count',
    'Cost': 'sum'
}).rename(columns={'Number': 'Volume'})
weight_dist['Volume %'] = weight_dist['Volume'] / len(df) * 100
weight_dist['Spend %'] = weight_dist['Cost'] / df['Cost'].sum() * 100
weight_dist['Avg Cost'] = weight_dist['Cost'] / weight_dist['Volume']

ws_weight.append(["Weight Category", "Volume", "Volume %", "Total Cost", "Spend %", "Avg Cost"])
for category, row in weight_dist.iterrows():
    ws_weight.append([
        category,
        row['Volume'],
        row['Volume %'],
        row['Cost'],
        row['Spend %'],
        row['Avg Cost']
    ])

# Add billable weight analysis
ws_weight.append([])
ws_weight.append(["BILLABLE WEIGHT IMPACT ANALYSIS"])
ws_weight.append(["Metric", "Value"])
ws_weight.append(["Average Actual Weight (oz)", df['weight_oz'].mean()])
ws_weight.append(["Average Billable Weight (oz)", df['billable_oz'].mean()])
ws_weight.append(["Average Difference (oz)", df['billable_oz'].mean() - df['weight_oz'].mean()])
ws_weight.append(["Average % Increase", ((df['billable_oz'] - df['weight_oz']) / df['weight_oz'] * 100).mean()])
ws_weight.append(["Packages with >25% increase", len(df[(df['billable_oz'] - df['weight_oz']) / df['weight_oz'] > 0.25])])

# 5. Zone Distribution Sheet
ws_zone = wb.create_sheet("Zone Analysis")

zone_dist = df.groupby('zone_estimate').agg({
    'Number': 'count',
    'Cost': 'sum'
}).rename(columns={'Number': 'Volume'})
zone_dist['Volume %'] = zone_dist['Volume'] / len(df) * 100
zone_dist['Spend %'] = zone_dist['Cost'] / df['Cost'].sum() * 100
zone_dist['Avg Cost'] = zone_dist['Cost'] / zone_dist['Volume']

ws_zone.append(["Zone", "Volume", "Volume %", "Total Cost", "Spend %", "Avg Cost"])
for zone, row in zone_dist.iterrows():
    ws_zone.append([
        zone,
        row['Volume'],
        row['Volume %'],
        row['Cost'],
        row['Spend %'],
        row['Avg Cost']
    ])

# 6. Geographic Distribution Sheet
ws_geo = wb.create_sheet("Geographic Analysis")

state_dist = df.groupby('dest_state').agg({
    'Number': 'count',
    'Cost': 'sum'
}).rename(columns={'Number': 'Volume'})
state_dist['Volume %'] = state_dist['Volume'] / len(df) * 100
state_dist['Spend %'] = state_dist['Cost'] / df['Cost'].sum() * 100
state_dist['Avg Cost'] = state_dist['Cost'] / state_dist['Volume']
state_dist = state_dist.sort_values('Volume', ascending=False).head(20)

ws_geo.append(["State", "Volume", "Volume %", "Total Cost", "Spend %", "Avg Cost"])
for state, row in state_dist.iterrows():
    ws_geo.append([
        state,
        row['Volume'],
        row['Volume %'],
        row['Cost'],
        row['Spend %'],
        row['Avg Cost']
    ])

# 7. Service Level Sheet
ws_service = wb.create_sheet("Service Levels")

service_dist = df.groupby('Service').agg({
    'Number': 'count',
    'Cost': 'sum',
    'Weight': 'mean'
}).rename(columns={'Number': 'Volume', 'Weight': 'Avg Weight'})
service_dist['Volume %'] = service_dist['Volume'] / len(df) * 100
service_dist['Spend %'] = service_dist['Cost'] / df['Cost'].sum() * 100
service_dist['Avg Cost'] = service_dist['Cost'] / service_dist['Volume']
service_dist = service_dist.sort_values('Volume', ascending=False)

ws_service.append(["Service", "Volume", "Volume %", "Total Cost", "Spend %", "Avg Cost", "Avg Weight (lbs)"])
for service, row in service_dist.iterrows():
    ws_service.append([
        service,
        row['Volume'],
        row['Volume %'],
        row['Cost'],
        row['Spend %'],
        row['Avg Cost'],
        row['Avg Weight']
    ])

# 8. Cost Analysis Sheet
ws_cost = wb.create_sheet("Cost Analysis")

# Cost ranges
cost_ranges = ['$0-5', '$5-7', '$7-9', '$9-11', '$11-15', '$15-20', '$20+']
cost_bins = [0, 5, 7, 9, 11, 15, 20, float('inf')]
df['cost_range'] = pd.cut(df['Cost'], bins=cost_bins, labels=cost_ranges)

cost_dist = df.groupby('cost_range').agg({
    'Number': 'count',
    'Cost': 'sum'
}).rename(columns={'Number': 'Volume'})
cost_dist['Volume %'] = cost_dist['Volume'] / len(df) * 100
cost_dist['Spend %'] = cost_dist['Cost'] / df['Cost'].sum() * 100
cost_dist['Avg Cost'] = cost_dist['Cost'] / cost_dist['Volume']

ws_cost.append(["Cost Range", "Volume", "Volume %", "Total Cost", "Spend %", "Avg Cost"])
for range_label, row in cost_dist.iterrows():
    ws_cost.append([
        range_label,
        row['Volume'],
        row['Volume %'],
        row['Cost'],
        row['Spend %'],
        row['Avg Cost']
    ])

# Format all sheets
for ws in wb.worksheets:
    # Apply header formatting
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            if hasattr(cell, 'column_letter'):
                column_letter = cell.column_letter
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Format numbers
    for row in ws.iter_rows(min_row=2, values_only=False):
        for cell in row:
            if cell.column > 1 and cell.value is not None:
                if isinstance(cell.value, (int, float)):
                    if 'Cost' in ws.cell(1, cell.column).value or '$' in str(ws.cell(1, cell.column).value):
                        cell.number_format = '$#,##0.00'
                    elif '%' in str(ws.cell(1, cell.column).value):
                        cell.number_format = '0.0%'
                        cell.value = cell.value / 100 if cell.value > 1 else cell.value
                    elif 'Volume' in str(ws.cell(1, cell.column).value):
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '#,##0.00'

# Save the workbook
output_file = 'Joshs_Frogs_PLD_Analysis.xlsx'
wb.save(output_file)

print(f"Excel report created successfully: {output_file}")
print(f"Total sheets created: {len(wb.worksheets)}")
for ws in wb.worksheets:
    print(f"  - {ws.title}: {ws.max_row - 1} data rows")