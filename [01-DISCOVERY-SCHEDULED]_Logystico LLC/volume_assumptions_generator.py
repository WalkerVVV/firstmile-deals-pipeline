#!/usr/bin/env python3
"""
Volume Assumptions Generator for FirstMile Xparcel Service Levels
Generates shipment volume distribution across Priority, Expedited, and Ground services
"""

import pandas as pd
import numpy as np
import argparse
import sys
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Try importing optional libraries
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel formatting features will be limited.")

try:
    import matplotlib.pyplot as plt
    import seaborn as sns
    CHARTS_AVAILABLE = True
except ImportError:
    CHARTS_AVAILABLE = False
    print("Warning: matplotlib/seaborn not installed. Charts will not be generated.")


class VolumeAssumptionsGenerator:
    """Generate volume assumptions table for Xparcel service levels"""
    
    def __init__(self, total_volume=10000, service_split=None, weight_distribution=None):
        """
        Initialize generator with volume parameters
        
        Args:
            total_volume: Total shipment volume to distribute
            service_split: Dict with service level percentages {'priority': 15, 'expedited': 60, 'ground': 25}
            weight_distribution: Dict with weight range percentages
        """
        self.total_volume = total_volume
        
        # Default service level distribution
        self.service_split = service_split or {
            'priority': 0.15,
            'expedited': 0.60,
            'ground': 0.25
        }
        
        # Default weight distribution (typical ecommerce)
        self.weight_distribution = weight_distribution or {
            'under_1lb': 0.35,
            '1_to_5lbs': 0.45,
            '5_to_10lbs': 0.15,
            '10_to_25lbs': 0.05
        }
        
        self.weight_ranges = []
        self.df = None
        
    def create_weight_ranges(self):
        """Create weight range labels and boundaries"""
        ranges = []
        
        # First range: 1oz-15.99oz
        ranges.append({
            'label': '1oz-15.99oz',
            'min_oz': 1,
            'max_oz': 15.99,
            'category': 'under_1lb'
        })
        
        # 16oz-31.99oz (1lb band)
        ranges.append({
            'label': '1lb-1.99lbs',
            'min_oz': 16,
            'max_oz': 31.99,
            'category': '1_to_5lbs'
        })
        
        # Continue in 1lb increments up to 400oz (25lbs)
        for lb in range(2, 26):
            if lb == 25:
                max_oz = 400
            else:
                max_oz = (lb + 1) * 16 - 0.01
                
            label = f'{lb}lbs-{lb}.99lbs' if lb < 25 else f'{lb}lbs'
            
            # Determine category
            if lb < 5:
                category = '1_to_5lbs'
            elif lb < 10:
                category = '5_to_10lbs'
            else:
                category = '10_to_25lbs'
                
            ranges.append({
                'label': label,
                'min_oz': lb * 16,
                'max_oz': max_oz,
                'category': category
            })
        
        self.weight_ranges = ranges
        return ranges
    
    def distribute_volume(self):
        """Distribute volume across weight ranges and service levels"""
        data = []
        
        # Count ranges by category
        category_counts = {}
        for r in self.weight_ranges:
            cat = r['category']
            category_counts[cat] = category_counts.get(cat, 0) + 1
        
        # Distribute volume
        for weight_range in self.weight_ranges:
            category = weight_range['category']
            
            # Get base volume for this category
            category_volume = self.total_volume * self.weight_distribution[category]
            
            # Distribute evenly within category (with some randomness for realism)
            range_count = category_counts[category]
            base_volume = category_volume / range_count
            
            # Add some variance
            variance = np.random.uniform(0.7, 1.3)
            range_volume = int(base_volume * variance)
            
            # Distribute across service levels
            priority_vol = int(range_volume * self.service_split['priority'])
            expedited_vol = int(range_volume * self.service_split['expedited'])
            ground_vol = int(range_volume * self.service_split['ground'])
            
            # Ensure total matches
            total_vol = priority_vol + expedited_vol + ground_vol
            
            data.append({
                'Weight_Range': weight_range['label'],
                'Priority_Volume': priority_vol,
                'Expedited_Volume': expedited_vol,
                'Ground_Volume': ground_vol,
                'Total_Volume': total_vol,
                'Category': category
            })
        
        # Create DataFrame
        self.df = pd.DataFrame(data)
        
        # Calculate percentages
        total_sum = self.df['Total_Volume'].sum()
        self.df['Percent_of_Total'] = (self.df['Total_Volume'] / total_sum * 100).round(2)
        
        return self.df
    
    def add_subtotals(self):
        """Add subtotal rows for weight categories"""
        if self.df is None:
            return
        
        subtotals = []
        categories = ['under_1lb', '1_to_5lbs', '5_to_10lbs', '10_to_25lbs']
        
        for cat in categories:
            cat_df = self.df[self.df['Category'] == cat]
            if len(cat_df) > 0:
                subtotal = {
                    'Weight_Range': f'SUBTOTAL: {cat.replace("_", " ").title()}',
                    'Priority_Volume': cat_df['Priority_Volume'].sum(),
                    'Expedited_Volume': cat_df['Expedited_Volume'].sum(),
                    'Ground_Volume': cat_df['Ground_Volume'].sum(),
                    'Total_Volume': cat_df['Total_Volume'].sum(),
                    'Percent_of_Total': cat_df['Percent_of_Total'].sum(),
                    'Category': 'subtotal'
                }
                subtotals.append(subtotal)
        
        # Add grand total
        grand_total = {
            'Weight_Range': 'GRAND TOTAL',
            'Priority_Volume': self.df['Priority_Volume'].sum(),
            'Expedited_Volume': self.df['Expedited_Volume'].sum(),
            'Ground_Volume': self.df['Ground_Volume'].sum(),
            'Total_Volume': self.df['Total_Volume'].sum(),
            'Percent_of_Total': 100.00,
            'Category': 'total'
        }
        
        # Combine with subtotals
        result_df = self.df.copy()
        for subtotal in subtotals:
            result_df = pd.concat([result_df, pd.DataFrame([subtotal])], ignore_index=True)
        result_df = pd.concat([result_df, pd.DataFrame([grand_total])], ignore_index=True)
        
        # Drop the Category column for final output
        result_df = result_df.drop('Category', axis=1)
        
        return result_df
    
    def format_excel(self, filename):
        """Apply Excel formatting with color coding"""
        if not EXCEL_AVAILABLE:
            print("Excel formatting not available. Install openpyxl for full features.")
            return
        
        # Save to Excel with formatting
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            self.df.to_excel(writer, sheet_name='Volume_Assumptions', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Volume_Assumptions']
            
            # Define fills
            priority_fill = PatternFill(start_color='FFB3B3', end_color='FFB3B3', fill_type='solid')  # Light red
            expedited_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')  # Light yellow
            ground_fill = PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid')  # Light green
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')  # Dark blue
            subtotal_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # Gray
            
            # Define fonts
            header_font = Font(color='FFFFFF', bold=True, size=11)
            bold_font = Font(bold=True)
            
            # Define borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format headers
            for col in range(1, 7):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Format data rows
            for row_num in range(2, worksheet.max_row + 1):
                weight_cell = worksheet.cell(row=row_num, column=1)
                
                # Check if it's a subtotal or total row
                if 'SUBTOTAL' in str(weight_cell.value) or 'TOTAL' in str(weight_cell.value):
                    for col in range(1, 7):
                        cell = worksheet.cell(row=row_num, column=col)
                        cell.fill = subtotal_fill
                        cell.font = bold_font
                        cell.border = thin_border
                else:
                    # Apply service level colors
                    worksheet.cell(row=row_num, column=2).fill = priority_fill  # Priority
                    worksheet.cell(row=row_num, column=3).fill = expedited_fill  # Expedited
                    worksheet.cell(row=row_num, column=4).fill = ground_fill  # Ground
                
                # Apply borders to all cells
                for col in range(1, 7):
                    worksheet.cell(row=row_num, column=col).border = thin_border
            
            # Format numbers with comma separators
            for row_num in range(2, worksheet.max_row + 1):
                for col in [2, 3, 4, 5]:  # Volume columns
                    cell = worksheet.cell(row=row_num, column=col)
                    cell.number_format = '#,##0'
                
                # Format percentage column
                cell_value = worksheet.cell(row=row_num, column=6).value
                if cell_value and not isinstance(cell_value, str):
                    worksheet.cell(row=row_num, column=6).number_format = '0.00%'
                    worksheet.cell(row=row_num, column=6).value = cell_value / 100
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 20
            worksheet.column_dimensions['B'].width = 15
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 15
            worksheet.column_dimensions['E'].width = 15
            worksheet.column_dimensions['F'].width = 15
    
    def create_charts(self, output_base):
        """Create charts for volume distribution"""
        if not CHARTS_AVAILABLE:
            print("Charts not available. Install matplotlib and seaborn for visualization.")
            return
        
        # Set style
        sns.set_style("whitegrid")
        fig, axes = plt.subplots(2, 2, figsize=(15, 12))
        
        # Filter out subtotals for charts
        chart_df = self.df[~self.df['Weight_Range'].str.contains('TOTAL')]
        
        # 1. Stacked bar chart by weight range
        ax1 = axes[0, 0]
        x_pos = np.arange(len(chart_df))
        
        ax1.bar(x_pos, chart_df['Ground_Volume'], label='Ground', color='#99FF99')
        ax1.bar(x_pos, chart_df['Expedited_Volume'], bottom=chart_df['Ground_Volume'], 
               label='Expedited', color='#FFFF99')
        ax1.bar(x_pos, chart_df['Priority_Volume'], 
               bottom=chart_df['Ground_Volume'] + chart_df['Expedited_Volume'],
               label='Priority', color='#FFB3B3')
        
        ax1.set_xlabel('Weight Range')
        ax1.set_ylabel('Volume')
        ax1.set_title('Volume Distribution by Weight Range')
        ax1.set_xticks(x_pos[::3])  # Show every 3rd label to avoid crowding
        ax1.set_xticklabels(chart_df['Weight_Range'].iloc[::3], rotation=45, ha='right')
        ax1.legend()
        ax1.grid(True, alpha=0.3)
        
        # 2. Pie chart for service level mix
        ax2 = axes[0, 1]
        service_totals = [
            chart_df['Priority_Volume'].sum(),
            chart_df['Expedited_Volume'].sum(),
            chart_df['Ground_Volume'].sum()
        ]
        colors = ['#FFB3B3', '#FFFF99', '#99FF99']
        labels = ['Priority (15%)', 'Expedited (60%)', 'Ground (25%)']
        
        ax2.pie(service_totals, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
        ax2.set_title('Service Level Mix')
        
        # 3. Weight category distribution
        ax3 = axes[1, 0]
        categories = ['Under 1lb', '1-5 lbs', '5-10 lbs', '10-25 lbs']
        cat_volumes = []
        
        for cat in ['under_1lb', '1_to_5lbs', '5_to_10lbs', '10_to_25lbs']:
            cat_df = self.df[self.df['Category'] == cat] if 'Category' in self.df.columns else pd.DataFrame()
            cat_volumes.append(cat_df['Total_Volume'].sum() if len(cat_df) > 0 else 0)
        
        ax3.bar(categories, cat_volumes, color=['#E8F4FD', '#B8D4E8', '#6FA3C5', '#366092'])
        ax3.set_xlabel('Weight Category')
        ax3.set_ylabel('Total Volume')
        ax3.set_title('Volume by Weight Category')
        ax3.grid(True, alpha=0.3)
        
        # 4. Cumulative distribution
        ax4 = axes[1, 1]
        cumulative = chart_df['Percent_of_Total'].cumsum()
        ax4.plot(range(len(cumulative)), cumulative, marker='o', linewidth=2, markersize=4)
        ax4.fill_between(range(len(cumulative)), cumulative, alpha=0.3)
        ax4.set_xlabel('Weight Range Index')
        ax4.set_ylabel('Cumulative Percentage')
        ax4.set_title('Cumulative Volume Distribution')
        ax4.grid(True, alpha=0.3)
        ax4.axhline(y=80, color='r', linestyle='--', alpha=0.5, label='80% threshold')
        ax4.legend()
        
        plt.tight_layout()
        plt.savefig(f'{output_base}_charts.png', dpi=300, bbox_inches='tight')
        plt.show()
        
        print(f"Charts saved to {output_base}_charts.png")
    
    def process_input_file(self, filename):
        """Process actual shipment data file"""
        try:
            # Try reading as CSV first
            if filename.endswith('.csv'):
                df = pd.read_csv(filename)
            else:
                df = pd.read_excel(filename)
            
            print(f"Loaded {len(df)} records from {filename}")
            
            # Expected columns: weight (in lbs or oz), service_level
            if 'Weight (lb)' in df.columns:
                df['weight_oz'] = pd.to_numeric(df['Weight (lb)'], errors='coerce') * 16
            elif 'weight_oz' in df.columns:
                pass  # Already in oz
            else:
                print("Warning: No weight column found. Using default distribution.")
                return
            
            # Calculate actual distribution
            total = len(df)
            
            # Calculate weight distribution
            under_1lb = len(df[df['weight_oz'] < 16]) / total
            lb_1_to_5 = len(df[(df['weight_oz'] >= 16) & (df['weight_oz'] < 80)]) / total
            lb_5_to_10 = len(df[(df['weight_oz'] >= 80) & (df['weight_oz'] < 160)]) / total
            lb_10_to_25 = len(df[(df['weight_oz'] >= 160) & (df['weight_oz'] <= 400)]) / total
            
            self.weight_distribution = {
                'under_1lb': under_1lb,
                '1_to_5lbs': lb_1_to_5,
                '5_to_10lbs': lb_5_to_10,
                '10_to_25lbs': lb_10_to_25
            }
            
            print(f"Actual weight distribution calculated from data:")
            print(f"  Under 1lb: {under_1lb:.1%}")
            print(f"  1-5 lbs: {lb_1_to_5:.1%}")
            print(f"  5-10 lbs: {lb_5_to_10:.1%}")
            print(f"  10-25 lbs: {lb_10_to_25:.1%}")
            
            self.total_volume = total
            
        except Exception as e:
            print(f"Error processing input file: {e}")
            print("Using default distribution instead.")


def main():
    """Main execution function"""
    parser = argparse.ArgumentParser(description='Generate volume assumptions table for Xparcel services')
    parser.add_argument('--input_file', type=str, help='Source data CSV/Excel file')
    parser.add_argument('--output_name', type=str, help='Output filename (without extension)')
    parser.add_argument('--total_volume', type=int, default=10000, help='Total shipment volume')
    parser.add_argument('--service_split', type=str, help='Service split as "priority,expedited,ground" (e.g., "15,60,25")')
    parser.add_argument('--weight_distribution', type=str, help='Weight distribution as "under1,1to5,5to10,10to25" (e.g., "35,45,15,5")')
    parser.add_argument('--no_charts', action='store_true', help='Skip chart generation')
    
    args = parser.parse_args()
    
    # Parse service split if provided
    service_split = None
    if args.service_split:
        splits = [float(x) for x in args.service_split.split(',')]
        if len(splits) == 3 and sum(splits) == 100:
            service_split = {
                'priority': splits[0] / 100,
                'expedited': splits[1] / 100,
                'ground': splits[2] / 100
            }
        else:
            print("Warning: Service split must be 3 values that sum to 100. Using defaults.")
    
    # Parse weight distribution if provided
    weight_dist = None
    if args.weight_distribution:
        weights = [float(x) for x in args.weight_distribution.split(',')]
        if len(weights) == 4 and sum(weights) == 100:
            weight_dist = {
                'under_1lb': weights[0] / 100,
                '1_to_5lbs': weights[1] / 100,
                '5_to_10lbs': weights[2] / 100,
                '10_to_25lbs': weights[3] / 100
            }
        else:
            print("Warning: Weight distribution must be 4 values that sum to 100. Using defaults.")
    
    # Create generator
    generator = VolumeAssumptionsGenerator(
        total_volume=args.total_volume,
        service_split=service_split,
        weight_distribution=weight_dist
    )
    
    # Process input file if provided
    if args.input_file:
        generator.process_input_file(args.input_file)
    
    # Generate weight ranges
    generator.create_weight_ranges()
    
    # Distribute volume
    generator.distribute_volume()
    
    # Add subtotals
    final_df = generator.add_subtotals()
    
    # Generate output filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_base = args.output_name or f'volume_assumptions_{timestamp}'
    
    # Save to CSV
    csv_filename = f'{output_base}.csv'
    final_df.to_csv(csv_filename, index=False)
    print(f"CSV saved to {csv_filename}")
    
    # Save to Excel with formatting
    if EXCEL_AVAILABLE:
        excel_filename = f'{output_base}.xlsx'
        generator.format_excel(excel_filename)
        print(f"Excel saved to {excel_filename}")
    
    # Create charts
    if not args.no_charts and CHARTS_AVAILABLE:
        generator.create_charts(output_base)
    
    # Display summary
    print("\n" + "="*60)
    print("VOLUME ASSUMPTIONS SUMMARY")
    print("="*60)
    print(f"Total Volume: {generator.total_volume:,}")
    print(f"\nService Level Distribution:")
    print(f"  Priority (15%): {int(generator.total_volume * generator.service_split['priority']):,}")
    print(f"  Expedited (60%): {int(generator.total_volume * generator.service_split['expedited']):,}")
    print(f"  Ground (25%): {int(generator.total_volume * generator.service_split['ground']):,}")
    print(f"\nWeight Category Distribution:")
    print(f"  Under 1lb: {generator.weight_distribution['under_1lb']:.1%}")
    print(f"  1-5 lbs: {generator.weight_distribution['1_to_5lbs']:.1%}")
    print(f"  5-10 lbs: {generator.weight_distribution['5_to_10lbs']:.1%}")
    print(f"  10-25 lbs: {generator.weight_distribution['10_to_25lbs']:.1%}")
    print("="*60)


if __name__ == '__main__':
    main()