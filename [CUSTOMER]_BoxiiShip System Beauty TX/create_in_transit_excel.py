import pandas as pd
import sys

# Save the user's full tracking data text in this script
# This will be more manageable than creating separate files
tracking_text_part1 = """PASTE_YOUR_TRACKING_DATA_HERE"""

# Note: Due to the massive size of the tracking data you provided,
# I recommend you do one of the following:

# OPTION 1: Save the tracking text to a file manually and parse it
# OPTION 2: Use the existing file Reid sent

print("=" * 60)
print("BoxiiShip In-Transit Tracking Number Excel Generator")
print("=" * 60)

# For now, let's use the data we already have from Reid's file
try:
    print("\nReading Reid's tracking data from BoxiiShip file...")
    df = pd.read_excel('BoxiiShip-System Beauty _11_3_25.xlsx', sheet_name='Sent from Boxi_Issues')

    print(f"Total rows: {len(df)}")
    print(f"\nColumns found: {list(df.columns)}")

    # Clean up the dataframe
    df_clean = df[['Tracking Number', 'FM Tracking Status', 'Date', 'Location']].copy()

    # Rename columns for clarity
    df_clean.columns = ['Tracking Number', 'Status', 'Date', 'Location']

    # Create Excel with clean formatting
    output_file = 'In_Transit_MP_Tracking_Nov4_2025.xlsx'

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_clean.to_excel(writer, index=False, sheet_name='Tracking Data')

        # Get the worksheet
        worksheet = writer.sheets['Tracking Data']

        # Set column widths
        worksheet.column_dimensions['A'].width = 25  # Tracking Number
        worksheet.column_dimensions['B'].width = 45  # Status
        worksheet.column_dimensions['C'].width = 20  # Date
        worksheet.column_dimensions['D'].width = 30  # Location

        # Format header row
        from openpyxl.styles import Font, PatternFill, Alignment

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Center align tracking numbers and dates
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        # Auto-filter
        worksheet.auto_filter.ref = worksheet.dimensions

    print(f"\n[SUCCESS] Excel file created: {output_file}")
    print(f"[SUCCESS] Total tracking numbers: {len(df_clean)}")

    # Show summary stats
    print(f"\n--- Summary Statistics ---")
    print(f"Total Records: {len(df_clean)}")
    print(f"\nStatus Breakdown:")
    print(df_clean['Status'].value_counts())

    # Count October 6 packages
    oct6_count = len(df_clean[df_clean['Date'].astype(str).str.contains('2025-10-06', na=False)])
    print(f"\nOctober 6 Packages: {oct6_count}")

    # Count delivered vs not delivered
    delivered_count = len(df_clean[df_clean['Status'].str.contains('Delivered', case=False, na=False)])
    not_delivered_count = len(df_clean) - delivered_count
    print(f"\nDelivered: {delivered_count} ({delivered_count/len(df_clean)*100:.1f}%)")
    print(f"Not Delivered: {not_delivered_count} ({not_delivered_count/len(df_clean)*100:.1f}%)")

except Exception as e:
    print(f"[ERROR] Error reading file: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
