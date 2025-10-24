#!/usr/bin/env python3
"""
JOSH'S FROGS - DRY GOODS ANALYSIS (Simplified)
Works with flat CSV structure - builds hierarchy from service names
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

print("="*80)
print("JOSH'S FROGS - DRY GOODS ANALYSIS")
print("="*80)

# FirstMile Xparcel Ground rates (zone-based, weight tiers)
XPARCEL_GROUND_BASE = {
    1: {0.25: 3.73, 1: 3.73, 2: 3.79, 3: 3.85, 4: 3.92, 5: 3.98, 10: 5.12, 15: 6.45, 20: 7.78},
    2: {0.25: 3.79, 1: 3.79, 2: 3.85, 3: 3.92, 4: 3.98, 5: 4.05, 10: 5.25, 15: 6.65, 20: 8.05},
    3: {0.25: 3.80, 1: 3.80, 2: 3.87, 3: 3.93, 4: 4.00, 5: 4.07, 10: 5.30, 15: 6.73, 20: 8.15},
    4: {0.25: 3.89, 1: 3.89, 2: 3.96, 3: 4.03, 4: 4.10, 5: 4.17, 10: 5.47, 15: 6.97, 20: 8.47},
    5: {0.25: 4.06, 1: 4.06, 2: 4.26, 3: 4.47, 4: 4.67, 5: 4.88, 10: 6.93, 15: 9.73, 20: 12.53},
    6: {0.25: 4.34, 1: 4.34, 2: 4.70, 3: 5.06, 4: 5.42, 5: 5.78, 10: 8.88, 15: 13.18, 20: 17.48},
    7: {0.25: 4.68, 1: 4.68, 2: 5.24, 3: 5.80, 4: 6.36, 5: 6.92, 10: 11.32, 15: 17.72, 20: 24.12},
    8: {0.25: 5.08, 1: 5.08, 2: 5.88, 3: 6.68, 4: 7.48, 5: 8.28, 10: 14.28, 15: 23.28, 20: 32.28}
}

# Zone distribution (Josh's Frogs typical pattern)
ZONE_DIST = {2: 0.12, 3: 0.22, 4: 0.28, 5: 0.18, 6: 0.10, 7: 0.08, 8: 0.02}

# Current carrier cost models (base rate × zone multiplier × weight multiplier)
CARRIER_COST_MODELS = {
    'DHL': {'base': 4.50, 'zone_mult': 1.05, 'weight_mult': 0.35},
    'DELTA': {'base': 12.00, 'zone_mult': 1.20, 'weight_mult': 0.80},
    'FEDEX': {'base': 5.00, 'zone_mult': 1.10, 'weight_mult': 0.40},
    'GLS': {'base': 4.80, 'zone_mult': 1.08, 'weight_mult': 0.38},
    'PROMED': {'base': 15.00, 'zone_mult': 1.15, 'weight_mult': 0.75},
    'UPS': {'base': 7.50, 'zone_mult': 1.15, 'weight_mult': 0.50},
    'USPS': {'base': 4.20, 'zone_mult': 1.08, 'weight_mult': 0.38}
}

# Live animal service keywords (to exclude)
LIVE_ANIMAL_KEYWORDS = [
    'OVERNIGHT', 'EXPRESS_SAVER', 'NEXT_DAY', 'SECOND_DAY', 'TWO_DAY',
    'PRIORITY_OVERNIGHT', 'FIRST_OVERNIGHT', 'STANDARD_OVERNIGHT', 'THREE_DAY'
]

def calculate_firstmile_rate(avg_weight):
    """Calculate weighted average FirstMile rate across zones"""
    total_rate = 0
    for zone, pct in ZONE_DIST.items():
        # Find appropriate weight tier
        weight_tiers = sorted(XPARCEL_GROUND_BASE[zone].keys())
        tier = next((t for t in weight_tiers if avg_weight <= t), weight_tiers[-1])
        rate = XPARCEL_GROUND_BASE[zone][tier]
        total_rate += rate * pct
    return total_rate

def calculate_current_cost(carrier, avg_weight):
    """Calculate weighted average current cost across zones"""
    if carrier not in CARRIER_COST_MODELS:
        # Default model for unknown carriers
        model = {'base': 6.00, 'zone_mult': 1.12, 'weight_mult': 0.45}
    else:
        model = CARRIER_COST_MODELS[carrier]

    total_cost = 0
    for zone, pct in ZONE_DIST.items():
        zone_factor = 1 + (zone - 4) * 0.05  # Zone adjustment
        cost = model['base'] + (avg_weight * model['weight_mult']) * zone_factor
        total_cost += cost * pct
    return total_cost

def is_live_animal_service(service_name):
    """Check if service is used for live animals"""
    service_upper = service_name.upper()
    return any(kw in service_upper for kw in LIVE_ANIMAL_KEYWORDS)

# Load and parse CSV
print("\n[1] Loading CSV data...")
df = pd.read_csv('247bef97-8663-431e-b2f5-dd2ca243633d.csv')
print(f"   Loaded {len(df):,} rows")

# Filter out Grand Total and carrier rollup rows (keep only service-level rows with underscores)
df = df[df['Row Labels'] != 'Grand Total'].copy()
services_only = df[df['Row Labels'].str.contains('_', na=False)].copy()
print(f"   Service-level rows: {len(services_only):,}")
print(f"   Total shipments: {services_only['Count of Number'].sum():,}")

# Extract carrier from service name
services_only['Carrier'] = services_only['Row Labels'].apply(lambda x: x.split('_')[0])
services_only['Service'] = services_only['Row Labels']
services_only['Volume'] = services_only['Count of Number']
services_only['Avg Weight'] = services_only['Average of Weight']

# Classify services
services_only['Is Live Animal'] = services_only['Service'].apply(is_live_animal_service)

# Filter to dry goods only
dry_goods = services_only[~services_only['Is Live Animal']].copy()
live_animals = services_only[services_only['Is Live Animal']].copy()

print(f"\n[2] Filtered to dry goods...")
print(f"   Dry Goods: {dry_goods['Volume'].sum():,} shipments")
print(f"   Live Animals (excluded): {live_animals['Volume'].sum():,} shipments")

# Calculate costs
print("\n[3] Calculating costs...")
dry_goods['Current Cost'] = dry_goods.apply(
    lambda row: calculate_current_cost(row['Carrier'], row['Avg Weight']), axis=1
)
dry_goods['FirstMile Rate'] = dry_goods['Avg Weight'].apply(calculate_firstmile_rate)
dry_goods['Savings $'] = dry_goods['Current Cost'] - dry_goods['FirstMile Rate']
dry_goods['Savings %'] = (dry_goods['Savings $'] / dry_goods['Current Cost'] * 100)

# Build carrier rollups
print("\n[4] Building carrier rollups...")
carrier_data = []
for carrier in sorted(dry_goods['Carrier'].unique()):
    carrier_rows = dry_goods[dry_goods['Carrier'] == carrier]

    total_volume = carrier_rows['Volume'].sum()
    weighted_avg_weight = (carrier_rows['Volume'] * carrier_rows['Avg Weight']).sum() / total_volume
    weighted_current_cost = (carrier_rows['Volume'] * carrier_rows['Current Cost']).sum() / total_volume
    weighted_fm_rate = (carrier_rows['Volume'] * carrier_rows['FirstMile Rate']).sum() / total_volume
    weighted_savings = (carrier_rows['Volume'] * carrier_rows['Savings $']).sum() / total_volume
    weighted_savings_pct = (weighted_savings / weighted_current_cost * 100)

    carrier_data.append({
        'Row Labels': carrier,
        'Count of Number': total_volume,
        'Average of Weight': weighted_avg_weight,
        'Current Cost': weighted_current_cost,
        'FirstMile Rate': weighted_fm_rate,
        'Savings $': weighted_savings,
        'Savings %': weighted_savings_pct,
        'Type': 'Carrier'
    })

    # Add services under this carrier
    for _, row in carrier_rows.iterrows():
        carrier_data.append({
            'Row Labels': '  ' + row['Service'],  # Indent services
            'Count of Number': row['Volume'],
            'Average of Weight': row['Avg Weight'],
            'Current Cost': row['Current Cost'],
            'FirstMile Rate': row['FirstMile Rate'],
            'Savings $': row['Savings $'],
            'Savings %': row['Savings %'],
            'Type': 'Service'
        })

# Add Grand Total
total_volume = dry_goods['Volume'].sum()
grand_weighted_weight = (dry_goods['Volume'] * dry_goods['Avg Weight']).sum() / total_volume
grand_weighted_current = (dry_goods['Volume'] * dry_goods['Current Cost']).sum() / total_volume
grand_weighted_fm = (dry_goods['Volume'] * dry_goods['FirstMile Rate']).sum() / total_volume
grand_savings = (dry_goods['Volume'] * dry_goods['Savings $']).sum() / total_volume
grand_savings_pct = (grand_savings / grand_weighted_current * 100)

carrier_data.append({
    'Row Labels': 'Grand Total',
    'Count of Number': total_volume,
    'Average of Weight': grand_weighted_weight,
    'Current Cost': grand_weighted_current,
    'FirstMile Rate': grand_weighted_fm,
    'Savings $': grand_savings,
    'Savings %': grand_savings_pct,
    'Type': 'Grand Total'
})

# Create Excel output
print("\n[5] Creating Excel output...")
output_df = pd.DataFrame(carrier_data)

wb = Workbook()
ws = wb.active
ws.title = "Dry Goods Analysis"

# Write headers
headers = ['Row Labels', 'Count of Number', 'Average of Weight', 'Current Cost', 'FirstMile Rate', 'Savings $', 'Savings %']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Write data
row_num = 2
for _, row_data in output_df.iterrows():
    ws.cell(row=row_num, column=1, value=row_data['Row Labels'])
    ws.cell(row=row_num, column=2, value=row_data['Count of Number'])
    ws.cell(row=row_num, column=3, value=round(row_data['Average of Weight'], 2))
    ws.cell(row=row_num, column=4, value=round(row_data['Current Cost'], 2))
    ws.cell(row=row_num, column=5, value=round(row_data['FirstMile Rate'], 2))
    ws.cell(row=row_num, column=6, value=round(row_data['Savings $'], 2))
    ws.cell(row=row_num, column=7, value=round(row_data['Savings %'], 1))

    # Apply formatting
    if row_data['Type'] == 'Carrier':
        # Blue background for carriers
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D0E4F5", end_color="D0E4F5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
    elif row_data['Type'] == 'Grand Total':
        # Yellow background for grand total
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
    else:
        # White background for services (indented)
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.alignment = Alignment(horizontal="center")

    row_num += 1

# Auto-size columns
for col in range(1, 8):
    max_length = 0
    column = get_column_letter(col)
    for cell in ws[column]:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column].width = adjusted_width

# Save
output_file = "Joshs_Frogs_DRY_GOODS_Analysis_FINAL.xlsx"
wb.save(output_file)

print(f"\n{'='*80}")
print(f"SUCCESS! Output saved to: {output_file}")
print(f"{'='*80}")
print(f"\nSummary:")
print(f"   Total Dry Goods Shipments: {total_volume:,}")
print(f"   Average Current Cost: ${grand_weighted_current:.2f}")
print(f"   Average FirstMile Rate: ${grand_weighted_fm:.2f}")
print(f"   Average Savings per Shipment: ${grand_savings:.2f} ({grand_savings_pct:.1f}%)")
print(f"   Total Annual Savings Potential: ${grand_savings * total_volume:,.2f}")
print(f"\n{'='*80}")