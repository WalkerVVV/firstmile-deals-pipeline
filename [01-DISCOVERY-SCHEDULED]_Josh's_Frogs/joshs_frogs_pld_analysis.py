#!/usr/bin/env python3
"""
JOSH'S FROGS - PLD ANALYSIS
Using actual per-shipment data with real costs
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

print("="*80)
print("JOSH'S FROGS - PLD DRY GOODS ANALYSIS")
print("Using Actual Per-Shipment Data")
print("="*80)

# FirstMile Xparcel Ground rates (zone-based, weight tiers)
XPARCEL_GROUND_BASE = {
    1: {0.25: 3.73, 1: 3.73, 2: 3.79, 3: 3.85, 4: 3.92, 5: 3.98, 10: 5.12, 15: 6.45, 20: 7.78},
    2: {0.25: 3.79, 1: 3.79, 2: 3.85, 3: 3.92, 4: 3.98, 5: 4.05, 10: 5.25, 15: 6.65, 20: 8.05},
    3: {0.25: 3.80, 1: 3.80, 2: 3.87, 3: 3.93, 4: 4.00, 5: 4.07, 10: 5.30, 15: 6.73, 20: 8.15},
    4: {0.25: 3.89, 1: 3.89, 2: 3.96, 3: 4.03, 4: 4.10, 5: 4.17, 10: 5.47, 15: 6.97, 20: 8.47},
    5: {0.25: 4.06, 1: 4.06, 2: 4.26, 3: 4.47, 4: 4.67, 5: 4.88, 10: 6.93, 15: 9.73, 20: 12.53},
    6: {0.25: 4.34, 1: 4.34, 2: 4.70, 3: 5.06, 4: 5.42, 5: 5.78, 10: 8.88, 15: 13.18, 20: 17.48},
    7: {0.25: 4.68, 1: 4.68, 2: 5.24, 3: 5.80, 4: 6.36, 5: 6.92, 10: 11.32, 15: 17.72, 20: 24.12},
    8: {0.25: 5.08, 1: 5.08, 2: 5.88, 3: 6.68, 4: 7.48, 5: 8.28, 10: 14.28, 15: 23.28, 20: 32.28}
}

# Zone lookup table (simplified - first digit of ZIP to zone mapping from PA)
def estimate_zone(origin_zip, dest_zip):
    """Estimate shipping zone based on ZIP codes"""
    try:
        origin_3 = str(origin_zip)[:3]
        dest_3 = str(dest_zip)[:3]

        # Pennsylvania origin (195xx)
        if origin_3 == '195':
            dest_first = dest_3[0]
            # Zone mapping (simplified)
            zone_map = {
                '0': 2, '1': 2, '2': 2, '3': 3, '4': 3,
                '5': 4, '6': 4, '7': 5, '8': 6, '9': 7
            }
            return zone_map.get(dest_first, 4)
    except:
        pass
    return 4  # Default to Zone 4

def calculate_firstmile_rate(weight, zone):
    """Calculate FirstMile rate for given weight and zone"""
    if zone not in XPARCEL_GROUND_BASE:
        zone = 4  # Default

    weight_tiers = sorted(XPARCEL_GROUND_BASE[zone].keys())
    tier = next((t for t in weight_tiers if weight <= t), weight_tiers[-1])
    return XPARCEL_GROUND_BASE[zone][tier]

# Live animal service keywords (to exclude)
LIVE_ANIMAL_KEYWORDS = [
    'OVERNIGHT', 'EXPRESS_SAVER', 'NEXT_DAY', 'SECOND_DAY', 'TWO_DAY',
    'PRIORITY_OVERNIGHT', 'FIRST_OVERNIGHT', 'STANDARD_OVERNIGHT', 'THREE_DAY'
]

def is_live_animal_service(service_name):
    """Check if service is used for live animals"""
    service_upper = str(service_name).upper()
    return any(kw in service_upper for kw in LIVE_ANIMAL_KEYWORDS)

# Load PLD data
print("\n[1] Loading PLD data...")
df = pd.read_csv('247bef97-8663-431e-b2f5-dd2ca243633d (1).csv')
print(f"   Loaded {len(df):,} shipments")

# Clean data
df['Origin'] = df['Origin'].astype(str).str.split('-').str[0]
df['Destination'] = df['Destination'].astype(str).str.split('-').str[0]

# Classify services
print("\n[2] Classifying services...")
df['Is_Live_Animal'] = df['Service'].apply(is_live_animal_service)

# Remove rows with missing Carrier or Service
df = df[df['Carrier'].notna() & df['Service'].notna()].copy()

dry_goods = df[~df['Is_Live_Animal']].copy()
live_animals = df[df['Is_Live_Animal']].copy()

print(f"   Dry Goods: {len(dry_goods):,} shipments")
print(f"   Live Animals (excluded): {len(live_animals):,} shipments")

# Calculate zones
print("\n[3] Calculating zones...")
dry_goods['Zone'] = dry_goods.apply(lambda row: estimate_zone(row['Origin'], row['Destination']), axis=1)

# Calculate FirstMile rates
print("\n[4] Calculating FirstMile rates...")
dry_goods['FirstMile_Rate'] = dry_goods.apply(
    lambda row: calculate_firstmile_rate(row['Weight'], row['Zone']), axis=1
)
dry_goods['Savings_Dollar'] = dry_goods['Cost'] - dry_goods['FirstMile_Rate']
dry_goods['Savings_Pct'] = (dry_goods['Savings_Dollar'] / dry_goods['Cost'] * 100)

# Create pivot table structure
print("\n[5] Creating pivot table...")

# Build carrier/service hierarchy
results = []

for carrier in sorted(dry_goods['Carrier'].unique()):
    carrier_data = dry_goods[dry_goods['Carrier'] == carrier]

    # Carrier rollup
    carrier_volume = len(carrier_data)
    carrier_avg_cost = carrier_data['Cost'].mean()
    carrier_avg_weight = carrier_data['Weight'].mean()
    carrier_avg_fm = carrier_data['FirstMile_Rate'].mean()
    carrier_savings = carrier_avg_cost - carrier_avg_fm
    carrier_savings_pct = (carrier_savings / carrier_avg_cost * 100)

    results.append({
        'Row Labels': carrier,
        'Count of Number': carrier_volume,
        'Average of Cost': carrier_avg_cost,
        'Average of Weight': carrier_avg_weight,
        'FirstMile Rate': carrier_avg_fm,
        'Savings $': carrier_savings,
        'Savings %': carrier_savings_pct,
        'Type': 'Carrier'
    })

    # Services under this carrier
    for service in sorted(carrier_data['Service'].unique()):
        service_data = carrier_data[carrier_data['Service'] == service]

        service_volume = len(service_data)
        service_avg_cost = service_data['Cost'].mean()
        service_avg_weight = service_data['Weight'].mean()
        service_avg_fm = service_data['FirstMile_Rate'].mean()
        service_savings = service_avg_cost - service_avg_fm
        service_savings_pct = (service_savings / service_avg_cost * 100)

        results.append({
            'Row Labels': '  ' + service,  # Indent services
            'Count of Number': service_volume,
            'Average of Cost': service_avg_cost,
            'Average of Weight': service_avg_weight,
            'FirstMile Rate': service_avg_fm,
            'Savings $': service_savings,
            'Savings %': service_savings_pct,
            'Type': 'Service'
        })

# Grand Total
grand_volume = len(dry_goods)
grand_avg_cost = dry_goods['Cost'].mean()
grand_avg_weight = dry_goods['Weight'].mean()
grand_avg_fm = dry_goods['FirstMile_Rate'].mean()
grand_savings = grand_avg_cost - grand_avg_fm
grand_savings_pct = (grand_savings / grand_avg_cost * 100)

results.append({
    'Row Labels': 'Grand Total',
    'Count of Number': grand_volume,
    'Average of Cost': grand_avg_cost,
    'Average of Weight': grand_avg_weight,
    'FirstMile Rate': grand_avg_fm,
    'Savings $': grand_savings,
    'Savings %': grand_savings_pct,
    'Type': 'Grand Total'
})

# Create Excel output
print("\n[6] Creating Excel output...")
output_df = pd.DataFrame(results)

wb = Workbook()
ws = wb.active
ws.title = "Dry Goods Analysis"

# Write headers
headers = ['Row Labels', 'Count of Number', 'Average of Cost', 'Average of Weight',
           'FirstMile Rate', 'Savings $', 'Savings %']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Write data
row_num = 2
for _, row_data in output_df.iterrows():
    ws.cell(row=row_num, column=1, value=row_data['Row Labels'])
    ws.cell(row=row_num, column=2, value=row_data['Count of Number'])
    ws.cell(row=row_num, column=3, value=round(row_data['Average of Cost'], 2))
    ws.cell(row=row_num, column=4, value=round(row_data['Average of Weight'], 2))
    ws.cell(row=row_num, column=5, value=round(row_data['FirstMile Rate'], 2))
    ws.cell(row=row_num, column=6, value=round(row_data['Savings $'], 2))
    ws.cell(row=row_num, column=7, value=round(row_data['Savings %'], 1))

    # Apply formatting
    if row_data['Type'] == 'Carrier':
        # Blue background for carriers
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D0E4F5", end_color="D0E4F5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
    elif row_data['Type'] == 'Grand Total':
        # Yellow background for grand total
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
    else:
        # White background for services (indented)
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.alignment = Alignment(horizontal="center")

    row_num += 1

# Auto-size columns
for col in range(1, 8):
    max_length = 0
    column = get_column_letter(col)
    for cell in ws[column]:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column].width = adjusted_width

# Save
output_file = "Joshs_Frogs_DRY_GOODS_PLD_Analysis.xlsx"
wb.save(output_file)

print(f"\n{'='*80}")
print(f"SUCCESS! Output saved to: {output_file}")
print(f"{'='*80}")
print(f"\nSummary:")
print(f"   Total Dry Goods Shipments: {grand_volume:,}")
print(f"   Average Current Cost: ${grand_avg_cost:.2f}")
print(f"   Average FirstMile Rate: ${grand_avg_fm:.2f}")
print(f"   Average Savings per Shipment: ${grand_savings:.2f} ({grand_savings_pct:.1f}%)")
print(f"   Total Annual Savings Potential: ${grand_savings * grand_volume:,.2f}")
print(f"\n{'='*80}")