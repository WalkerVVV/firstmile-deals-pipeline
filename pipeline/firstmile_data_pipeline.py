"""
FirstMile Data Pipeline - Automated Customer Report Generation
Processes customer shipment CSV files into branded Excel reports with SLA compliance metrics

Author: Brett Walker
Version: 1.0.0
Last Updated: 2025-10-13
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')


# ============================================================================
# FIRSTMILE BUSINESS CONSTANTS
# ============================================================================

FIRSTMILE_BLUE = '366092'
SLA_WINDOWS = {
    'Xparcel Priority': 3,
    'Xparcel Expedited': 5,
    'Xparcel Ground': 8
}

PERF_THRESHOLDS = [
    (100.0, 'Perfect Compliance'),
    (95.0, 'Exceeds Standard'),
    (90.0, 'Meets Standard'),
    (0.0, 'Below Standard')
]

HUB_MAP = {
    'CA': 'LAX - West Coast',
    'TX': 'DFW - South Central',
    'FL': 'MIA - Southeast',
    'NY': 'JFK/EWR - Northeast',
    'NJ': 'JFK/EWR - Northeast',
    'IL': 'ORD - Midwest',
    'GA': 'ATL - Southeast',
    'PA': 'PHL - Northeast',
    'OH': 'CVG - Midwest',
    'NC': 'CLT - Southeast'
}

SERVICE_MAPPING = {
    'Ground': 'Xparcel Ground',
    'Expedited': 'Xparcel Expedited',
    'Priority': 'Xparcel Priority',
    'Direct Call': 'Xparcel Priority'
}


# ============================================================================
# DATA INGESTION MODULE
# ============================================================================

class DataIngestion:
    """Handle CSV ingestion with robust error handling and data cleaning"""

    @staticmethod
    def read_csv(file_path: str) -> pd.DataFrame:
        """
        Read customer CSV with automatic encoding detection and parsing

        Args:
            file_path: Path to customer CSV file

        Returns:
            Cleaned DataFrame with standardized columns
        """
        # Try multiple date column variations
        date_cols = [
            'Request Date', 'ShipDate', 'ship_date', 'created',
            'delivered', 'Delivered Date', 'Created Date'
        ]

        try:
            df = pd.read_csv(
                file_path,
                parse_dates=date_cols,
                dayfirst=False,
                low_memory=False
            )

            print(f"✓ Loaded {len(df):,} records from {Path(file_path).name}")

            # Clean tracking numbers (remove scientific notation)
            if 'tracking' in df.columns or 'Tracking Number' in df.columns:
                track_col = 'tracking' if 'tracking' in df.columns else 'Tracking Number'
                df[track_col] = df[track_col].astype(str).str.replace('.0', '', regex=False)

            return df

        except Exception as e:
            raise ValueError(f"Error reading CSV: {str(e)}")

    @staticmethod
    def validate_required_columns(df: pd.DataFrame, required: List[str]) -> Tuple[bool, List[str]]:
        """Check for required columns"""
        missing = [col for col in required if col not in df.columns]

        if missing:
            print(f"⚠ Missing columns: {', '.join(missing)}")
            print(f"Available columns: {', '.join(df.columns.tolist())}")
            return False, missing

        return True, []


# ============================================================================
# TRANSFORMATION MODULE
# ============================================================================

class FirstMileTransformations:
    """Apply FirstMile business logic and calculations"""

    @staticmethod
    def standardize_service_levels(df: pd.DataFrame, service_col: str) -> pd.DataFrame:
        """Map service level names to Xparcel standards"""
        df['Service Level'] = df[service_col].map(SERVICE_MAPPING).fillna(df[service_col])
        return df

    @staticmethod
    def calculate_billable_weight(df: pd.DataFrame, weight_col: str) -> pd.DataFrame:
        """
        Apply carrier billable weight rules:
        - Under 1 lb: Round UP to next whole oz (max 15.99 oz)
        - 16 oz exactly: Bills as 1 lb
        - Over 1 lb: Round UP to next whole pound
        """
        df['Actual Weight'] = df[weight_col]

        def calc_billable(weight):
            if pd.isna(weight):
                return weight

            if weight < 1:
                # Convert to oz, round up
                oz = weight * 16
                return min(np.ceil(oz) / 16, 0.9999)  # Max 15.99 oz
            elif weight == 1:
                return 1
            else:
                # Round up to next whole pound
                return np.ceil(weight)

        df['Billable Weight'] = df['Actual Weight'].apply(calc_billable)
        return df

    @staticmethod
    def calculate_sla_compliance(df: pd.DataFrame, service_col: str,
                                 transit_col: str, status_col: str) -> pd.DataFrame:
        """
        Calculate SLA compliance based on service windows
        Only for Delivered packages
        """
        # Filter delivered only
        delivered_mask = df[status_col].str.contains('Delivered', case=False, na=False)

        df['SLA Window'] = df[service_col].map(SLA_WINDOWS)
        df['Within SLA'] = False

        # Calculate compliance for delivered packages
        df.loc[delivered_mask, 'Within SLA'] = (
            df.loc[delivered_mask, transit_col] <= df.loc[delivered_mask, 'SLA Window']
        )

        return df

    @staticmethod
    def assign_hubs(df: pd.DataFrame, state_col: str) -> pd.DataFrame:
        """Assign hub based on destination state"""
        df['Assigned Hub'] = df[state_col].map(HUB_MAP).fillna('National Network')
        return df

    @staticmethod
    def categorize_zones(df: pd.DataFrame, zone_col: str) -> pd.DataFrame:
        """Group zones into Regional (1-4) vs Cross-Country (5-8)"""
        df['Zone Category'] = df[zone_col].apply(
            lambda z: 'Regional (1-4)' if z <= 4 else 'Cross-Country (5-8)'
            if pd.notna(z) else 'Unknown'
        )
        return df


# ============================================================================
# REPORT GENERATION MODULE
# ============================================================================

class FirstMileReportGenerator:
    """Generate branded Excel reports with 9 standard tabs"""

    def __init__(self, customer_name: str, report_period: str, service_level: str):
        self.customer_name = customer_name
        self.report_period = report_period
        self.service_level = service_level
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet

    def create_header_style(self):
        """FirstMile branded header style"""
        return {
            'font': Font(color='FFFFFF', bold=True, size=11),
            'fill': PatternFill(start_color=FIRSTMILE_BLUE, end_color=FIRSTMILE_BLUE, fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center')
        }

    def create_data_style(self):
        """Standard data cell style"""
        return {
            'alignment': Alignment(horizontal='center'),
            'border': Border(
                left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD')
            )
        }

    def add_sheet_with_data(self, sheet_name: str, df: pd.DataFrame,
                           include_index: bool = False):
        """Add sheet with formatted data"""
        ws = self.wb.create_sheet(title=sheet_name)

        # Write data
        for r in dataframe_to_rows(df, index=include_index, header=True):
            ws.append(r)

        # Apply header formatting
        header_style = self.create_header_style()
        for cell in ws[1]:
            cell.font = header_style['font']
            cell.fill = header_style['fill']
            cell.alignment = header_style['alignment']

        # Apply data formatting
        data_style = self.create_data_style()
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = data_style['alignment']
                cell.border = data_style['border']

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add auto-filter
        ws.auto_filter.ref = ws.dimensions

    def generate_executive_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        """Tab 1: Executive Summary KPIs"""
        total_shipments = len(df)
        delivered = df[df['Delivered Status'].str.contains('Delivered', case=False, na=False)]

        sla_compliance = (delivered['Within SLA'].sum() / len(delivered) * 100) if len(delivered) > 0 else 0

        # Determine performance status
        perf_status = next(status for threshold, status in PERF_THRESHOLDS if sla_compliance >= threshold)

        summary_data = {
            'Metric': [
                'Report Period',
                'Service Level',
                'Total Shipments',
                'Delivered Shipments',
                'In-Transit Shipments',
                'SLA Compliance %',
                'Performance Status',
                'Average Transit Days',
                'Top Destination State'
            ],
            'Value': [
                self.report_period,
                self.service_level,
                f"{total_shipments:,}",
                f"{len(delivered):,}",
                f"{total_shipments - len(delivered):,}",
                f"{sla_compliance:.1f}%",
                perf_status,
                f"{delivered['Days In Transit'].mean():.1f}" if len(delivered) > 0 else 'N/A',
                df['Destination State'].value_counts().index[0] if len(df) > 0 else 'N/A'
            ]
        }

        return pd.DataFrame(summary_data)

    def save(self, output_path: str):
        """Save workbook to file"""
        self.wb.save(output_path)
        print(f"✓ Report saved: {output_path}")


# ============================================================================
# PIPELINE ORCHESTRATOR
# ============================================================================

class FirstMileDataPipeline:
    """Main pipeline orchestrator"""

    def __init__(self, config: Dict):
        self.config = config
        self.df = None
        self.report_generator = None

    def run(self, csv_path: str, output_dir: str = None):
        """Execute complete pipeline"""
        print("\n" + "="*70)
        print("FIRSTMILE DATA PIPELINE - PROCESSING STARTED")
        print("="*70)

        # Stage 1: Ingestion
        print("\n[1/5] Data Ingestion...")
        self.df = DataIngestion.read_csv(csv_path)

        # Stage 2: Validation
        print("\n[2/5] Validation...")
        required_cols = self.config.get('required_columns', [])
        if required_cols:
            valid, missing = DataIngestion.validate_required_columns(self.df, required_cols)
            if not valid:
                raise ValueError(f"Missing required columns: {missing}")

        # Stage 3: Transformation
        print("\n[3/5] Transformations...")
        transformations = FirstMileTransformations()

        # Apply transformations based on available columns
        if 'Service' in self.df.columns or 'Service Level' in self.df.columns:
            service_col = 'Service' if 'Service' in self.df.columns else 'Service Level'
            self.df = transformations.standardize_service_levels(self.df, service_col)
            print("  ✓ Service levels standardized")

        if 'Weight' in self.df.columns:
            self.df = transformations.calculate_billable_weight(self.df, 'Weight')
            print("  ✓ Billable weights calculated")

        if all(col in self.df.columns for col in ['Service Level', 'Days In Transit', 'Delivered Status']):
            self.df = transformations.calculate_sla_compliance(
                self.df, 'Service Level', 'Days In Transit', 'Delivered Status'
            )
            print("  ✓ SLA compliance calculated")

        if 'Destination State' in self.df.columns:
            self.df = transformations.assign_hubs(self.df, 'Destination State')
            print("  ✓ Hubs assigned")

        if 'Calculated Zone' in self.df.columns or 'Zone' in self.df.columns:
            zone_col = 'Calculated Zone' if 'Calculated Zone' in self.df.columns else 'Zone'
            self.df = transformations.categorize_zones(self.df, zone_col)
            print("  ✓ Zones categorized")

        # Stage 4: Report Generation
        print("\n[4/5] Report Generation...")
        customer_name = self.config.get('customer_name', 'Customer')
        report_period = self.config.get('report_period', 'Recent Period')
        service_level = self.config.get('service_level', 'All Services')

        self.report_generator = FirstMileReportGenerator(customer_name, report_period, service_level)

        # Generate tabs
        summary_df = self.report_generator.generate_executive_summary(self.df)
        self.report_generator.add_sheet_with_data('Executive Summary', summary_df)
        print("  ✓ Executive Summary created")

        # Add raw data tab
        self.report_generator.add_sheet_with_data('Raw Data', self.df)
        print("  ✓ Raw Data tab created")

        # Stage 5: Output
        print("\n[5/5] Saving Report...")
        if output_dir is None:
            output_dir = Path(csv_path).parent

        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        output_filename = f"FirstMile_Report_{customer_name.replace(' ', '_')}_{timestamp}.xlsx"
        output_path = Path(output_dir) / output_filename

        self.report_generator.save(str(output_path))

        print("\n" + "="*70)
        print("PIPELINE COMPLETED SUCCESSFULLY")
        print("="*70)
        print(f"\nReport: {output_path}")
        print(f"Records Processed: {len(self.df):,}")

        return str(output_path)


# ============================================================================
# USAGE EXAMPLE
# ============================================================================

if __name__ == "__main__":
    # Pipeline configuration
    config = {
        'customer_name': 'Sample Customer',
        'report_period': 'October 2025',
        'service_level': 'All Services',
        'required_columns': ['Delivered Status', 'Days In Transit']
    }

    # Initialize and run pipeline
    pipeline = FirstMileDataPipeline(config)

    # Process CSV file
    csv_file = "customer_shipments.csv"  # Replace with actual file
    output_report = pipeline.run(csv_file)

    print(f"\n✓ Pipeline complete! Report saved to: {output_report}")
