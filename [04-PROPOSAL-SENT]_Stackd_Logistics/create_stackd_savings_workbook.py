#!/usr/bin/env python
"""
STACKD LOGISTICS - CUSTOMER-FACING XPARCEL SAVINGS ANALYSIS WORKBOOK
Professional Excel deliverable for Landon meeting
Date: October 7, 2025
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

print("="*80)
print("STACKD LOGISTICS - CUSTOMER-FACING XPARCEL SAVINGS WORKBOOK")
print("Creating Professional Excel Deliverable")
print("="*80)

# Load the corrected analysis data
df = pd.read_csv('Stackd_Logistics_Xparcel_Analysis_CORRECTED.csv')

print(f"\nLoaded {len(df):,} shipments from analysis")

# Create workbook
wb = Workbook()

# Define FirstMile brand colors and styles
FIRSTMILE_BLUE = "366092"
LIGHT_GRAY = "F2F2F2"
DARK_GRAY = "666666"
GREEN = "00B050"
RED = "FF0000"

# Header style
header_style = NamedStyle(name="fm_header")
header_style.font = Font(bold=True, color="FFFFFF", size=12)
header_style.fill = PatternFill(start_color=FIRSTMILE_BLUE, end_color=FIRSTMILE_BLUE, fill_type="solid")
header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
header_style.border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# Subheader style
subheader_style = NamedStyle(name="fm_subheader")
subheader_style.font = Font(bold=True, size=11, color="FFFFFF")
subheader_style.fill = PatternFill(start_color=DARK_GRAY, end_color=DARK_GRAY, fill_type="solid")
subheader_style.alignment = Alignment(horizontal="left", vertical="center")

# Currency style
currency_style = NamedStyle(name="fm_currency")
currency_style.number_format = '$#,##0.00'
currency_style.alignment = Alignment(horizontal="right")

# Percent style
percent_style = NamedStyle(name="fm_percent")
percent_style.number_format = '0.0%'
percent_style.alignment = Alignment(horizontal="center")

# Data style
data_style = NamedStyle(name="fm_data")
data_style.alignment = Alignment(horizontal="center")

# Add styles to workbook
try:
    wb.add_named_style(header_style)
    wb.add_named_style(subheader_style)
    wb.add_named_style(currency_style)
    wb.add_named_style(percent_style)
    wb.add_named_style(data_style)
except:
    pass  # Styles may already exist

# ============================================
# TAB 1: EXECUTIVE SUMMARY
# ============================================
ws1 = wb.active
ws1.title = "Executive Summary"

# Company header
ws1.merge_cells('A1:J1')
ws1['A1'] = "STACKD LOGISTICS - FIRSTMILE XPARCEL SAVINGS ANALYSIS"
ws1['A1'].font = Font(bold=True, size=18, color="FFFFFF")
ws1['A1'].fill = PatternFill(start_color=FIRSTMILE_BLUE, end_color=FIRSTMILE_BLUE, fill_type="solid")
ws1['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 30

# Analysis info
ws1['A3'] = "Analysis Date:"
ws1['B3'] = datetime.now().strftime("%B %d, %Y")
ws1['A4'] = "Data Period:"
ws1['B4'] = "August - September 2025 (2-week sample)"
ws1['A5'] = "Total Shipments:"
ws1['B5'] = f"{len(df):,}"

for row in [3, 4, 5]:
    ws1[f'A{row}'].font = Font(bold=True)

# Financial Summary
ws1['A7'] = "SAVINGS SUMMARY"
ws1.merge_cells('A7:E7')
ws1['A7'].font = Font(bold=True, size=14, color="FFFFFF")
ws1['A7'].fill = PatternFill(start_color=FIRSTMILE_BLUE, end_color=FIRSTMILE_BLUE, fill_type="solid")
ws1['A7'].alignment = Alignment(horizontal="center")

summary_data = [
    ["Metric", "Current (DHL)", "With FirstMile Xparcel", "Monthly Savings", "% Savings"],
    ["Monthly Spend", df['Label Cost'].sum(), df['xparcel_cost'].sum(),
     df['savings'].sum(), df['savings'].sum() / df['Label Cost'].sum()],
    ["Annual Projection", df['Label Cost'].sum() * 12, df['xparcel_cost'].sum() * 12,
     df['savings'].sum() * 12, df['savings'].sum() / df['Label Cost'].sum()],
    ["Average Cost/Label", df['Label Cost'].mean(), df['xparcel_cost'].mean(),
     df['savings'].mean(), df['savings'].mean() / df['Label Cost'].mean()],
]

for row_idx, row_data in enumerate(summary_data, 9):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)

        if row_idx == 9:  # Header row
            cell.style = "fm_header"
        else:
            if col_idx in [2, 3, 4]:  # Currency columns
                cell.style = "fm_currency"
            elif col_idx == 5:  # Percent column
                cell.style = "fm_percent"

            # Highlight savings
            if col_idx == 4 and row_idx > 9:
                cell.font = Font(bold=True, color=GREEN, size=12)

# Key Metrics
ws1['G7'] = "KEY METRICS"
ws1.merge_cells('G7:J7')
ws1['G7'].font = Font(bold=True, size=14, color="FFFFFF")
ws1['G7'].fill = PatternFill(start_color=FIRSTMILE_BLUE, end_color=FIRSTMILE_BLUE, fill_type="solid")
ws1['G7'].alignment = Alignment(horizontal="center")

key_metrics = [
    ["Total Shipments", f"{len(df):,}"],
    ["Under 1 lb", f"{len(df[df['Weight (lb)'] <= 1]):,} ({len(df[df['Weight (lb)'] <= 1])/len(df)*100:.1f}%)"],
    ["Average Weight", f"{df['Weight (lb)'].mean():.2f} lbs"],
    ["Most Common Zone", f"Zone {df['Zone'].mode()[0]:.0f}"],
    ["Implementation Time", "2-3 weeks"],
    ["Contract Minimum", "None required"],
]

for idx, (label, value) in enumerate(key_metrics, 9):
    ws1.cell(row=idx, column=7, value=label).font = Font(bold=True, size=10)
    ws1.cell(row=idx, column=8, value=value).font = Font(size=10)

# Value Proposition
ws1['A15'] = "FIRSTMILE VALUE PROPOSITION"
ws1.merge_cells('A15:J15')
ws1['A15'].font = Font(bold=True, size=12, color="FFFFFF")
ws1['A15'].fill = PatternFill(start_color=DARK_GRAY, end_color=DARK_GRAY, fill_type="solid")

value_props = [
    "Cost Savings: 10.2% average savings across all shipments ($4,522/month)",
    "Lightweight Specialty: 92.5% of your volume is under 1 lb - FirstMile's sweet spot",
    "Single Integration: One ShipHero connection for Ground + Expedited + Priority",
    "Audit Queue: Blocks mis-rated labels before invoice hits your accounting",
    "3PL Focus: We understand your business model and customer margin needs",
    "Unified Support: Single team handles claims, returns, and exceptions",
]

for idx, prop in enumerate(value_props, 17):
    ws1.cell(row=idx, column=1, value=f"  {prop}").font = Font(size=10)
    ws1.merge_cells(f'A{idx}:J{idx}')

# ============================================
# TAB 2: WEIGHT ANALYSIS
# ============================================
ws2 = wb.create_sheet("Weight Analysis")

ws2.merge_cells('A1:H1')
ws2['A1'] = "SAVINGS BY WEIGHT RANGE"
ws2['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws2['A1'].fill = PatternFill(start_color=FIRSTMILE_BLUE, end_color=FIRSTMILE_BLUE, fill_type="solid")
ws2['A1'].alignment = Alignment(horizontal="center")

# Weight breakdown
weight_bins = [0, 0.25, 0.5, 1, 2, 5, 10, 100]
weight_labels = ['0-4oz', '4-8oz', '8oz-1lb', '1-2lb', '2-5lb', '5-10lb', '10lb+']
df['weight_bucket'] = pd.cut(df['Weight (lb)'], bins=weight_bins, labels=weight_labels)

weight_headers = ["Weight Range", "Shipments", "% of Total", "Avg DHL Cost",
                  "Avg Xparcel Cost", "Avg Savings", "Savings %", "Total Savings"]

for col_idx, header in enumerate(weight_headers, 1):
    cell = ws2.cell(row=3, column=col_idx, value=header)
    cell.style = "fm_header"

row_idx = 4
for bucket in weight_labels:
    bucket_df = df[df['weight_bucket'] == bucket]
    if len(bucket_df) > 0:
        ws2.cell(row=row_idx, column=1, value=bucket).style = "fm_data"
        ws2.cell(row=row_idx, column=2, value=len(bucket_df)).style = "fm_data"
        ws2.cell(row=row_idx, column=3, value=len(bucket_df)/len(df)).style = "fm_percent"
        ws2.cell(row=row_idx, column=4, value=bucket_df['Label Cost'].mean()).style = "fm_currency"
        ws2.cell(row=row_idx, column=5, value=bucket_df['xparcel_cost'].mean()).style = "fm_currency"
        ws2.cell(row=row_idx, column=6, value=bucket_df['savings'].mean()).style = "fm_currency"

        savings_pct = bucket_df['savings'].mean() / bucket_df['Label Cost'].mean()
        pct_cell = ws2.cell(row=row_idx, column=7, value=savings_pct)
        pct_cell.style = "fm_percent"

        # Color code savings percentage
        if savings_pct >= 0.15:
            pct_cell.font = Font(bold=True, color=GREEN)
        elif savings_pct >= 0.10:
            pct_cell.font = Font(bold=True, color="0066CC")
        elif savings_pct < 0:
            pct_cell.font = Font(bold=True, color=RED)

        ws2.cell(row=row_idx, column=8, value=bucket_df['savings'].sum()).style = "fm_currency"
        row_idx += 1

# Add totals row
ws2.cell(row=row_idx, column=1, value="TOTAL").font = Font(bold=True)
ws2.cell(row=row_idx, column=2, value=len(df)).font = Font(bold=True)
ws2.cell(row=row_idx, column=3, value=1.0).style = "fm_percent"
ws2.cell(row=row_idx, column=4, value=df['Label Cost'].mean()).style = "fm_currency"
ws2.cell(row=row_idx, column=5, value=df['xparcel_cost'].mean()).style = "fm_currency"
ws2.cell(row=row_idx, column=6, value=df['savings'].mean()).style = "fm_currency"
ws2.cell(row=row_idx, column=7, value=df['savings'].mean() / df['Label Cost'].mean()).style = "fm_percent"
ws2.cell(row=row_idx, column=8, value=df['savings'].sum()).style = "fm_currency"

# Key insights
ws2[f'A{row_idx+3}'] = "KEY INSIGHTS:"
ws2[f'A{row_idx+3}'].font = Font(bold=True, size=12)

insights = [
    f"92.5% of your shipments are under 1 lb - FirstMile's lightweight specialty",
    f"Best savings in 4-8 oz range: 27.2% of volume at 13.8% average savings",
    f"Heavier packages (1-5 lbs) save 19%+ despite being only 7% of volume",
    f"Only 6 packages over 10 lbs show slight cost increase (can keep with DHL)",
]

for idx, insight in enumerate(insights, row_idx+4):
    ws2.cell(row=idx, column=1, value=f"  {insight}").font = Font(size=10)
    ws2.merge_cells(f'A{idx}:H{idx}')

# ============================================
# TAB 3: ZONE ANALYSIS
# ============================================
ws3 = wb.create_sheet("Zone Analysis")

ws3.merge_cells('A1:H1')
ws3['A1'] = "SAVINGS BY DESTINATION ZONE"
ws3['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws3['A1'].fill = PatternFill(start_color=FIRSTMILE_BLUE, end_color=FIRSTMILE_BLUE, fill_type="solid")
ws3['A1'].alignment = Alignment(horizontal="center")

zone_headers = ["Zone", "Shipments", "% of Total", "Avg DHL Cost",
                "Avg Xparcel Cost", "Avg Savings/Pkg", "Total Savings"]

for col_idx, header in enumerate(zone_headers, 1):
    cell = ws3.cell(row=3, column=col_idx, value=header)
    cell.style = "fm_header"

row_idx = 4
for zone in range(1, 9):
    zone_df = df[df['Zone'] == zone]
    if len(zone_df) > 0:
        ws3.cell(row=row_idx, column=1, value=f"Zone {zone}").style = "fm_data"
        ws3.cell(row=row_idx, column=2, value=len(zone_df)).style = "fm_data"
        ws3.cell(row=row_idx, column=3, value=len(zone_df)/len(df)).style = "fm_percent"
        ws3.cell(row=row_idx, column=4, value=zone_df['Label Cost'].mean()).style = "fm_currency"
        ws3.cell(row=row_idx, column=5, value=zone_df['xparcel_cost'].mean()).style = "fm_currency"
        ws3.cell(row=row_idx, column=6, value=zone_df['savings'].mean()).style = "fm_currency"
        ws3.cell(row=row_idx, column=7, value=zone_df['savings'].sum()).style = "fm_currency"
        row_idx += 1

# ============================================
# TAB 4: RATE COMPARISON
# ============================================
ws4 = wb.create_sheet("Rate Comparison")

ws4.merge_cells('A1:H1')
ws4['A1'] = "DHL vs FIRSTMILE XPARCEL RATE COMPARISON"
ws4['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws4['A1'].fill = PatternFill(start_color=FIRSTMILE_BLUE, end_color=FIRSTMILE_BLUE, fill_type="solid")
ws4['A1'].alignment = Alignment(horizontal="center")

ws4['A3'] = "LIGHTWEIGHT PACKAGES (Where 92.5% of Your Volume Sits)"
ws4['A3'].font = Font(bold=True, size=12)
ws4.merge_cells('A3:H3')

comparison_headers = ["Weight", "Zone", "Avg DHL Rate", "Xparcel Ground Rate",
                     "Savings $", "Savings %", "Sample Count"]

for col_idx, header in enumerate(comparison_headers, 1):
    cell = ws4.cell(row=5, column=col_idx, value=header)
    cell.style = "fm_header"

# Sample rates by weight and zone
row_idx = 6
sample_weights = [(0.0625, '1 oz'), (0.25, '4 oz'), (0.5, '8 oz'), (1, '1 lb'), (2, '2 lb'), (5, '5 lb')]
sample_zones = [3, 4, 5, 6, 7]

for weight, weight_label in sample_weights:
    for zone in sample_zones:
        # Get samples from actual data
        samples = df[(df['Weight (lb)'] >= weight-0.1) &
                    (df['Weight (lb)'] <= weight+0.1) &
                    (df['Zone'] == zone)]

        if len(samples) > 0:
            dhl_avg = samples['Label Cost'].mean()
            xparcel_avg = samples['xparcel_cost'].mean()
            savings_amt = samples['savings'].mean()
            savings_pct = savings_amt / dhl_avg if dhl_avg > 0 else 0

            ws4.cell(row=row_idx, column=1, value=weight_label).style = "fm_data"
            ws4.cell(row=row_idx, column=2, value=f"Zone {zone}").style = "fm_data"
            ws4.cell(row=row_idx, column=3, value=dhl_avg).style = "fm_currency"
            ws4.cell(row=row_idx, column=4, value=xparcel_avg).style = "fm_currency"
            ws4.cell(row=row_idx, column=5, value=savings_amt).style = "fm_currency"

            pct_cell = ws4.cell(row=row_idx, column=6, value=savings_pct)
            pct_cell.style = "fm_percent"
            if savings_pct >= 0.15:
                pct_cell.font = Font(bold=True, color=GREEN)
            elif savings_pct >= 0.10:
                pct_cell.font = Font(bold=True, color="0066CC")

            ws4.cell(row=row_idx, column=7, value=len(samples)).style = "fm_data"
            row_idx += 1

# ============================================
# TAB 5: IMPLEMENTATION PLAN
# ============================================
ws5 = wb.create_sheet("Implementation Plan")

ws5.merge_cells('A1:G1')
ws5['A1'] = "FIRSTMILE IMPLEMENTATION ROADMAP"
ws5['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws5['A1'].fill = PatternFill(start_color=FIRSTMILE_BLUE, end_color=FIRSTMILE_BLUE, fill_type="solid")
ws5['A1'].alignment = Alignment(horizontal="center")

# Timeline
ws5['A3'] = "IMPLEMENTATION TIMELINE (2-3 Weeks)"
ws5['A3'].font = Font(bold=True, size=12)
ws5.merge_cells('A3:G3')

timeline_data = [
    ["Phase", "Activities", "Duration", "Owner"],
    ["1. Kickoff", "Contract execution, account setup, credentials", "1-2 days", "FirstMile + Stackd"],
    ["2. Integration", "ShipHero API configuration, label testing", "3-5 days", "FirstMile Tech + Stackd"],
    ["3. Testing", "Test labels for all service levels, rate validation", "2-3 days", "Stackd + FirstMile"],
    ["4. Parallel Run", "Run FirstMile + DHL in parallel to validate", "3-5 days", "Stackd"],
    ["5. Full Cutover", "Switch all volume to FirstMile, monitor closely", "1 week", "Stackd + FirstMile"],
]

for row_idx, row_data in enumerate(timeline_data, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws5.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 5:
            cell.style = "fm_header"
        else:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

# Requirements
ws5['A13'] = "REQUIREMENTS FROM STACKD"
ws5['A13'].font = Font(bold=True, size=12)
ws5.merge_cells('A13:G13')

requirements = [
    "ShipHero account credentials (API access)",
    "Current DHL rate card (for final validation)",
    "3PL customer list (Chin Mounts, New Age Performance, Tapeher)",
    "Service level SLA requirements from your customers",
    "Go-live target date preference",
]

for idx, req in enumerate(requirements, 15):
    ws5.cell(row=idx, column=1, value=f"  {req}").font = Font(size=10)
    ws5.merge_cells(f'A{idx}:G{idx}')

# Support
ws5['A22'] = "FIRSTMILE SUPPORT"
ws5['A22'].font = Font(bold=True, size=12)
ws5.merge_cells('A22:G22')

support_info = [
    "Account Manager: Brett Walker | brett.walker@firstmile.com | 402-718-4727",
    "Technical Support: Available 24/7 for integration and troubleshooting",
    "Customer Success: Dedicated 3PL team for ongoing optimization",
    "Claims & Returns: Single-point-of-contact for all exception handling",
]

for idx, info in enumerate(support_info, 24):
    ws5.cell(row=idx, column=1, value=f"  {info}").font = Font(size=10)
    ws5.merge_cells(f'A{idx}:G{idx}')

# ============================================
# TAB 6: FIRSTMILE ADVANTAGES
# ============================================
ws6 = wb.create_sheet("FirstMile Advantages")

ws6.merge_cells('A1:H1')
ws6['A1'] = "WHY FIRSTMILE FOR 3PL PROVIDERS"
ws6['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws6['A1'].fill = PatternFill(start_color=FIRSTMILE_BLUE, end_color=FIRSTMILE_BLUE, fill_type="solid")
ws6['A1'].alignment = Alignment(horizontal="center")

advantages = [
    ("Cost Optimization", [
        "10.2% average savings on your current volume",
        "No hidden fees or surcharges",
        "Transparent pricing with rate card guarantee",
        "Volume discounts available as you scale",
    ]),
    ("Operational Excellence", [
        "Audit Queue: Blocks mis-rated labels before invoicing",
        "Single integration: One ShipHero connection for all services",
        "Unified platform: Ground, Expedited, Priority from one provider",
        "Real-time tracking and visibility",
    ]),
    ("3PL-Focused Support", [
        "We understand your multi-customer business model",
        "Dedicated 3PL customer success team",
        "Custom reporting for your customer base",
        "Margin protection strategies and pricing guidance",
    ]),
    ("Risk Mitigation", [
        "No long-term contracts or commitments",
        "Parallel run option to validate performance",
        "Easy integration with ShipHero (your current platform)",
        "Keep DHL for exceptions if needed",
    ]),
]

row_idx = 3
for category, items in advantages:
    ws6.cell(row=row_idx, column=1, value=category).font = Font(bold=True, size=12, color="FFFFFF")
    ws6.cell(row=row_idx, column=1).fill = PatternFill(start_color=DARK_GRAY, end_color=DARK_GRAY, fill_type="solid")
    ws6.merge_cells(f'A{row_idx}:H{row_idx}')
    row_idx += 1

    for item in items:
        ws6.cell(row=row_idx, column=1, value=f"  {item}").font = Font(size=10)
        ws6.merge_cells(f'A{row_idx}:H{row_idx}')
        row_idx += 1

    row_idx += 1

# ============================================
# AUTO-ADJUST COLUMN WIDTHS
# ============================================
for sheet in wb.worksheets:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width

# ============================================
# SAVE WORKBOOK
# ============================================
timestamp = datetime.now().strftime("%Y%m%d_%H%M")
output_file = f'Stackd_Logistics_FirstMile_Xparcel_Savings_Analysis_{timestamp}.xlsx'
wb.save(output_file)

print(f"\n{'='*80}")
print(f"WORKBOOK CREATED: {output_file}")
print(f"{'='*80}")
print(f"\nWorkbook Contents:")
print(f"  1. Executive Summary - Overall savings and key metrics")
print(f"  2. Weight Analysis - Savings breakdown by weight range")
print(f"  3. Zone Analysis - Savings breakdown by destination zone")
print(f"  4. Rate Comparison - DHL vs FirstMile rate examples")
print(f"  5. Implementation Plan - 2-3 week roadmap with timeline")
print(f"  6. FirstMile Advantages - Value proposition for 3PL providers")
print(f"\nKey Numbers:")
print(f"  - Monthly Savings: ${df['savings'].sum():,.2f}")
print(f"  - Annual Savings: ${df['savings'].sum() * 12:,.2f}")
print(f"  - Savings Percentage: {df['savings'].sum() / df['Label Cost'].sum() * 100:.1f}%")
print(f"  - Total Shipments Analyzed: {len(df):,}")
print(f"\n{'='*80}")
print(f"READY FOR LANDON MEETING")
print(f"{'='*80}")
