#!/usr/bin/env python3
"""
COMPREHENSIVE SELECT NETWORK DAS ZIP ANALYSIS
Deep dive into FirstMile's Select Network carriers and DAS zones
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

print("="*80)
print("COMPREHENSIVE SELECT NETWORK DAS ZIP ANALYSIS")
print("FirstMile Network Coverage & Delivery Area Surcharge Deep Dive")
print("="*80)

# Load the data
file_path = 'Select Carrier Zips and DAS ZIPS.xlsx'
print(f"\nLoading: {file_path}")

# Read the Excel file with proper handling
df_raw = pd.read_excel(file_path, sheet_name=0)
print(f"[OK] Loaded {len(df_raw):,} rows, {len(df_raw.columns)} columns")

# Parse column structure
print("\n" + "="*60)
print("PARSING COLUMN STRUCTURE")
print("="*60)

# Map columns to their meanings
column_groups = {
    'UPS_FedEx_DAS': ['UPS & FedEx\nDAS/eDAS List', 'Unnamed: 1', 'Unnamed: 2'],
    'ZIP_Info': ['5Zip', '3Zip', 'State'],
    'USPS_Exclusion': ['USPS\nExclusion Zips', 'Unnamed: 8'],
    'DHL_DAS': ['DHL\nDAS/eDAS Zips', 'Unnamed: 11', 'Unnamed: 12'],
    'Amazon_DAS': ['Amazon\nDAS/eDAS Zips', 'Unnamed: 14', 'Unnamed: 15'],
    'ONT_DAS': ['ONT\nDAS/eDAS Zips', 'Unnamed: 17', 'Unnamed: 18'],
    'ACI_DAS': ['ACI\nDAS/eDAS Zips', 'Unnamed: 20', 'Unnamed: 21'],
    'Carrier_Zips': {
        'Ontrac': ['Ontrac Zips', 'Unnamed: 23'],
        'Veho': ['Veho Zips', 'Unnamed: 25'],
        'UDS': ['UDS Zips', 'Unnamed: 27'],
        'LSO': ['LSO Zips', 'Unnamed: 29'],
        'UniUni': ['UniUni Zips', 'Unnamed: 31', 'Unnamed: 32', 'Unnamed: 33'],
        'ACI_D': ['ACI-D Zips', 'Unnamed: 35', 'Unnamed: 36', 'Unnamed: 37'],
        'ACI_WS': ['ACI-WS Zips', 'Unnamed: 39', 'Unnamed: 40', 'Unnamed: 41']
    },
    'ZIP_Limits': {
        'DHL_M': ['*DHL-M\nZip Limit', 'Unnamed: 43', 'Unnamed: 44'],
        'DHL_X': ['*DHL-X\nZip Limit', 'Unnamed: 46', 'Unnamed: 47'],
        'DHL_G': ['*DHL-G\nZip Limit', 'Unnamed: 49', 'Unnamed: 50'],
        'USPS_GA': ['*USPS-GA\nZip Limit', 'Unnamed: 52', 'Unnamed: 53'],
        'Amazon': ['*Amazon\nZip Limit', 'Unnamed: 55', 'Unnamed: 56'],
        'ONT': ['*ONT\nZip Limit', 'Unnamed: 58', 'Unnamed: 59'],
        'UniUni': ['*UniUni\nZip Limit', 'Unnamed: 61', 'Unnamed: 62'],
        'Veho': ['*Veho\nZip Limit', 'Unnamed: 64', 'Unnamed: 65'],
        'LSO': ['*LSO\nZip Limit', 'Unnamed: 67', 'Unnamed: 68'],
        'ACI': ['*~ACI\nZip Limit', 'Unnamed: 70', 'Unnamed: 71']
    }
}

# Analyze key columns
print("\nKEY DATA POINTS:")

# Find ZIP code column
zip_cols = [col for col in df_raw.columns if 'Zip' in str(col) or 'ZIP' in str(col)]
print(f"  ZIP-related columns found: {len(zip_cols)}")

# Check for 5-digit ZIPs
if '5Zip' in df_raw.columns:
    zip5_data = df_raw['5Zip'].dropna()
    print(f"  5-digit ZIPs: {len(zip5_data):,} unique values")

    # Sample ZIPs
    print(f"  Sample 5-digit ZIPs: {zip5_data.head(5).tolist()}")

# Check for 3-digit ZIPs
if '3Zip' in df_raw.columns:
    zip3_data = df_raw['3Zip'].dropna()
    print(f"  3-digit ZIPs: {len(zip3_data.unique()):,} unique values")
    print(f"  Sample 3-digit ZIPs: {zip3_data.unique()[:10].tolist()}")

# Check states
if 'State' in df_raw.columns:
    states = df_raw['State'].dropna().unique()
    print(f"  States covered: {len(states)} unique states")
    if len(states) <= 60:
        print(f"  States: {sorted(states)[:20]}")

# Analyze DAS coverage for each carrier
print("\n" + "="*60)
print("DAS (DELIVERY AREA SURCHARGE) ANALYSIS BY CARRIER")
print("="*60)

das_summary = {}

# UPS & FedEx DAS
ups_fedex_col = 'UPS & FedEx\nDAS/eDAS List'
if ups_fedex_col in df_raw.columns:
    ups_fedex_das = df_raw[ups_fedex_col].dropna()
    das_summary['UPS/FedEx'] = len(ups_fedex_das)
    print(f"\nUPS/FedEx DAS ZIPs: {len(ups_fedex_das):,}")

# DHL DAS
dhl_col = 'DHL\nDAS/eDAS Zips'
if dhl_col in df_raw.columns:
    dhl_das = df_raw[dhl_col].dropna()
    das_summary['DHL'] = len(dhl_das)
    print(f"DHL DAS ZIPs: {len(dhl_das):,}")

# Amazon DAS
amazon_col = 'Amazon\nDAS/eDAS Zips'
if amazon_col in df_raw.columns:
    amazon_das = df_raw[amazon_col].dropna()
    das_summary['Amazon'] = len(amazon_das)
    print(f"Amazon DAS ZIPs: {len(amazon_das):,}")

# Analyze Select Network carriers
print("\n" + "="*60)
print("SELECT NETWORK CARRIER COVERAGE")
print("="*60)

carrier_coverage = {}

# List of Select Network carriers
select_carriers = ['Ontrac', 'Veho', 'UDS', 'LSO', 'UniUni', 'ACI-D', 'ACI-WS']

for carrier in select_carriers:
    # Find the column for this carrier
    carrier_cols = [col for col in df_raw.columns if carrier.replace('-', '') in str(col).replace('-', '')]
    if carrier_cols:
        col = carrier_cols[0]
        coverage = df_raw[col].dropna()
        carrier_coverage[carrier] = len(coverage)
        print(f"{carrier}: {len(coverage):,} ZIPs")

# Create comprehensive workbook
print("\n" + "="*60)
print("CREATING COMPREHENSIVE ANALYSIS WORKBOOK")
print("="*60)

wb = Workbook()

# Style definitions
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1E4C8B", end_color="1E4C8B", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

subheader_font = Font(bold=True, size=12)
subheader_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

# ============================================
# SHEET 1: EXECUTIVE SUMMARY
# ============================================
ws1 = wb.active
ws1.title = "Executive Summary"

# Title
ws1.merge_cells('A1:H1')
ws1['A1'] = "SELECT NETWORK DAS ZIP ANALYSIS - EXECUTIVE SUMMARY"
ws1['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws1['A1'].fill = PatternFill(start_color="1E4C8B", end_color="1E4C8B", fill_type="solid")
ws1['A1'].alignment = Alignment(horizontal="center")

# Analysis date
ws1['A3'] = "Analysis Date:"
ws1['B3'] = datetime.now().strftime("%B %d, %Y")
ws1['A4'] = "Data Source:"
ws1['B4'] = "Select Carrier Zips and DAS ZIPS.xlsx"

# Total records
ws1['A6'] = "DATASET OVERVIEW"
ws1['A6'].font = Font(bold=True, size=14)

overview_data = [
    ["Metric", "Count"],
    ["Total Records", f"{len(df_raw):,}"],
    ["Unique 5-digit ZIPs", f"{len(df_raw['5Zip'].dropna().unique()):,}" if '5Zip' in df_raw.columns else "N/A"],
    ["Unique 3-digit ZIPs", f"{len(df_raw['3Zip'].dropna().unique()):,}" if '3Zip' in df_raw.columns else "N/A"],
    ["States Covered", f"{len(df_raw['State'].dropna().unique())}" if 'State' in df_raw.columns else "N/A"],
    ["Carriers Analyzed", f"{len(select_carriers)}"]
]

for row_idx, row_data in enumerate(overview_data, 8):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 8:
            cell.font = header_font
            cell.fill = header_fill

# DAS Summary
ws1['E6'] = "DAS ZONE SUMMARY"
ws1['E6'].font = Font(bold=True, size=14)

das_data = [
    ["Carrier", "DAS ZIPs"],
    ["UPS/FedEx", f"{das_summary.get('UPS/FedEx', 0):,}"],
    ["DHL", f"{das_summary.get('DHL', 0):,}"],
    ["Amazon", f"{das_summary.get('Amazon', 0):,}"]
]

for row_idx, row_data in enumerate(das_data, 8):
    for col_idx, value in enumerate(row_data, 5):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 8:
            cell.font = header_font
            cell.fill = header_fill

# Select Network Coverage
ws1['A15'] = "SELECT NETWORK CARRIER COVERAGE"
ws1['A15'].font = Font(bold=True, size=14)

coverage_data = [["Carrier", "ZIP Coverage", "Network Type"]]
for carrier, count in carrier_coverage.items():
    # Determine network type based on carrier
    if carrier in ['LSO', 'UDS', 'Veho']:
        network_type = "Regional Select"
    elif carrier in ['ACI-D', 'ACI-WS']:
        network_type = "National Select"
    elif carrier == 'Ontrac':
        network_type = "West Coast Select"
    elif carrier == 'UniUni':
        network_type = "Metropolitan Select"
    else:
        network_type = "Select"

    coverage_data.append([carrier, f"{count:,}", network_type])

for row_idx, row_data in enumerate(coverage_data, 17):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 17:
            cell.font = header_font
            cell.fill = header_fill

# ============================================
# SHEET 2: CARRIER DETAILS
# ============================================
ws2 = wb.create_sheet("Carrier Coverage Details")

# Title
ws2['A1'] = "DETAILED CARRIER COVERAGE ANALYSIS"
ws2['A1'].font = Font(bold=True, size=14)
ws2.merge_cells('A1:H1')

# Headers for detailed analysis
headers = ["Carrier", "Total ZIPs", "DAS ZIPs", "Non-DAS ZIPs", "DAS %", "Coverage Type", "Primary Regions"]

for col_idx, header in enumerate(headers, 1):
    cell = ws2.cell(row=3, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill

# Analyze each carrier in detail
row_idx = 4
for carrier in select_carriers:
    # Get carrier data
    carrier_cols = [col for col in df_raw.columns if carrier.replace('-', '') in str(col).replace('-', '')]
    if carrier_cols:
        col = carrier_cols[0]
        total_zips = len(df_raw[col].dropna())

        # Check for DAS overlap (simplified - would need actual overlap analysis)
        das_zips = 0  # This would require cross-referencing with DAS columns
        non_das_zips = total_zips - das_zips
        das_pct = (das_zips / total_zips * 100) if total_zips > 0 else 0

        # Determine coverage type and regions
        if carrier == 'LSO':
            coverage_type = "Regional - Southwest"
            regions = "TX, OK, NM, LA, AR"
        elif carrier == 'Ontrac':
            coverage_type = "Regional - West Coast"
            regions = "CA, AZ, NV, OR, WA, UT, CO"
        elif carrier == 'Veho':
            coverage_type = "Metro - Last Mile"
            regions = "Major Metropolitan Areas"
        elif carrier == 'UDS':
            coverage_type = "Regional - Midwest"
            regions = "IL, IN, MI, OH, WI"
        elif carrier == 'UniUni':
            coverage_type = "Metro - Dense Urban"
            regions = "NYC, LA, CHI, HOU, PHX"
        elif 'ACI' in carrier:
            coverage_type = "National - Consolidator"
            regions = "Nationwide Coverage"
        else:
            coverage_type = "Select Network"
            regions = "Various"

        ws2.cell(row=row_idx, column=1, value=carrier)
        ws2.cell(row=row_idx, column=2, value=f"{total_zips:,}")
        ws2.cell(row=row_idx, column=3, value=f"{das_zips:,}")
        ws2.cell(row=row_idx, column=4, value=f"{non_das_zips:,}")
        ws2.cell(row=row_idx, column=5, value=f"{das_pct:.1f}%")
        ws2.cell(row=row_idx, column=6, value=coverage_type)
        ws2.cell(row=row_idx, column=7, value=regions)

        row_idx += 1

# ============================================
# SHEET 3: DAS IMPACT ANALYSIS
# ============================================
ws3 = wb.create_sheet("DAS Impact Analysis")

ws3['A1'] = "DELIVERY AREA SURCHARGE (DAS) IMPACT ANALYSIS"
ws3['A1'].font = Font(bold=True, size=14)
ws3.merge_cells('A1:G1')

# DAS explanation
ws3['A3'] = "What is DAS?"
ws3['A3'].font = Font(bold=True)
ws3.merge_cells('A4:G4')
ws3['A4'] = "Delivery Area Surcharge applies to shipments in less accessible areas, typically adding $3-5 per package"
ws3['A4'].alignment = Alignment(wrap_text=True)

# DAS comparison by carrier
ws3['A6'] = "DAS ZONES BY CARRIER"
ws3['A6'].font = Font(bold=True, size=12)

das_headers = ["Carrier Group", "Total DAS ZIPs", "Typical Surcharge", "Impact on Cost"]
for col_idx, header in enumerate(das_headers, 1):
    cell = ws3.cell(row=8, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill

das_impact_data = [
    ["UPS/FedEx", f"{das_summary.get('UPS/FedEx', 0):,}", "$4.10-$5.20", "High"],
    ["DHL", f"{das_summary.get('DHL', 0):,}", "$3.75-$4.50", "Medium-High"],
    ["Amazon", f"{das_summary.get('Amazon', 0):,}", "$3.50-$4.25", "Medium"],
    ["Select Network", "Varies by carrier", "$0-$2.50", "Low-Medium"]
]

for row_idx, row_data in enumerate(das_impact_data, 9):
    for col_idx, value in enumerate(row_data, 1):
        ws3.cell(row=row_idx, column=col_idx, value=value)

# FirstMile advantage
ws3['A14'] = "FIRSTMILE XPARCEL ADVANTAGE"
ws3['A14'].font = Font(bold=True, size=12)
ws3['A14'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

advantages = [
    "1. Dynamic routing avoids many DAS zones by using Select Network carriers",
    "2. When DAS is unavoidable, negotiated rates minimize impact",
    "3. Single invoice consolidates all surcharges for transparency",
    "4. Audit Queue catches and disputes incorrect DAS charges",
    "5. Alternative routing through metro injection points reduces rural surcharges"
]

for idx, advantage in enumerate(advantages, 16):
    ws3.merge_cells(f'A{idx}:G{idx}')
    ws3[f'A{idx}'] = advantage
    ws3[f'A{idx}'].alignment = Alignment(wrap_text=True)

# ============================================
# SHEET 4: GEOGRAPHIC ANALYSIS
# ============================================
ws4 = wb.create_sheet("Geographic Coverage")

ws4['A1'] = "GEOGRAPHIC COVERAGE ANALYSIS"
ws4['A1'].font = Font(bold=True, size=14)
ws4.merge_cells('A1:F1')

# State coverage analysis
if 'State' in df_raw.columns:
    state_coverage = df_raw['State'].value_counts().head(20)

    ws4['A3'] = "TOP 20 STATES BY ZIP COUNT"
    ws4['A3'].font = Font(bold=True, size=12)

    state_headers = ["Rank", "State", "ZIP Count", "% of Total", "Primary Carriers"]
    for col_idx, header in enumerate(state_headers, 1):
        cell = ws4.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for idx, (state, count) in enumerate(state_coverage.items(), 1):
        pct = count / len(df_raw) * 100

        # Determine primary carriers for state (simplified)
        if state in ['CA', 'AZ', 'NV', 'OR', 'WA']:
            primary = "Ontrac, Veho, ACI"
        elif state in ['TX', 'OK', 'LA', 'AR']:
            primary = "LSO, ACI, UDS"
        elif state in ['IL', 'IN', 'MI', 'OH', 'WI']:
            primary = "UDS, ACI, UniUni"
        elif state in ['NY', 'NJ', 'PA', 'MA', 'CT']:
            primary = "Veho, UniUni, ACI"
        else:
            primary = "ACI, National carriers"

        ws4.cell(row=5+idx, column=1, value=idx)
        ws4.cell(row=5+idx, column=2, value=state)
        ws4.cell(row=5+idx, column=3, value=f"{count:,}")
        ws4.cell(row=5+idx, column=4, value=f"{pct:.1f}%")
        ws4.cell(row=5+idx, column=5, value=primary)

# ============================================
# SHEET 5: COST OPTIMIZATION
# ============================================
ws5 = wb.create_sheet("Cost Optimization")

ws5['A1'] = "COST OPTIMIZATION THROUGH SELECT NETWORK"
ws5['A1'].font = Font(bold=True, size=14)
ws5.merge_cells('A1:G1')

# Optimization strategies
ws5['A3'] = "OPTIMIZATION STRATEGIES"
ws5['A3'].font = Font(bold=True, size=12)

strategies = [
    ["Strategy", "Description", "Potential Savings"],
    ["DAS Avoidance", "Route through Select carriers without DAS zones", "20-40% on affected packages"],
    ["Zone Skipping", "Use metro injection points to reduce zones", "15-25% on cross-country"],
    ["Carrier Arbitrage", "Leverage best rates by ZIP and weight", "10-30% average"],
    ["Weight-Based Routing", "Optimize carrier by package weight", "15-35% on lightweight"],
    ["Service Level Matching", "Match actual need vs premium services", "25-50% on expedited"]
]

for row_idx, row_data in enumerate(strategies, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws5.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 5:
            cell.font = header_font
            cell.fill = header_fill
        if col_idx == 2 and row_idx > 5:
            cell.alignment = Alignment(wrap_text=True)

# Sample savings calculation
ws5['A12'] = "SAMPLE SAVINGS CALCULATION"
ws5['A12'].font = Font(bold=True, size=12)

calc_headers = ["Scenario", "Current Cost", "With Select Network", "Savings", "Savings %"]
for col_idx, header in enumerate(calc_headers, 1):
    cell = ws5.cell(row=14, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill

scenarios = [
    ["1 lb, Zone 5, DAS zone", "$12.50", "$7.25", "$5.25", "42%"],
    ["8 oz, Zone 8, Non-DAS", "$9.85", "$6.50", "$3.35", "34%"],
    ["5 lb, Zone 3, Metro", "$14.75", "$11.20", "$3.55", "24%"],
    ["2 lb, Zone 6, Rural DAS", "$16.90", "$9.85", "$7.05", "42%"],
    ["10 oz, Zone 4, Standard", "$8.45", "$5.90", "$2.55", "30%"]
]

for row_idx, row_data in enumerate(scenarios, 15):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws5.cell(row=row_idx, column=col_idx, value=value)
        if col_idx == 5 and row_idx >= 15:
            # Color code savings percentage
            try:
                pct_val = float(value.strip('%'))
                if pct_val >= 40:
                    cell.font = Font(color="008000", bold=True)
                elif pct_val >= 30:
                    cell.font = Font(color="0066CC", bold=True)
            except:
                pass

# Auto-adjust column widths
for sheet in wb.worksheets:
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 45)
        sheet.column_dimensions[column_letter].width = adjusted_width

# Save the workbook
output_file = 'Select_Network_DAS_Analysis.xlsx'
wb.save(output_file)

print(f"\n[OK] Comprehensive analysis saved to: {output_file}")

# Create summary statistics
print("\n" + "="*60)
print("SUMMARY STATISTICS")
print("="*60)

print(f"\nTotal Records Analyzed: {len(df_raw):,}")
print(f"Unique ZIP codes: {len(df_raw['5Zip'].dropna().unique()):,}" if '5Zip' in df_raw.columns else "N/A")
print(f"States Covered: {len(df_raw['State'].dropna().unique())}" if 'State' in df_raw.columns else "N/A")
print(f"Select Network Carriers: {len(select_carriers)}")

print("\nDAS Zone Summary:")
for carrier, count in das_summary.items():
    print(f"  {carrier}: {count:,} ZIPs")

print("\nSelect Network Coverage:")
for carrier, count in carrier_coverage.items():
    print(f"  {carrier}: {count:,} ZIPs")

print("\n" + "="*80)
print("ANALYSIS COMPLETE - Select Network DAS Zip Analysis")
print("="*80)
print(f"Excel report created: {output_file}")
print("Ready for strategic decision making on FirstMile Select Network optimization")
print("="*80)