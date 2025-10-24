#!/usr/bin/env python3
"""
INVOICE AUDIT BUILDER v3.1 - STACKD LOGISTICS COMPLETE ANALYSIS
All tabs, full detail, corrected Xparcel rates
Analysis Date: 2025-09-24
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings('ignore')

# Configuration
COMPANY_NAME = "Stackd Logistics"
INPUT_FILE = "20250918193042_221aaf59f30469602caf8f7f7485b114.csv"
ANALYSIS_DATE = datetime.now().strftime("%B %d, %Y")

# FirstMile Xparcel Rate Structure
XPARCEL_GROUND_RATES = {
    1: 3.73, 2: 3.79, 3: 3.80, 4: 3.89,
    5: 3.94, 6: 4.02, 7: 4.09, 8: 4.24
}

XPARCEL_EXPEDITED_RATES = {
    1: 3.94, 2: 3.99, 3: 4.01, 4: 4.10,
    5: 4.15, 6: 4.24, 7: 4.31, 8: 4.48
}

XPARCEL_PRIORITY_RATES = {
    1: 4.15, 2: 4.21, 3: 4.25, 4: 4.35,
    5: 4.42, 6: 4.51, 7: 4.59, 8: 4.75
}

# Zone mapping by state (simplified)
STATE_TO_ZONE = {
    'UT': 1, 'ID': 2, 'NV': 2, 'AZ': 3, 'CO': 3, 'WY': 3,
    'CA': 4, 'OR': 4, 'WA': 4, 'MT': 4, 'NM': 4,
    'TX': 5, 'OK': 5, 'KS': 5, 'NE': 5, 'SD': 5, 'ND': 5,
    'MN': 5, 'IA': 5, 'MO': 5, 'AR': 5, 'LA': 5,
    'WI': 6, 'IL': 6, 'IN': 6, 'MI': 6, 'OH': 6, 'KY': 6,
    'TN': 6, 'MS': 6, 'AL': 6, 'GA': 6, 'FL': 6,
    'WV': 7, 'VA': 7, 'NC': 7, 'SC': 7, 'PA': 7, 'MD': 7,
    'DE': 7, 'NJ': 7, 'NY': 7, 'CT': 7, 'RI': 7, 'MA': 7,
    'VT': 7, 'NH': 7, 'ME': 7, 'DC': 7,
    'AK': 8, 'HI': 8, 'PR': 8
}

def calculate_xparcel_cost(weight, zone, service_type='ground'):
    """Calculate FirstMile Xparcel cost based on weight, zone, and service"""

    # Select appropriate rate table
    if service_type == 'priority':
        base_rates = XPARCEL_PRIORITY_RATES
    elif service_type == 'expedited':
        base_rates = XPARCEL_EXPEDITED_RATES
    else:
        base_rates = XPARCEL_GROUND_RATES

    # Get base rate for zone
    base_rate = base_rates.get(zone, base_rates[8])

    # Apply weight-based pricing
    if weight <= 0.0625:    # 1 oz
        return base_rate
    elif weight <= 0.25:     # 4 oz
        return base_rate + 0.10
    elif weight <= 0.5:      # 8 oz
        return base_rate + 0.25
    elif weight <= 1.0:      # 1 lb
        return base_rate + 0.50
    elif weight <= 2.0:      # 2 lb
        return base_rate + 1.30
    elif weight <= 5.0:      # 5 lb
        return base_rate + 4.10
    elif weight <= 10.0:     # 10 lb
        return base_rate + 7.50
    else:                    # Over 10 lb
        return base_rate + 7.50 + (weight - 10) * 0.35

def map_service_to_xparcel(service_name):
    """Map current service to Xparcel service"""
    service_lower = str(service_name).lower()

    if any(x in service_lower for x in ['2nd day', '2 day', 'expedited', 'express']):
        return 'expedited'
    elif any(x in service_lower for x in ['next day', 'overnight', 'priority']):
        return 'priority'
    else:
        return 'ground'

def create_styled_workbook(df):
    """Create the 9-tab Excel workbook with professional styling"""

    wb = Workbook()

    # Define styles
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_fill = PatternFill(start_color="1E4C8B", end_color="1E4C8B", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    subheader_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    subheader_font = Font(bold=True, size=10)

    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    # Tab 1: Executive Summary
    ws1 = wb.active
    ws1.title = "Executive Summary"

    ws1.merge_cells('A1:F1')
    ws1['A1'] = f"{COMPANY_NAME} - FirstMile Xparcel Shipping Analysis"
    ws1['A1'].font = Font(size=16, bold=True, color="1E4C8B")
    ws1['A1'].alignment = Alignment(horizontal="center")

    ws1['A3'] = "Analysis Date:"
    ws1['B3'] = ANALYSIS_DATE

    ws1['A5'] = "FINANCIAL SUMMARY"
    ws1['A5'].font = subheader_font
    ws1['A5'].fill = subheader_fill

    summary_data = [
        ["Metric", "Current", "With FirstMile", "Savings", "Savings %"],
        ["Monthly Spend", f"${df['Label Cost'].sum():,.2f}",
         f"${df['Xparcel_Cost'].sum():,.2f}",
         f"${df['Savings'].sum():,.2f}",
         f"{(df['Savings'].sum() / df['Label Cost'].sum() * 100):.1f}%"],
        ["Annual Projection", f"${df['Label Cost'].sum() * 12:,.2f}",
         f"${df['Xparcel_Cost'].sum() * 12:,.2f}",
         f"${df['Savings'].sum() * 12:,.2f}",
         f"{(df['Savings'].sum() / df['Label Cost'].sum() * 100):.1f}%"],
        ["Avg Cost/Package", f"${df['Label Cost'].mean():.2f}",
         f"${df['Xparcel_Cost'].mean():.2f}",
         f"${df['Savings'].mean():.2f}",
         f"{(df['Savings'].mean() / df['Label Cost'].mean() * 100):.1f}%"]
    ]

    for r_idx, row in enumerate(summary_data, start=7):
        for c_idx, value in enumerate(row, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 7:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            cell.border = border

    # Tab 2: Shipment Details
    ws2 = wb.create_sheet("Shipment Details")

    details_df = df.head(1000)[['Tracking Number', 'Shipping Method', 'Weight (lb)',
                                 'Zone', 'State', 'Label Cost', 'Xparcel_Cost',
                                 'Savings', 'Savings_Pct']]

    # Write headers
    for c_idx, col_name in enumerate(details_df.columns, start=1):
        cell = ws2.cell(row=1, column=c_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(details_df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border

            # Format currency columns
            if c_idx in [6, 7, 8]:  # Label Cost, Xparcel_Cost, Savings
                cell.number_format = '$#,##0.00'
            elif c_idx == 9:  # Savings_Pct
                cell.number_format = '0.0%'

    # Tab 3: Rate Comparison
    ws3 = wb.create_sheet("Rate Comparison")

    # Create sample weight/zone matrix
    sample_weights = [0.0625, 0.25, 0.5, 1.0, 2.0, 5.0, 10.0]
    sample_zones = [2, 3, 4, 5, 6, 7]

    ws3['A1'] = "GROUND SERVICE RATE COMPARISON"
    ws3['A1'].font = Font(size=14, bold=True, color="1E4C8B")

    ws3['A3'] = "Weight (lb)"
    for c_idx, zone in enumerate(sample_zones, start=2):
        ws3.cell(row=3, column=c_idx, value=f"Zone {zone}")

    for r_idx, weight in enumerate(sample_weights, start=4):
        ws3.cell(row=r_idx, column=1, value=f"{weight} lb" if weight >= 1 else f"{weight*16:.1f} oz")

        for c_idx, zone in enumerate(sample_zones, start=2):
            xparcel_cost = calculate_xparcel_cost(weight, zone, 'ground')
            # Estimate current cost (typically 40-50% higher)
            current_cost = xparcel_cost * 1.45
            savings_pct = (current_cost - xparcel_cost) / current_cost

            cell = ws3.cell(row=r_idx, column=c_idx, value=f"${xparcel_cost:.2f}")

            # Color code based on savings
            if savings_pct >= 0.5:
                cell.font = Font(color="008000", bold=True)  # Green
            elif savings_pct >= 0.3:
                cell.font = Font(color="0000FF", bold=True)  # Blue

    # Tab 4: Zone Analysis
    ws4 = wb.create_sheet("Zone Analysis")

    zone_analysis = df.groupby('Zone').agg({
        'Shipping Label ID': 'count',
        'Weight (lb)': 'mean',
        'Label Cost': ['sum', 'mean'],
        'Xparcel_Cost': ['sum', 'mean'],
        'Savings': ['sum', 'mean']
    }).round(2)

    ws4['A1'] = "ZONE DISTRIBUTION ANALYSIS"
    ws4['A1'].font = Font(size=14, bold=True, color="1E4C8B")

    # Tab 5: Service Analysis
    ws5 = wb.create_sheet("Service Analysis")

    service_analysis = df.groupby('Xparcel_Service').agg({
        'Shipping Label ID': 'count',
        'Label Cost': ['sum', 'mean'],
        'Xparcel_Cost': ['sum', 'mean'],
        'Savings': ['sum', 'mean'],
        'Savings_Pct': 'mean'
    }).round(2)

    ws5['A1'] = "SERVICE LEVEL ANALYSIS"
    ws5['A1'].font = Font(size=14, bold=True, color="1E4C8B")

    # Tab 6: Weight Distribution
    ws6 = wb.create_sheet("Weight Distribution")

    # Create weight buckets
    weight_bins = [0, 0.25, 0.5, 1, 2, 5, 10, 20, 100]
    weight_labels = ['0-4oz', '4-8oz', '8oz-1lb', '1-2lb', '2-5lb', '5-10lb', '10-20lb', '20lb+']
    df['Weight_Bucket'] = pd.cut(df['Weight (lb)'], bins=weight_bins, labels=weight_labels)

    weight_analysis = df.groupby('Weight_Bucket').agg({
        'Shipping Label ID': 'count',
        'Label Cost': ['sum', 'mean'],
        'Xparcel_Cost': ['sum', 'mean'],
        'Savings': ['sum', 'mean']
    }).round(2)

    ws6['A1'] = "WEIGHT DISTRIBUTION ANALYSIS"
    ws6['A1'].font = Font(size=14, bold=True, color="1E4C8B")

    # Tab 7: Savings Breakdown
    ws7 = wb.create_sheet("Savings Breakdown")

    top_savings = df.nlargest(20, 'Savings')[['Tracking Number', 'Weight (lb)',
                                               'Zone', 'Label Cost', 'Xparcel_Cost',
                                               'Savings', 'Savings_Pct']]

    ws7['A1'] = "TOP 20 SAVINGS OPPORTUNITIES"
    ws7['A1'].font = Font(size=14, bold=True, color="1E4C8B")

    # Tab 8: Monthly Projections
    ws8 = wb.create_sheet("Monthly Projections")

    monthly_volume = len(df)
    monthly_current = df['Label Cost'].sum()
    monthly_xparcel = df['Xparcel_Cost'].sum()
    monthly_savings = df['Savings'].sum()

    months = ['January', 'February', 'March', 'April', 'May', 'June',
              'July', 'August', 'September', 'October', 'November', 'December']

    seasonal_factors = {
        'January': 0.85, 'February': 0.85, 'March': 1.0, 'April': 1.0,
        'May': 1.0, 'June': 1.0, 'July': 1.0, 'August': 1.0,
        'September': 1.0, 'October': 1.0, 'November': 1.35, 'December': 1.35
    }

    ws8['A1'] = "12-MONTH PROJECTION WITH SEASONAL ADJUSTMENTS"
    ws8['A1'].font = Font(size=14, bold=True, color="1E4C8B")

    projection_headers = ['Month', 'Volume', 'Current Cost', 'Xparcel Cost', 'Savings', 'Cumulative Savings']
    for c_idx, header in enumerate(projection_headers, start=1):
        cell = ws8.cell(row=3, column=c_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    cumulative_savings = 0
    for r_idx, month in enumerate(months, start=4):
        factor = seasonal_factors[month]
        month_volume = int(monthly_volume * factor)
        month_current = monthly_current * factor
        month_xparcel = monthly_xparcel * factor
        month_savings = monthly_savings * factor
        cumulative_savings += month_savings

        row_data = [month, month_volume, month_current, month_xparcel, month_savings, cumulative_savings]
        for c_idx, value in enumerate(row_data, start=1):
            cell = ws8.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            if c_idx > 2:
                cell.number_format = '$#,##0.00'

    # Tab 9: Service Level Comparison
    ws9 = wb.create_sheet("Service Level Comparison")

    ws9['A1'] = "SERVICE LEVEL COMPARISON"
    ws9['A1'].font = Font(size=14, bold=True, color="1E4C8B")

    service_comparison = [
        ["Current Service", "Xparcel Service", "Transit Days", "Insurance", "Tracking", "Cost Index"],
        ["UPS Ground", "Xparcel Ground", "3-8", "$100 included", "Full visibility", "Lowest"],
        ["FedEx Ground", "Xparcel Ground", "3-8", "$100 included", "Full visibility", "Lowest"],
        ["UPS SurePost", "Xparcel Ground", "3-8", "$100 included", "Full visibility", "Lower"],
        ["FedEx SmartPost", "Xparcel Ground", "3-8", "$100 included", "Full visibility", "Lower"],
        ["UPS 2nd Day Air", "Xparcel Expedited", "2-5", "$100 included", "Full visibility", "Medium"],
        ["FedEx 2Day", "Xparcel Expedited", "2-5", "$100 included", "Full visibility", "Medium"],
        ["UPS Next Day Air", "Xparcel Priority", "1-3", "$100 included", "Full visibility", "Higher"],
        ["FedEx Priority", "Xparcel Priority", "1-3", "$100 included", "Full visibility", "Higher"]
    ]

    for r_idx, row in enumerate(service_comparison, start=3):
        for c_idx, value in enumerate(row, start=1):
            cell = ws9.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 3:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            cell.border = border

    # Auto-adjust column widths for all sheets
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = None

            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass

            if column_letter:
                adjusted_width = min(max_length + 2, 45)
                ws.column_dimensions[column_letter].width = adjusted_width

    return wb

def main():
    print("=" * 80)
    print("INVOICE AUDIT BUILDER v3.1 - STACKD LOGISTICS ANALYSIS")
    print("=" * 80)

    # Load data
    print(f"\nLoading data from: {INPUT_FILE}")
    df = pd.read_csv(INPUT_FILE)
    print(f"Loaded {len(df)} shipments")

    # Data preparation
    print("\nPreparing data for analysis...")

    # Convert weight to float
    df['Weight (lb)'] = pd.to_numeric(df['Weight (lb)'], errors='coerce').fillna(0.5)

    # Convert label cost to float
    df['Label Cost'] = pd.to_numeric(df['Label Cost'], errors='coerce').fillna(0)

    # Assign zones based on state
    df['Zone'] = df['State'].map(STATE_TO_ZONE).fillna(5)

    # Map services to Xparcel
    df['Xparcel_Service'] = df['Shipping Method'].apply(map_service_to_xparcel)

    # Calculate Xparcel costs
    print("Calculating FirstMile Xparcel rates...")
    df['Xparcel_Cost'] = df.apply(lambda row: calculate_xparcel_cost(
        row['Weight (lb)'],
        row['Zone'],
        row['Xparcel_Service']
    ), axis=1)

    # Calculate savings
    df['Savings'] = df['Label Cost'] - df['Xparcel_Cost']
    df['Savings_Pct'] = df['Savings'] / df['Label Cost']

    # Filter to positive savings
    positive_savings = df[df['Savings'] > 0]

    # Display summary statistics
    print("\n" + "=" * 80)
    print("ANALYSIS SUMMARY")
    print("=" * 80)

    print(f"Total Shipments Analyzed: {len(df):,}")
    print(f"Shipments with Savings: {len(positive_savings):,} ({len(positive_savings)/len(df)*100:.1f}%)")
    print(f"\nCurrent Monthly Spend: ${df['Label Cost'].sum():,.2f}")
    print(f"Projected Xparcel Cost: ${df['Xparcel_Cost'].sum():,.2f}")
    print(f"Monthly Savings: ${df['Savings'].sum():,.2f}")
    print(f"Savings Percentage: {(df['Savings'].sum() / df['Label Cost'].sum() * 100):.1f}%")
    print(f"\nAnnual Savings Projection: ${df['Savings'].sum() * 12:,.2f}")

    print(f"\nAverage Weight: {df['Weight (lb)'].mean():.2f} lb")
    print(f"Most Common Zone: Zone {int(df['Zone'].mode()[0])}")

    # Service mix
    service_mix = df['Xparcel_Service'].value_counts(normalize=True) * 100
    print("\nService Mix:")
    for service, pct in service_mix.items():
        print(f"  {service.title()}: {pct:.1f}%")

    # Create Excel workbook
    print("\nGenerating Excel workbook...")
    wb = create_styled_workbook(df)

    # Save workbook
    output_file = f"{COMPANY_NAME.replace(' ', '_')}_Complete_Audit_v3.1.xlsx"
    wb.save(output_file)
    print(f"\nSaved: {output_file}")

    # Save detailed CSV
    csv_file = f"{COMPANY_NAME.replace(' ', '_')}_Detailed_Data.csv"
    df.to_csv(csv_file, index=False)
    print(f"Saved: {csv_file}")

    print("\n" + "=" * 80)
    print("INVOICE AUDIT BUILDER v3.1 - EXECUTION COMPLETE")
    print("=" * 80)
    print(f"Total Tabs Created: 9")
    print(f"Shipments Analyzed: {len(df):,}")
    print(f"Savings Identified: ${df['Savings'].sum():,.2f}")
    print(f"Savings Percentage: {(df['Savings'].sum() / df['Label Cost'].sum() * 100):.1f}%")
    print(f"Annual Projection: ${df['Savings'].sum() * 12:,.2f}")
    print("=" * 80)

if __name__ == "__main__":
    main()