"""
FirstMile Xparcel Performance Report Generator
Creates professional Excel deliverable analyzing FirstMile carrier performance
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule

# Configuration
DATA_FILE = r"C:\Users\BrettWalker\FirstMile_Deals\[07-CLOSED-WON]_JM_Group_NY\JM Group NY_FirstMile_Domestic_Tracking_9.30.25.xlsx"
CUSTOMER_NAME = "JM Group NY"

# Brand Colors
FIRSTMILE_BLUE = "366092"
LIGHT_GRAY = "DDDDDD"
RED_BG = "FFC7CE"
YELLOW_BG = "FFEB84"
GREEN_BG = "C6EFCE"

# SLA Windows
SLA_WINDOWS = {
    "Xparcel Priority": 3,
    "Xparcel Expedited": 5,
    "Xparcel Ground": 8
}

# Performance Thresholds
PERF_THRESHOLDS = [
    (100.0, "Perfect Compliance"),
    (95.0, "Exceeds Standard"),
    (90.0, "Meets Standard"),
    (0.0, "Below Standard")
]

# Hub Mapping
HUB_MAP = {
    "CA": "LAX - West Coast",
    "TX": "DFW - South Central",
    "FL": "MIA - Southeast",
    "NY": "JFK/EWR - Northeast",
    "IL": "ORD - Midwest",
    "GA": "ATL - Southeast",
    "NJ": "JFK/EWR - Northeast",
    "PA": "JFK/EWR - Northeast",
    "OH": "ORD - Midwest",
    "MI": "ORD - Midwest",
    "WI": "ORD - Midwest"
}

def load_data(file_path):
    """Load and standardize the input Excel file"""
    print(f"Loading data from {file_path}...")
    df = pd.read_excel(file_path)

    # Display columns for debugging
    print(f"Columns found: {df.columns.tolist()}")
    print(f"Total rows: {len(df)}")

    return df

def detect_service_level(df):
    """Detect the service level from data patterns"""
    # Check for Xparcel Type column first (most reliable)
    if 'Xparcel Type' in df.columns:
        xparcel_counts = df['Xparcel Type'].value_counts()
        print(f"\nXparcel types detected: {xparcel_counts.to_dict()}")
        most_common = xparcel_counts.idxmax()

        # Map Xparcel Type to service level
        xparcel_map = {
            'Ground': 'Xparcel Ground',
            'Expedited': 'Xparcel Expedited',
            'Priority': 'Xparcel Priority',
            'Direct Call': 'Xparcel Priority'
        }

        for key, value in xparcel_map.items():
            if key.lower() in str(most_common).lower():
                print(f"Detected service level: {value}")
                return value

    # Fallback to Service Level column
    if 'Service Level' in df.columns:
        service_counts = df['Service Level'].value_counts()
        print(f"\nService levels detected: {service_counts.to_dict()}")
        most_common = service_counts.idxmax()

        # Map to Xparcel naming
        service_map = {
            'Priority': 'Xparcel Priority',
            'Direct Call': 'Xparcel Priority',
            'Expedited': 'Xparcel Expedited',
            'Ground': 'Xparcel Ground'
        }

        for key, value in service_map.items():
            if key.lower() in most_common.lower():
                return value

    # Default to analyzing transit days of DELIVERED packages only
    if 'Days In Transit' in df.columns:
        delivered = df[df['Delivered Status'] == 'Delivered']
        if len(delivered) > 0:
            avg_transit = delivered['Days In Transit'].mean()
            print(f"\nAverage transit days (delivered only): {avg_transit:.1f}")
            if avg_transit <= 3:
                return 'Xparcel Priority'
            elif avg_transit <= 5:
                return 'Xparcel Expedited'
            else:
                return 'Xparcel Ground'

    return 'Xparcel Ground'  # Default

def detect_report_period(df):
    """Detect the report period from the data"""
    date_columns = ['Request Date', 'ShipDate', 'Ship Date', 'Created Date', 'Delivery Date']

    for col in date_columns:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
            valid_dates = df[col].dropna()
            if len(valid_dates) > 0:
                min_date = valid_dates.min()
                max_date = valid_dates.max()
                return f"{min_date.strftime('%B %d')} - {max_date.strftime('%B %d, %Y')}"

    return "Date Range Not Available"

def calculate_sla_compliance(df, service_level):
    """Calculate SLA compliance for delivered packages only"""
    sla_window = SLA_WINDOWS[service_level]

    # Filter for delivered only
    delivered = df[df['Delivered Status'] == 'Delivered'].copy()

    if len(delivered) == 0:
        return {
            'total_delivered': 0,
            'within_sla': 0,
            'compliance_pct': 0.0,
            'performance_status': 'No Data'
        }

    within_sla = len(delivered[delivered['Days In Transit'] <= sla_window])
    compliance_pct = (within_sla / len(delivered)) * 100

    # Determine performance status
    performance_status = next(
        status for threshold, status in PERF_THRESHOLDS
        if compliance_pct >= threshold
    )

    return {
        'total_delivered': len(delivered),
        'within_sla': within_sla,
        'compliance_pct': compliance_pct,
        'performance_status': performance_status
    }

def calculate_transit_performance(df):
    """Calculate transit performance distribution for delivered packages"""
    delivered = df[df['Delivered Status'] == 'Delivered'].copy()

    if len(delivered) == 0:
        return pd.DataFrame()

    # Create distribution
    transit_dist = []
    for day in range(8):
        count = len(delivered[delivered['Days In Transit'] == day])
        pct = (count / len(delivered)) * 100
        transit_dist.append({
            'Transit Days': f'{day} Days',
            'Shipment Count': count,
            'Percentage': pct
        })

    # 8+ days
    count_8plus = len(delivered[delivered['Days In Transit'] >= 8])
    pct_8plus = (count_8plus / len(delivered)) * 100
    transit_dist.append({
        'Transit Days': '8+ Days',
        'Shipment Count': count_8plus,
        'Percentage': pct_8plus
    })

    df_dist = pd.DataFrame(transit_dist)

    # Add statistics
    stats = {
        'Average Transit Days': delivered['Days In Transit'].mean(),
        'Median Transit Days': delivered['Days In Transit'].median(),
        '90th Percentile': delivered['Days In Transit'].quantile(0.90),
        '95th Percentile': delivered['Days In Transit'].quantile(0.95)
    }

    return df_dist, stats

def calculate_geographic_distribution(df):
    """Calculate geographic distribution with hub mapping"""
    if 'Destination State' not in df.columns:
        return pd.DataFrame()

    state_counts = df['Destination State'].value_counts().head(15)

    geo_data = []
    for state, count in state_counts.items():
        pct = (count / len(df)) * 100
        hub = HUB_MAP.get(state, "National Network")
        network_type = "Select Network" if state in HUB_MAP else "National Network"

        geo_data.append({
            'State': state,
            'Shipment Count': count,
            'Percentage': pct,
            'Primary Hub': hub,
            'Network Type': network_type
        })

    return pd.DataFrame(geo_data)

def calculate_zone_analysis(df):
    """Calculate zone distribution and regional vs cross-country split"""
    if 'Calculated Zone' not in df.columns:
        return pd.DataFrame(), {}

    # Clean zone data
    df['Calculated Zone'] = df['Calculated Zone'].astype(str).str.extract(r'(\d+)')[0].astype(float)

    zone_data = []
    for zone in range(1, 9):
        zone_df = df[df['Calculated Zone'] == zone]
        count = len(zone_df)
        pct = (count / len(df)) * 100

        if count > 0 and 'Days In Transit' in zone_df.columns:
            avg_transit = zone_df['Days In Transit'].mean()
        else:
            avg_transit = 0

        zone_data.append({
            'Zone': f'Zone {zone}',
            'Shipment Count': count,
            'Percentage': pct,
            'Avg Transit Days': avg_transit
        })

    df_zones = pd.DataFrame(zone_data)

    # Regional vs Cross-Country
    regional = df[df['Calculated Zone'].isin([1, 2, 3, 4])]
    cross_country = df[df['Calculated Zone'].isin([5, 6, 7, 8])]

    summary = {
        'Regional (Zones 1-4)': {
            'count': len(regional),
            'pct': (len(regional) / len(df)) * 100
        },
        'Cross-Country (Zones 5-8)': {
            'count': len(cross_country),
            'pct': (len(cross_country) / len(df)) * 100
        }
    }

    return df_zones, summary

def calculate_in_transit_status(df, service_level):
    """Calculate in-transit shipment status with SLA window indicator"""
    sla_window = SLA_WINDOWS[service_level]

    # Filter for non-delivered
    in_transit = df[df['Delivered Status'] != 'Delivered'].copy()

    if len(in_transit) == 0:
        return pd.DataFrame(), {}

    # Calculate days since ship using Start Date (most accurate)
    today = pd.Timestamp.now()

    # Try different date columns in priority order
    date_col = None
    for col in ['Start Date', 'ShipDate', 'Ship Date', 'Request Date', 'Created Date']:
        if col in in_transit.columns:
            date_col = col
            print(f"Using '{col}' for in-transit calculations")
            break

    if date_col:
        in_transit[date_col] = pd.to_datetime(in_transit[date_col], errors='coerce')
        in_transit['Days Since Ship'] = (today - in_transit[date_col]).dt.days
        in_transit['Within SLA Window'] = in_transit['Days Since Ship'].apply(
            lambda x: 'Yes' if x <= sla_window else 'No'
        )
    else:
        in_transit['Days Since Ship'] = 0
        in_transit['Within SLA Window'] = 'Unknown'

    # Calculate summary statistics
    within_sla_count = len(in_transit[in_transit['Within SLA Window'] == 'Yes'])
    outside_sla_count = len(in_transit[in_transit['Within SLA Window'] == 'No'])

    summary = {
        'total_in_transit': len(in_transit),
        'within_sla_window': within_sla_count,
        'outside_sla_window': outside_sla_count,
        'within_sla_pct': (within_sla_count / len(in_transit)) * 100 if len(in_transit) > 0 else 0
    }

    # Select relevant columns
    cols_to_keep = ['Delivered Status', 'Days Since Ship', 'Within SLA Window']
    if 'Destination State' in in_transit.columns:
        cols_to_keep.append('Destination State')
    if 'Calculated Zone' in in_transit.columns:
        cols_to_keep.append('Calculated Zone')
    if date_col and date_col in in_transit.columns:
        cols_to_keep.insert(1, date_col)

    return in_transit[cols_to_keep], summary

def apply_header_style(ws, row=1):
    """Apply FirstMile blue header styling"""
    blue_fill = PatternFill(start_color=FIRSTMILE_BLUE, end_color=FIRSTMILE_BLUE, fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[row]:
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center_align

def apply_data_style(ws, start_row=2):
    """Apply data cell styling with borders"""
    thin_border = Border(
        left=Side(style='thin', color=LIGHT_GRAY),
        right=Side(style='thin', color=LIGHT_GRAY),
        top=Side(style='thin', color=LIGHT_GRAY),
        bottom=Side(style='thin', color=LIGHT_GRAY)
    )
    center_align = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align

def auto_size_columns(ws, max_width=50):
    """Auto-size columns based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_executive_summary(wb, df, service_level, report_period, sla_data, in_transit_summary):
    """Create Executive Summary tab"""
    ws = wb.create_sheet("Executive Summary", 0)

    # Header
    ws['A1'] = 'FirstMile Xparcel Performance Report'
    ws['A1'].font = Font(size=14, bold=True, color=FIRSTMILE_BLUE)

    ws['A2'] = f'Customer: {CUSTOMER_NAME}'
    ws['A3'] = f'Report Period: {report_period}'
    ws['A4'] = f'Service Level: {service_level}'

    # KPI Table
    ws['A6'] = 'Key Performance Indicator'
    ws['B6'] = 'Value'
    apply_header_style(ws, row=6)

    kpis = [
        ('Total Shipments', len(df)),
        ('Total Delivered', sla_data['total_delivered']),
        ('SLA Compliance Rate (Delivered Only)', f"{sla_data['compliance_pct']:.1f}%"),
        ('Performance Status', sla_data['performance_status']),
        ('SLA Window', f"{SLA_WINDOWS[service_level]} Days"),
        ('Shipments Within SLA', sla_data['within_sla']),
        ('', ''),
        ('In-Transit Shipments', in_transit_summary.get('total_in_transit', 0)),
        ('In-Transit Within SLA Window', in_transit_summary.get('within_sla_window', 0)),
        ('In-Transit Outside SLA Window', in_transit_summary.get('outside_sla_window', 0))
    ]

    for i, (label, value) in enumerate(kpis, start=7):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value

    apply_data_style(ws, start_row=7)
    auto_size_columns(ws)

def create_sla_compliance_tab(wb, sla_data):
    """Create SLA Compliance tab with conditional formatting"""
    ws = wb.create_sheet("SLA Compliance")

    # Headers
    ws['A1'] = 'Metric'
    ws['B1'] = 'Value'
    apply_header_style(ws, row=1)

    # Data
    data = [
        ('Total Delivered Shipments', sla_data['total_delivered']),
        ('Shipments Within SLA', sla_data['within_sla']),
        ('SLA Compliance Rate', f"{sla_data['compliance_pct']:.1f}%"),
        ('Performance Status', sla_data['performance_status'])
    ]

    for i, (label, value) in enumerate(data, start=2):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value

    # Apply conditional formatting to compliance rate
    # Color scale: red (80-89), yellow (90-94), green (95-100)
    ws['B4'].number_format = '0.0'

    apply_data_style(ws, start_row=2)
    auto_size_columns(ws)

def create_transit_performance_tab(wb, transit_df, stats):
    """Create Transit Performance tab"""
    ws = wb.create_sheet("Transit Performance")

    # Add distribution table
    for r_idx, row in enumerate(dataframe_to_rows(transit_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    apply_header_style(ws, row=1)
    apply_data_style(ws, start_row=2)

    # Add statistics below
    stats_start = len(transit_df) + 3
    ws[f'A{stats_start}'] = 'Statistical Summary'
    ws[f'A{stats_start}'].font = Font(bold=True, size=12)

    stats_start += 1
    ws[f'A{stats_start}'] = 'Metric'
    ws[f'B{stats_start}'] = 'Value'
    apply_header_style(ws, row=stats_start)

    for i, (label, value) in enumerate(stats.items(), start=stats_start+1):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = f"{value:.1f}"

    auto_size_columns(ws)

def create_geographic_tab(wb, geo_df):
    """Create Geographic Distribution tab"""
    ws = wb.create_sheet("Geographic Distribution")

    for r_idx, row in enumerate(dataframe_to_rows(geo_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    apply_header_style(ws, row=1)
    apply_data_style(ws, start_row=2)
    auto_size_columns(ws)

def create_zone_analysis_tab(wb, zone_df, summary):
    """Create Zone Analysis tab"""
    ws = wb.create_sheet("Zone Analysis")

    # Zone distribution
    for r_idx, row in enumerate(dataframe_to_rows(zone_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    apply_header_style(ws, row=1)
    apply_data_style(ws, start_row=2)

    # Regional summary
    summary_start = len(zone_df) + 3
    ws[f'A{summary_start}'] = 'Regional vs Cross-Country Summary'
    ws[f'A{summary_start}'].font = Font(bold=True, size=12)

    summary_start += 1
    ws[f'A{summary_start}'] = 'Category'
    ws[f'B{summary_start}'] = 'Shipment Count'
    ws[f'C{summary_start}'] = 'Percentage'
    apply_header_style(ws, row=summary_start)

    for i, (category, data) in enumerate(summary.items(), start=summary_start+1):
        ws[f'A{i}'] = category
        ws[f'B{i}'] = data['count']
        ws[f'C{i}'] = f"{data['pct']:.1f}%"

    auto_size_columns(ws)

def create_operational_metrics_tab(wb, df, in_transit_summary):
    """Create Operational Metrics tab"""
    ws = wb.create_sheet("Operational Metrics")

    # Volume metrics
    ws['A1'] = 'Volume Metrics'
    ws['A1'].font = Font(bold=True, size=12)

    ws['A2'] = 'Metric'
    ws['B2'] = 'Value'
    apply_header_style(ws, row=2)

    metrics = [
        ('Total Shipments', len(df)),
        ('Daily Average', len(df) / 30),  # Approximate
        ('Delivered Shipments', len(df[df['Delivered Status'] == 'Delivered'])),
        ('In-Transit Shipments', in_transit_summary.get('total_in_transit', 0)),
        ('In-Transit Within SLA Window', in_transit_summary.get('within_sla_window', 0)),
        ('In-Transit Outside SLA Window', in_transit_summary.get('outside_sla_window', 0))
    ]

    for i, (label, value) in enumerate(metrics, start=3):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = f"{value:.0f}" if isinstance(value, float) else value

    apply_data_style(ws, start_row=3)
    auto_size_columns(ws)

def create_in_transit_tab(wb, in_transit_df):
    """Create In-Transit Detail tab"""
    ws = wb.create_sheet("In-Transit Detail")

    if len(in_transit_df) == 0:
        ws['A1'] = 'No in-transit shipments found'
        return

    for r_idx, row in enumerate(dataframe_to_rows(in_transit_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    apply_header_style(ws, row=1)
    apply_data_style(ws, start_row=2)
    auto_size_columns(ws)

def create_notes_tab(wb, service_level):
    """Create Notes & Assumptions tab"""
    ws = wb.create_sheet("Notes & Assumptions")

    notes = [
        ('Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('', ''),
        ('Business Rules', ''),
        ('SLA Calculation', 'Based on delivered packages only'),
        ('SLA Window', f'{SLA_WINDOWS[service_level]} days for {service_level}'),
        ('In-Transit Status', 'Shipments not yet delivered are excluded from SLA calculations'),
        ('Performance Thresholds', '100%=Perfect, ≥95%=Exceeds, ≥90%=Meets, <90%=Below'),
        ('', ''),
        ('Definitions', ''),
        ('FirstMile', 'The carrier providing shipping services'),
        ('Xparcel', 'The ship method (Priority, Expedited, Ground)'),
        ('National Network', 'Nationwide coverage for all ZIP codes'),
        ('Select Network', 'Metro-focused injection points in major hubs')
    ]

    for i, (label, value) in enumerate(notes, start=1):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value

    auto_size_columns(ws)

def create_brand_style_guide_tab(wb):
    """Create Brand Style Guide tab"""
    ws = wb.create_sheet("Brand Style Guide")

    ws['A1'] = 'FirstMile Brand Colors'
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = 'Color Name'
    ws['B3'] = 'HEX'
    ws['C3'] = 'RGB'
    ws['D3'] = 'Usage'
    apply_header_style(ws, row=3)

    colors = [
        ('FirstMile Blue', '#366092', 'RGB(54, 96, 146)', 'Primary brand color, headers'),
        ('Light Gray', '#DDDDDD', 'RGB(221, 221, 221)', 'Borders, dividers'),
        ('Red', '#FFC7CE', 'RGB(255, 199, 206)', 'Below standard performance (80-89%)'),
        ('Yellow', '#FFEB84', 'RGB(255, 235, 132)', 'Meets standard performance (90-94%)'),
        ('Green', '#C6EFCE', 'RGB(198, 239, 206)', 'Exceeds standard performance (95-100%)')
    ]

    for i, (name, hex_code, rgb, usage) in enumerate(colors, start=4):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = hex_code
        ws[f'C{i}'] = rgb
        ws[f'D{i}'] = usage

    apply_data_style(ws, start_row=4)
    auto_size_columns(ws)

def generate_report():
    """Main report generation function"""
    print("="*60)
    print("FirstMile Xparcel Performance Report Generator")
    print("="*60)

    # Load data
    df = load_data(DATA_FILE)

    # Detect service level and report period
    service_level = detect_service_level(df)
    report_period = detect_report_period(df)

    print(f"\nDetected Service Level: {service_level}")
    print(f"Report Period: {report_period}")

    # Calculate all analytics
    print("\nCalculating SLA compliance...")
    sla_data = calculate_sla_compliance(df, service_level)

    print("Calculating transit performance...")
    transit_df, transit_stats = calculate_transit_performance(df)

    print("Calculating geographic distribution...")
    geo_df = calculate_geographic_distribution(df)

    print("Calculating zone analysis...")
    zone_df, zone_summary = calculate_zone_analysis(df)

    print("Calculating in-transit status...")
    in_transit_df, in_transit_summary = calculate_in_transit_status(df, service_level)

    # Print in-transit summary
    print(f"\nIn-Transit Summary:")
    print(f"  Total In-Transit: {in_transit_summary.get('total_in_transit', 0)}")
    print(f"  Within SLA Window: {in_transit_summary.get('within_sla_window', 0)}")
    print(f"  Outside SLA Window: {in_transit_summary.get('outside_sla_window', 0)}")

    # Create Excel workbook
    print("\nGenerating Excel report...")
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create all tabs
    create_executive_summary(wb, df, service_level, report_period, sla_data, in_transit_summary)
    create_sla_compliance_tab(wb, sla_data)
    create_transit_performance_tab(wb, transit_df, transit_stats)
    create_geographic_tab(wb, geo_df)
    create_zone_analysis_tab(wb, zone_df, zone_summary)
    create_operational_metrics_tab(wb, df, in_transit_summary)
    create_in_transit_tab(wb, in_transit_df)
    create_notes_tab(wb, service_level)
    create_brand_style_guide_tab(wb)

    # Save workbook
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    output_file = f"FirstMile_Xparcel_Performance_{CUSTOMER_NAME.replace(' ', '_')}_{timestamp}.xlsx"
    output_path = r"C:\Users\BrettWalker\FirstMile_Deals\[07-CLOSED-WON]_JM_Group_NY\\" + output_file

    wb.save(output_path)

    print(f"\n{'='*60}")
    print(f"Report generated successfully!")
    print(f"{'='*60}")
    print(f"\nOutput file: {output_path}")
    print(f"\nSummary:")
    print(f"  Total Shipments: {len(df)}")
    print(f"  Total Delivered: {sla_data['total_delivered']}")
    print(f"  SLA Compliance: {sla_data['compliance_pct']:.1f}%")
    print(f"  Performance Status: {sla_data['performance_status']}")
    print(f"\n{'='*60}")

if __name__ == "__main__":
    generate_report()
