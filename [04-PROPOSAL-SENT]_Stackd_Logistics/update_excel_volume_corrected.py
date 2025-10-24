"""
Update Stackd Logistics Excel file with CORRECTED volume (2 weeks → monthly)
October 14, 2025
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# Load the existing workbook
wb = openpyxl.load_workbook('Stackd_Logistics_FirstMile_Xparcel_Savings_Analysis_20251014_1126.xlsx')

# CORRECTED METRICS
monthly_volume = 16_836
annual_volume = 202_032
monthly_savings = 6_962.22
annual_savings = 83_546.64
savings_pct = 9.5
avg_savings_pkg = 0.41

select_monthly_volume = 8_992
select_monthly_savings = 6_003.82
select_pct = 15.7

national_monthly_volume = 7_844
national_monthly_savings = 958.40
national_pct = 2.7

# Update Executive Summary sheet
exec_sheet = wb['Executive Summary']

# Find and update the key metrics
for row in exec_sheet.iter_rows(min_row=1, max_row=50):
    for cell in row:
        if cell.value:
            cell_text = str(cell.value)

            # Update volume figures
            if 'Monthly Package Volume' in cell_text or 'Addressable Volume (Monthly)' in cell_text:
                next_cell = exec_sheet.cell(row=cell.row, column=cell.column + 1)
                next_cell.value = f"{monthly_volume:,} packages/month"

            if 'Annual Package Volume' in cell_text:
                next_cell = exec_sheet.cell(row=cell.row, column=cell.column + 1)
                next_cell.value = f"{annual_volume:,} packages/year"

            # Update savings figures
            if 'Monthly Savings' in cell_text and 'Select' not in cell_text and 'National' not in cell_text:
                next_cell = exec_sheet.cell(row=cell.row, column=cell.column + 1)
                next_cell.value = f"${monthly_savings:,.2f}"

            if 'Annual Savings' in cell_text and '(' not in cell_text:
                next_cell = exec_sheet.cell(row=cell.row, column=cell.column + 1)
                next_cell.value = f"${annual_savings:,.2f} ({savings_pct}%)"

            if 'Average Savings per Package' in cell_text:
                next_cell = exec_sheet.cell(row=cell.row, column=cell.column + 1)
                next_cell.value = f"${avg_savings_pkg:.2f}"

print("Updated Executive Summary")

# Update FirstMile Advantages sheet - network breakdown
advantages_sheet = wb['FirstMile Advantages']

for row in advantages_sheet.iter_rows(min_row=1, max_row=100):
    for cell in row:
        if cell.value:
            cell_text = str(cell.value)

            # Update Select Network figures
            if 'Select Network' in cell_text and 'Monthly' in cell_text:
                # Find the savings cell
                for search_row in advantages_sheet.iter_rows(min_row=cell.row, max_row=cell.row+5):
                    for search_cell in search_row:
                        if search_cell.value and ('Savings' in str(search_cell.value) or '$' in str(search_cell.value)):
                            if search_cell.column == cell.column + 1:
                                search_cell.value = f"${select_monthly_savings:,.2f} ({select_pct}%)"

            # Update National Network figures
            if 'National Network' in cell_text and 'Monthly' in cell_text:
                for search_row in advantages_sheet.iter_rows(min_row=cell.row, max_row=cell.row+5):
                    for search_cell in search_row:
                        if search_cell.value and ('Savings' in str(search_cell.value) or '$' in str(search_cell.value)):
                            if search_cell.column == cell.column + 1:
                                search_cell.value = f"${national_monthly_savings:,.2f} ({national_pct}%)"

print("Updated FirstMile Advantages")

# Update Weight Analysis sheet header
weight_sheet = wb['Weight Analysis']
# Update title if it contains volume info
if weight_sheet['A1'].value:
    weight_sheet['A1'].value = f"Stackd Logistics Weight Analysis (Monthly Volume: {monthly_volume:,} packages)"

print("Updated Weight Analysis")

# Update Zone Analysis sheet header
zone_sheet = wb['Zone Analysis']
if zone_sheet['A1'].value:
    zone_sheet['A1'].value = f"Stackd Logistics Zone Analysis (Monthly Volume: {monthly_volume:,} packages)"

print("Updated Zone Analysis")

# Save with new timestamp
timestamp = datetime.now().strftime("%Y%m%d_%H%M")
new_filename = f"Stackd_Logistics_FirstMile_Xparcel_Savings_Analysis_CORRECTED_{timestamp}.xlsx"
wb.save(new_filename)

print(f"\nSaved corrected Excel file: {new_filename}")
print(f"\nCORRECTED FIGURES:")
print(f"  Monthly Volume: {monthly_volume:,} packages")
print(f"  Annual Volume: {annual_volume:,} packages")
print(f"  Monthly Savings: ${monthly_savings:,.2f}")
print(f"  Annual Savings: ${annual_savings:,.2f} ({savings_pct}%)")
print(f"  Select Network: ${select_monthly_savings:,.2f}/month ({select_pct}%)")
print(f"  National Network: ${national_monthly_savings:,.2f}/month ({national_pct}%)")
