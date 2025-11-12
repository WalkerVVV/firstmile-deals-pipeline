import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

print("=" * 70)
print("BoxiiShip BLANK -MP Complete Tracking Parser")
print("=" * 70)

# Read the complete tracking data file
with open('BLANK -MP BoxiiShip SB Logistics LLC TX Oct 6.txt', 'r', encoding='utf-8') as f:
    lines = [line.strip() for line in f.readlines()]

print(f"\nTotal lines in file: {len(lines)}")

# Parse tracking data
tracking_data = []
i = 0

while i < len(lines):
    line = lines[i]

    # Skip empty lines and header
    if not line or 'BoxiiShip' in line or 'Tracking IDs' in line:
        i += 1
        continue

    # Check if this is a tracking number (starts with digit and is long enough)
    if line and line[0].isdigit() and len(line) >= 18:
        tracking_num = line
        status = ''
        date = ''
        location = ''

        # Get status (next non-empty line)
        j = i + 1
        while j < len(lines) and not lines[j]:
            j += 1

        if j < len(lines) and 'View More' not in lines[j]:
            status = lines[j]
            j += 1

        # Get date (look for day of week)
        while j < len(lines) and not lines[j]:
            j += 1

        if j < len(lines):
            potential_date = lines[j]
            days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            if any(day in potential_date for day in days):
                date = potential_date
                j += 1

                # Check for location (next non-empty line before "View More")
                while j < len(lines) and not lines[j]:
                    j += 1

                if j < len(lines) and 'View More' not in lines[j]:
                    location = lines[j]

        tracking_data.append({
            'Tracking Number': tracking_num,
            'Status': status,
            'Date': date,
            'Location': location
        })

    i += 1

# Create DataFrame
df = pd.DataFrame(tracking_data)

print(f"\n[SUCCESS] Parsed {len(df)} tracking numbers")

# Show status distribution
print(f"\n--- Status Distribution ---")
status_counts = df['Status'].value_counts()
print(status_counts.head(20))

# Count specific statuses
blank_status = len(df[df['Status'] == ''])
shipment_info = len(df[df['Status'].str.contains('Shipment info received', case=False, na=False)])
delivered = len(df[df['Status'].str.contains('Delivered', case=False, na=False)])
in_transit = len(df[df['Status'].str.contains('transit|Accepted|Departed', case=False, na=False)])

print(f"\n--- Summary ---")
print(f"Total Tracking Numbers: {len(df)}")
print(f"Blank/No Status: {blank_status} ({blank_status/len(df)*100:.1f}%)")
print(f"Shipment info received: {shipment_info} ({shipment_info/len(df)*100:.1f}%)")
print(f"Delivered: {delivered} ({delivered/len(df)*100:.1f}%)")
print(f"In Transit: {in_transit} ({in_transit/len(df)*100:.1f}%)")

# Save to Excel with professional formatting
output_file = 'BLANK_MP_Tracking_Nov4_2025.xlsx'

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='BLANK -MP Tracking')

    # Get the worksheet
    worksheet = writer.sheets['BLANK -MP Tracking']

    # Set column widths
    worksheet.column_dimensions['A'].width = 28  # Tracking Number
    worksheet.column_dimensions['B'].width = 50  # Status
    worksheet.column_dimensions['C'].width = 30  # Date
    worksheet.column_dimensions['D'].width = 35  # Location

    # Format header row (FirstMile blue #366092)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Center align tracking numbers and dates
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    # Auto-filter
    worksheet.auto_filter.ref = worksheet.dimensions

print(f"\n[SUCCESS] Excel file created: {output_file}")
print(f"[SUCCESS] Contains {len(df)} BLANK -MP tracking numbers with complete data")
print("=" * 70)
