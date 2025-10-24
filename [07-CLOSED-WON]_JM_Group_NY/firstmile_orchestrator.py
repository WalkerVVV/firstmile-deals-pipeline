"""
FirstMile Xparcel Performance Report Orchestrator
Version: 1.0 (Customer-facing, brand-consistent, enterprise-grade)

Purpose
-------
Create a consistent, professional Excel deliverable for customers analyzing
FirstMile (carrier) performance using Xparcel ship methods (Priority, Expedited, Ground).

Key brand rules enforced in output:
- FirstMile is the carrier. Xparcel is the ship method.
- Service windows: Priority (1–3), Expedited (2–5), Ground (3–8)
- SLA is calculated on Delivered only. In-Transit items are excluded from SLA math
  and shown with an "Within SLA Window" indicator based on ship date + SLA window.
- Plain, factual language. No emojis or slogans. Consistent sheet names & formats.
"""

from __future__ import annotations
import os
from datetime import datetime, timedelta
from typing import Dict, Any, Tuple, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

# -----------------------------
# Configuration & Constants
# -----------------------------
PRIMARY_BLUE = "366092"  # FirstMile brand primary
WHITE = "FFFFFF"
LIGHT_BORDER = "DDDDDD"

HEADER_FILL = PatternFill(start_color=PRIMARY_BLUE, end_color=PRIMARY_BLUE, fill_type="solid")
WHITE_BOLD = Font(color=WHITE, bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(style="thin", color=LIGHT_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SLA_WINDOWS: Dict[str, int] = {
    "Xparcel Priority": 3,     # 1–3 day ship method
    "Xparcel Expedited": 5,    # 2–5 day ship method
    "Xparcel Ground": 8        # 3–8 day ship method
}

PERF_THRESHOLDS = [
    (100.0, "Perfect Compliance"),
    (95.0, "Exceeds Standard"),
    (90.0, "Meets Standard"),
    (0.0,  "Below Standard"),
]

REQUIRED_COLUMNS = [
    "Delivered Status",
    "Days In Transit",
    "Destination State",
    "Calculated Zone"
]

SHEET_ORDER = [
    "Executive Summary",
    "SLA Compliance",
    "Transit Performance",
    "Geographic Distribution",
    "Zone Analysis",
    "Operational Metrics",
    "In-Transit Detail",
    "Notes & Assumptions",
    "Brand Style Guide"
]

# Hub mapping
HUB_MAP = {
    "CA": "LAX - West Coast",
    "TX": "DFW - South Central",
    "FL": "MIA - Southeast",
    "NY": "JFK/EWR - Northeast",
    "IL": "ORD - Midwest",
    "GA": "ATL - Southeast",
    "NJ": "EWR - Northeast",
    "PA": "PHL - Mid-Atlantic",
    "OH": "CMH - Midwest",
    "NC": "CLT - Southeast"
}

# -----------------------------
# Utility functions
# -----------------------------

def _validate_columns(df: pd.DataFrame) -> None:
    """Check for required columns with flexible naming"""
    missing = []
    for col in REQUIRED_COLUMNS:
        if col not in df.columns:
            # Check for variations
            found = False
            for df_col in df.columns:
                if col.lower().replace(" ", "") in df_col.lower().replace(" ", ""):
                    found = True
                    break
            if not found:
                missing.append(col)

    if missing:
        print(f"Warning: Missing columns (will handle gracefully): {', '.join(missing)}")


def _coerce_numeric(series: pd.Series) -> pd.Series:
    """Convert to numeric, handling errors gracefully"""
    return pd.to_numeric(series, errors="coerce")


def _parse_date(series: pd.Series) -> pd.Series:
    """Parse dates, handling various formats"""
    return pd.to_datetime(series, errors="coerce")


def _status_label(compliance_pct: float) -> str:
    """Get performance status label based on percentage"""
    for threshold, label in PERF_THRESHOLDS:
        if compliance_pct >= threshold:
            return label
    return "Below Standard"


# -----------------------------
# Core analytics
# -----------------------------

def compute_sla(df: pd.DataFrame, service_level: str) -> Dict[str, Any]:
    """Compute SLA metrics using Delivered only."""
    if service_level not in SLA_WINDOWS:
        # Default to Expedited if not specified
        service_level = "Xparcel Expedited"

    window = SLA_WINDOWS[service_level]

    # Filter for delivered packages
    delivered = df[df["Delivered Status"].astype(str).str.lower() == "delivered"].copy()
    delivered["Days In Transit"] = _coerce_numeric(delivered["Days In Transit"])
    delivered = delivered.dropna(subset=["Days In Transit"])

    total = len(delivered)
    if total == 0:
        return {
            "service_level": service_level,
            "sla_window": window,
            "total_delivered": 0,
            "within_sla": 0,
            "exceeding_sla": 0,
            "compliance_percentage": 0.0,
            "performance_status": "No Delivered Volume"
        }

    within = int((delivered["Days In Transit"] <= window).sum())
    exceeding = total - within
    pct = round((within / total) * 100.0, 1)

    return {
        "service_level": service_level,
        "sla_window": window,
        "total_delivered": total,
        "within_sla": within,
        "exceeding_sla": exceeding,
        "compliance_percentage": pct,
        "performance_status": _status_label(pct)
    }


def compute_transit(df: pd.DataFrame) -> Dict[str, Any]:
    """Compute transit time distribution for delivered packages"""
    d = df.copy()
    d["Days In Transit"] = _coerce_numeric(d["Days In Transit"])
    d = d.dropna(subset=["Days In Transit"])

    # Distribution 0–7 individually, 8+ aggregated
    daily_distribution = []
    total = len(d)
    cum = 0

    if total > 0:
        for day in range(8):
            count = int((d["Days In Transit"] == day).sum())
            pct = (count / total * 100.0)
            cum += pct
            daily_distribution.append({
                "day": day,
                "count": count,
                "pct": round(pct, 1),
                "cumulative_pct": round(cum, 1)
            })

        # 8+ days
        count_8plus = int((d["Days In Transit"] >= 8).sum())
        if count_8plus > 0:
            pct = (count_8plus / total * 100.0)
            cum += pct
            daily_distribution.append({
                "day": "8+",
                "count": count_8plus,
                "pct": round(pct, 1),
                "cumulative_pct": round(cum, 1)
            })

        transit_days = d["Days In Transit"].to_numpy()
        stats = {
            "average": round(float(np.mean(transit_days)), 2),
            "median": round(float(np.median(transit_days)), 1),
            "p90": round(float(np.percentile(transit_days, 90)), 1),
            "p95": round(float(np.percentile(transit_days, 95)), 1)
        }
    else:
        stats = {
            "average": 0.0,
            "median": 0.0,
            "p90": 0.0,
            "p95": 0.0
        }

    return {"daily_distribution": daily_distribution, "statistics": stats}


def compute_geography(df: pd.DataFrame) -> Dict[str, Any]:
    """Compute geographic distribution"""
    d = df.copy()
    state_col = "Destination State"

    if state_col not in d.columns:
        return {"top_states": [], "select_network_pct": None, "national_network_pct": None}

    top = d.groupby(state_col).size().sort_values(ascending=False).head(15)
    total = len(d)
    rows = []

    select_states = ["CA", "TX", "FL", "NY", "IL", "GA"]
    select_count = 0
    national_count = 0

    for state, count in top.items():
        hub = HUB_MAP.get(str(state).upper(), "Regional Hub")
        pct = round((count / total * 100.0) if total else 0.0, 1)

        if str(state).upper() in select_states:
            network = "Select Network"
            select_count += count
        else:
            network = "National Network"
            national_count += count

        rows.append({
            "state": state,
            "count": int(count),
            "pct": pct,
            "hub": hub,
            "network": network
        })

    # Calculate network percentages
    total_network = select_count + national_count
    select_pct = round(select_count / total_network * 100.0, 1) if total_network > 0 else 0
    national_pct = round(national_count / total_network * 100.0, 1) if total_network > 0 else 0

    return {
        "top_states": rows,
        "select_network_pct": select_pct,
        "national_network_pct": national_pct
    }


def compute_zones(df: pd.DataFrame) -> Dict[str, Any]:
    """Compute zone distribution"""
    d = df.copy()

    if "Calculated Zone" not in d.columns:
        return {"zone_distribution": [], "regional_percentage": None, "cross_country_percentage": None}

    d["Calculated Zone"] = _coerce_numeric(d["Calculated Zone"])
    d = d.dropna(subset=["Calculated Zone"]).copy()
    d["Calculated Zone"] = d["Calculated Zone"].astype(int)

    total = len(d)
    dist_rows = []

    for z in range(1, 9):
        zdf = d[d["Calculated Zone"] == z]
        count = len(zdf)
        if count > 0:
            pct = round((count / total * 100.0) if total else 0.0, 1)

            # Calculate average transit for zone
            if "Days In Transit" in zdf.columns:
                transit_clean = _coerce_numeric(zdf["Days In Transit"]).dropna()
                avg_transit = round(float(transit_clean.mean()), 2) if len(transit_clean) > 0 else 0.0
            else:
                avg_transit = 0.0

            dist_rows.append({
                "zone": int(z),
                "count": int(count),
                "pct": pct,
                "avg_transit": avg_transit
            })

    # Calculate regional vs cross-country
    regional = int(d[(d["Calculated Zone"] >= 1) & (d["Calculated Zone"] <= 4)].shape[0])
    cross = int(d[(d["Calculated Zone"] >= 5) & (d["Calculated Zone"] <= 8)].shape[0])
    denom = regional + cross
    regional_pct = round(regional / denom * 100.0, 1) if denom else 0
    cross_pct = round(cross / denom * 100.0, 1) if denom else 0

    return {
        "zone_distribution": dist_rows,
        "regional_percentage": regional_pct,
        "cross_country_percentage": cross_pct
    }


def compute_ops(df: pd.DataFrame) -> Dict[str, Any]:
    """Compute operational metrics"""
    total = len(df)
    delivered = int((df["Delivered Status"].astype(str).str.lower() == "delivered").sum())
    in_transit = total - delivered  # Simplified - anything not delivered is in-transit

    # Calculate daily average (22 business days per month)
    daily_avg = round(total / 22) if total > 0 else 0

    # Optimization opportunities
    opps = []
    if in_transit > 0:
        opps.append("Monitor in-transit shipments approaching SLA window")
    if delivered > 0:
        delivered_df = df[df["Delivered Status"].astype(str).str.lower() == "delivered"]
        if "Days In Transit" in delivered_df.columns:
            transit_clean = _coerce_numeric(delivered_df["Days In Transit"]).dropna()
            if len(transit_clean) > 0:
                pct_over_5 = ((transit_clean > 5).sum() / len(transit_clean) * 100)
                if pct_over_5 > 10:
                    opps.append(f"Review {pct_over_5:.0f}% of shipments exceeding 5-day window")

    return {
        "total_volume": total,
        "delivered": delivered,
        "in_transit": in_transit,
        "daily_average": daily_avg,
        "optimization_opportunities": opps
    }


def split_in_transit_window(df: pd.DataFrame, service_level: str) -> pd.DataFrame:
    """Identify in-transit shipments and check if within SLA window"""
    window = SLA_WINDOWS.get(service_level, 5)
    today = pd.Timestamp.today().normalize()

    d = df.copy()

    # Try multiple date column names
    date_cols = ["Request Date", "ShipDate", "Ship Date", "Start Date"]
    date_col_used = None

    for col in date_cols:
        if col in d.columns:
            d["ShipDate"] = _parse_date(d[col])
            date_col_used = col
            break

    if date_col_used is None:
        # No ship date found, return empty
        return pd.DataFrame()

    # Calculate days since ship
    d["Days Since Ship"] = (today - d["ShipDate"]).dt.days

    # Filter for in-transit only
    in_transit_mask = d["Delivered Status"].astype(str).str.lower() != "delivered"
    in_transit_df = d.loc[in_transit_mask].copy()

    if len(in_transit_df) > 0:
        in_transit_df["Within SLA Window"] = in_transit_df["Days Since Ship"].apply(
            lambda x: "Yes" if pd.notnull(x) and x <= window else "No" if pd.notnull(x) else "Unknown"
        )

    return in_transit_df


# -----------------------------
# Excel writer helpers
# -----------------------------

def _hex_to_rgb(hex_color: str) -> Tuple[int, int, int]:
    """Convert HEX to RGB"""
    h = hex_color.strip().lstrip('#')
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))


def _rgb_to_cmyk(r: int, g: int, b: int) -> Tuple[int, int, int, int]:
    """Convert RGB to CMYK"""
    if (r, g, b) == (0, 0, 0):
        return (0, 0, 0, 100)
    r_, g_, b_ = r/255.0, g/255.0, b/255.0
    k = 1 - max(r_, g_, b_)
    c = (1 - r_ - k) / (1 - k) if (1 - k) else 0
    m = (1 - g_ - k) / (1 - k) if (1 - k) else 0
    y = (1 - b_ - k) / (1 - k) if (1 - k) else 0
    return (int(round(c*100)), int(round(m*100)), int(round(y*100)), int(round(k*100)))


def _style_table(ws, header_row: int = 1):
    """Apply FirstMile styling to table"""
    # Style header
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTER
        cell.border = BORDER

    # Style body
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell.alignment = CENTER
                cell.border = BORDER

    # Add auto filter if there's data
    if ws.max_row > header_row:
        ws.auto_filter.ref = ws.dimensions


def _autosize(ws):
    """Auto-size columns based on content"""
    for column in ws.columns:
        max_length = 0
        column = list(column)
        column_letter = None

        for cell in column:
            try:
                # Skip merged cells
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass

        if column_letter and max_length > 0:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


def _add_color_row(ws, name: str, hex_color: str, row: int):
    """Add a color swatch row to Brand Style Guide"""
    # Swatch
    swatch = ws.cell(row=row, column=1, value=" ")
    swatch.fill = PatternFill(start_color=hex_color.lstrip('#'), end_color=hex_color.lstrip('#'), fill_type="solid")
    swatch.border = BORDER

    # Name & HEX
    ws.cell(row=row, column=2, value=name)
    ws.cell(row=row, column=3, value=f"#{hex_color.lstrip('#').upper()}")

    # RGB & CMYK
    R, G, B = _hex_to_rgb(hex_color)
    C, M, Y, K = _rgb_to_cmyk(R, G, B)
    ws.cell(row=row, column=4, value=f"{R}, {G}, {B}")
    ws.cell(row=row, column=5, value=f"{C}, {M}, {Y}, {K}")


# -----------------------------
# Main Excel writer
# -----------------------------

def write_excel(
    results: Dict[str, Any],
    delivered_df: pd.DataFrame,
    in_transit_df: pd.DataFrame,
    customer_name: str,
    report_period: str,
    service_level: str,
    output_path: Optional[str] = None
) -> str:
    """Generate the complete Excel report"""

    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # 1) Executive Summary
    ws = wb.create_sheet("Executive Summary")
    rows = [
        ["Executive Summary"],
        ["Metric", "Value"],
        ["Customer", customer_name],
        ["Report Period", report_period],
        ["Carrier", "FirstMile"],
        ["Ship Method", service_level],
        ["SLA Window (days)", results["sla"]["sla_window"]],
        ["Total Volume", results["ops"]["total_volume"]],
        ["Delivered Volume", results["sla"]["total_delivered"]],
        ["In-Transit Volume", results["ops"]["in_transit"]],
        ["SLA Compliance (%)", f"{results['sla']['compliance_percentage']}%"],
        ["Performance Status", results["sla"]["performance_status"]]
    ]

    for r in rows:
        ws.append(r)

    # Merge title
    ws.merge_cells("A1:B1")
    ws["A1"].font = Font(size=14, bold=True, color=PRIMARY_BLUE)
    ws["A1"].alignment = CENTER

    _style_table(ws, header_row=2)
    _autosize(ws)

    # 2) SLA Compliance
    ws = wb.create_sheet("SLA Compliance")
    sla_df = pd.DataFrame({
        "Service Level": [results["sla"]["service_level"]],
        "SLA Window (days)": [results["sla"]["sla_window"]],
        "Delivered (count)": [results["sla"]["total_delivered"]],
        "Within SLA (count)": [results["sla"]["within_sla"]],
        "Exceeding SLA (count)": [results["sla"]["exceeding_sla"]],
        "Compliance %": [results["sla"]["compliance_percentage"]],
        "Status": [results["sla"]["performance_status"]]
    })

    for r in dataframe_to_rows(sla_df, index=False, header=True):
        ws.append(r)

    _style_table(ws)

    # Add conditional formatting on Compliance %
    if ws.max_row > 1:
        ws.conditional_formatting.add(
            f"F2:F{ws.max_row}",
            ColorScaleRule(
                start_type='num', start_value=80, start_color='FFC7CE',
                mid_type='num', mid_value=92, mid_color='FFEB84',
                end_type='num', end_value=100, end_color='C6EFCE'
            )
        )

    _autosize(ws)

    # 3) Transit Performance
    ws = wb.create_sheet("Transit Performance")

    # Daily distribution
    dist_df = pd.DataFrame(results["transit"]["daily_distribution"])
    if not dist_df.empty:
        dist_df.columns = ["Day", "Count", "%", "Cumulative %"]
        for r in dataframe_to_rows(dist_df, index=False, header=True):
            ws.append(r)

    # Statistics
    start_row = ws.max_row + 2 if ws.max_row > 0 else 1
    ws.cell(row=start_row, column=1, value="Statistical Summary")
    ws.cell(row=start_row, column=1).font = Font(bold=True, color=PRIMARY_BLUE)

    stats_data = [
        ["Metric", "Value"],
        ["Average (days)", results["transit"]["statistics"]["average"]],
        ["Median (days)", results["transit"]["statistics"]["median"]],
        ["90th Percentile (days)", results["transit"]["statistics"]["p90"]],
        ["95th Percentile (days)", results["transit"]["statistics"]["p95"]]
    ]

    for i, row_data in enumerate(stats_data, start=start_row+1):
        for j, value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=value)

    _style_table(ws)
    _autosize(ws)

    # 4) Geographic Distribution
    ws = wb.create_sheet("Geographic Distribution")

    if results["geography"]["top_states"]:
        geo_df = pd.DataFrame(results["geography"]["top_states"])
        geo_df.columns = ["State", "Volume", "%", "Hub Assignment", "Network"]

        for r in dataframe_to_rows(geo_df, index=False, header=True):
            ws.append(r)

        # Network summary
        sr = ws.max_row + 2
        ws.cell(row=sr, column=1, value="Network Distribution")
        ws.cell(row=sr, column=1).font = Font(bold=True, color=PRIMARY_BLUE)
        ws.cell(row=sr+1, column=1, value="Select Network (%)")
        ws.cell(row=sr+1, column=2, value=f"{results['geography']['select_network_pct']}%")
        ws.cell(row=sr+2, column=1, value="National Network (%)")
        ws.cell(row=sr+2, column=2, value=f"{results['geography']['national_network_pct']}%")

    _style_table(ws)
    _autosize(ws)

    # 5) Zone Analysis
    ws = wb.create_sheet("Zone Analysis")

    if results["zones"]["zone_distribution"]:
        zone_df = pd.DataFrame(results["zones"]["zone_distribution"])
        zone_df.columns = ["Zone", "Volume", "%", "Avg Transit (days)"]

        for r in dataframe_to_rows(zone_df, index=False, header=True):
            ws.append(r)

        # Summary
        sr = ws.max_row + 2
        ws.cell(row=sr, column=1, value="Zone Summary")
        ws.cell(row=sr, column=1).font = Font(bold=True, color=PRIMARY_BLUE)
        ws.cell(row=sr+1, column=1, value="Regional (Zones 1-4) %")
        ws.cell(row=sr+1, column=2, value=f"{results['zones']['regional_percentage']}%")
        ws.cell(row=sr+2, column=1, value="Cross-Country (Zones 5-8) %")
        ws.cell(row=sr+2, column=2, value=f"{results['zones']['cross_country_percentage']}%")

    _style_table(ws)
    _autosize(ws)

    # 6) Operational Metrics
    ws = wb.create_sheet("Operational Metrics")
    ops_df = pd.DataFrame({
        "Metric": ["Total Volume", "Delivered", "In Transit", "Daily Average"],
        "Value": [
            results["ops"]["total_volume"],
            results["ops"]["delivered"],
            results["ops"]["in_transit"],
            results["ops"]["daily_average"]
        ]
    })

    for r in dataframe_to_rows(ops_df, index=False, header=True):
        ws.append(r)

    # Opportunities
    if results["ops"]["optimization_opportunities"]:
        sr = ws.max_row + 2
        ws.cell(row=sr, column=1, value="Optimization Opportunities")
        ws.cell(row=sr, column=1).font = Font(bold=True, color=PRIMARY_BLUE)

        for i, opp in enumerate(results["ops"]["optimization_opportunities"], start=1):
            ws.cell(row=sr + i, column=1, value=f"• {opp}")
            ws.merge_cells(start_row=sr + i, start_column=1, end_row=sr + i, end_column=2)
            ws[f"A{sr + i}"].alignment = LEFT

    _style_table(ws)
    _autosize(ws)

    # 7) In-Transit Detail
    ws = wb.create_sheet("In-Transit Detail")

    if not in_transit_df.empty:
        # Select relevant columns
        cols_to_show = []
        for col in ["Request Date", "ShipDate", "Delivered Status", "Destination State",
                    "Calculated Zone", "Days Since Ship", "Within SLA Window"]:
            if col in in_transit_df.columns:
                cols_to_show.append(col)

        if cols_to_show:
            detail_df = in_transit_df[cols_to_show].head(100)  # Limit to 100 rows
            for r in dataframe_to_rows(detail_df, index=False, header=True):
                ws.append(r)
        else:
            ws.append(["No detailed data available for in-transit shipments"])
    else:
        ws.append(["All shipments are delivered for the report period"])

    _style_table(ws)
    _autosize(ws)

    # 8) Notes & Assumptions
    ws = wb.create_sheet("Notes & Assumptions")
    notes = [
        ["Note #", "Detail"],
        ["1.", "FirstMile is the carrier. Xparcel is the ship method."],
        ["2.", "SLA is calculated on Delivered shipments only."],
        ["3.", "In-Transit shipments are excluded from SLA calculations."],
        ["4.", "Service windows: Priority (1-3 days), Expedited (2-5 days), Ground (3-8 days)."],
        ["5.", "Zone definitions: Regional = Zones 1-4; Cross-Country = Zones 5-8."],
        ["6.", "Report generated: " + datetime.now().strftime("%B %d, %Y at %I:%M %p")],
        ["7.", "Data quality: Any missing or invalid values are excluded from calculations."]
    ]

    for r in notes:
        ws.append(r)

    _style_table(ws)
    _autosize(ws)

    # 9) Brand Style Guide
    ws = wb.create_sheet("Brand Style Guide")
    ws.append(["Swatch", "Name", "HEX", "RGB", "CMYK"])
    _style_table(ws)

    # Add color swatches
    _add_color_row(ws, "Primary Blue", PRIMARY_BLUE, 2)
    _add_color_row(ws, "White", "FFFFFF", 3)
    _add_color_row(ws, "Light Border", LIGHT_BORDER, 4)
    _add_color_row(ws, "Compliance Green", "C6EFCE", 5)
    _add_color_row(ws, "Compliance Yellow", "FFEB84", 6)
    _add_color_row(ws, "Compliance Red", "FFC7CE", 7)

    # Add note
    ws.append([])
    ws.append(["Note", "Color values are provided for brand consistency. CMYK values are computed from HEX."])

    _autosize(ws)

    # Generate filename
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    safe_customer = "".join([c for c in customer_name if c.isalnum() or c in (" ", "_", "-")]).strip().replace(" ", "_")
    fname = f"FirstMile_Xparcel_Performance_{safe_customer}_{ts}.xlsx"

    if output_path:
        out_file = os.path.join(output_path, fname)
    else:
        out_file = fname

    wb.save(out_file)
    return out_file


# -----------------------------
# Main Orchestration
# -----------------------------

def run_orchestrator(
    data_file: str,
    customer_name: str,
    report_period: str,
    service_level: str,
    output_path: Optional[str] = None
) -> Tuple[str, Dict[str, Any]]:
    """Main orchestration function"""

    print(f"[LOADING] Reading {data_file}...")
    df = pd.read_excel(data_file)

    print(f"[INFO] Total records: {len(df)}")
    print(f"[INFO] Columns: {', '.join(df.columns[:5])}...")

    _validate_columns(df)

    # Clean data types
    if "Days In Transit" in df.columns:
        df["Days In Transit"] = _coerce_numeric(df["Days In Transit"])

    # Parse dates - try multiple column names
    date_cols = ["Request Date", "ShipDate", "Ship Date", "Start Date"]
    for col in date_cols:
        if col in df.columns:
            df["ShipDate"] = _parse_date(df[col])
            break

    # Filter delivered for SLA calculations
    delivered_df = df[df["Delivered Status"].astype(str).str.lower() == "delivered"].copy()
    print(f"[INFO] Delivered packages: {len(delivered_df)}")

    # Compute all analytics
    print("[COMPUTING] Running analytics...")
    sla = compute_sla(df, service_level)
    transit = compute_transit(delivered_df)
    geo = compute_geography(df)
    zones = compute_zones(df)
    ops = compute_ops(df)

    # Get in-transit details
    in_transit_df = split_in_transit_window(df, service_level)
    print(f"[INFO] In-transit packages: {len(in_transit_df)}")

    results = {
        "sla": sla,
        "transit": transit,
        "geography": geo,
        "zones": zones,
        "ops": ops
    }

    # Generate Excel
    print("[GENERATING] Creating Excel report...")
    out_file = write_excel(
        results=results,
        delivered_df=delivered_df,
        in_transit_df=in_transit_df,
        customer_name=customer_name,
        report_period=report_period,
        service_level=service_level,
        output_path=output_path
    )

    print(f"[SUCCESS] Report saved: {out_file}")

    # Print summary
    print("\n" + "="*60)
    print("REPORT SUMMARY")
    print("="*60)
    print(f"Customer: {customer_name}")
    print(f"Period: {report_period}")
    print(f"Service: {service_level}")
    print(f"Total Volume: {ops['total_volume']}")
    print(f"SLA Compliance: {sla['compliance_percentage']}% ({sla['performance_status']})")
    print(f"Average Transit: {transit['statistics']['average']} days")
    print("="*60)

    return out_file, results


# -----------------------------
# CLI Entry Point
# -----------------------------

if __name__ == "__main__":
    # Run for JM Group
    output_file, results = run_orchestrator(
        data_file="Domestic_Tracking_JM_Group_Aug_2025.xlsx",
        customer_name="JM Group of NY",
        report_period="August 7 - September 19, 2025",
        service_level="Xparcel Expedited",
        output_path=None
    )

    print(f"\n[COMPLETE] FirstMile Xparcel Performance Report generated successfully!")
    print(f"[FILE] {output_file}")