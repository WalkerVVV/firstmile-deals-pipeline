#!/usr/bin/env python3
"""
JOSH'S FROGS - DRY GOODS ANALYSIS v2.0
Based on AI Agent Blueprint v2
Hierarchical pivot table analysis matching exact output format
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from collections import OrderedDict

print("="*80)
print("JOSH'S FROGS - DRY GOODS ANALYSIS v2.0")
print("Hierarchical Pivot Table Analysis")
print("="*80)

# ============================================
# SUB-AGENT 1: HIERARCHICAL CSV PARSER
# ============================================

def parse_hierarchical_csv(file_path):
    """Parse CSV with carrier → service hierarchy"""
    print("\n[Sub-Agent 1] Parsing hierarchical CSV...")

    carriers = OrderedDict()
    current_carrier = None
    grand_total = None

    with open(file_path, 'r') as f:
        lines = f.readlines()

    for line in lines[1:]:  # Skip header
        if not line.strip():
            continue

        parts = line.strip().split(',')
        if len(parts) < 3:
            continue

        row_label = parts[0]

        try:
            volume = int(parts[1])
            avg_weight = float(parts[2])
        except:
            continue

        if row_label == 'Grand Total':
            grand_total = {'volume': volume, 'avg_weight': avg_weight}
        elif row_label.startswith('  '):  # Service (child) - has leading spaces
            service_name = row_label.strip()
            if current_carrier and current_carrier in carriers:
                carriers[current_carrier]['services'][service_name] = {
                    'volume': volume,
                    'avg_weight': avg_weight
                }
        else:  # Carrier (parent)
            current_carrier = row_label.strip()
            carriers[current_carrier] = {
                'volume': volume,
                'avg_weight': avg_weight,
                'services': OrderedDict()
            }

    print(f"   Parsed {len(carriers)} carriers")
    print(f"   Grand Total: {grand_total['volume']:,} shipments")

    return {'carriers': carriers, 'grand_total': grand_total}

# ============================================
# SUB-AGENT 2: SERVICE CLASSIFIER
# ============================================

def classify_services(data):
    """Classify DRY_GOODS vs LIVE_ANIMALS"""
    print("\n[Sub-Agent 2] Classifying services...")

    live_animal_keywords = [
        'OVERNIGHT', 'EXPRESS_SAVER', 'NEXT_DAY', 'SECOND_DAY', 'TWO_DAY',
        'PRIORITY_OVERNIGHT', 'FIRST_OVERNIGHT', 'STANDARD_OVERNIGHT'
    ]

    dry_goods_count = 0
    live_animal_count = 0

    for carrier, carrier_data in data['carriers'].items():
        for service, service_data in carrier_data['services'].items():
            service_upper = service.upper()
            is_live_animal = any(kw in service_upper for kw in live_animal_keywords)
            service_data['category'] = 'LIVE_ANIMALS' if is_live_animal else 'DRY_GOODS'

            if is_live_animal:
                live_animal_count += service_data['volume']
            else:
                dry_goods_count += service_data['volume']

    print(f"   Dry Goods: {dry_goods_count:,} shipments")
    print(f"   Live Animals (excluded): {live_animal_count:,} shipments")

    return data

# ============================================
# SUB-AGENT 3: CURRENT COST CALCULATOR
# ============================================

def calculate_current_costs(data):
    """Calculate current carrier costs using weighted zone averages"""
    print("\n[Sub-Agent 3] Calculating current carrier costs...")

    # Base rates by service
    base_rates = {
        'USPS_GROUND_ADVANTAGE': 5.25,
        'USPS_PRIORITY_MAIL': 7.95,
        'USPS_FIRST_CLASS_MAIL': 4.50,
        'USPS_PARCEL_SELECT': 5.80,
        'USPS_GROUND': 6.00,
        'UPS_GROUND': 6.50,
        'UPS_SURE_POST': 5.80,
        'UPS_THREE_DAY_SELECT': 12.50,
        'FEDEX_GROUND': 6.45,
        'FEDEX_HOME_DELIVERY': 6.75,
        'DHL_GROUND': 6.20,
        'DHL_EXPEDITED': 8.50,
        'DHL_EXPEDITED_MAX': 8.20,
        'GLS_GROUND_PRIORITY': 7.00,
    }

    # Typical ecommerce zone distribution
    zone_dist = {2: 0.12, 3: 0.22, 4: 0.28, 5: 0.18, 6: 0.10, 7: 0.08, 8: 0.02}

    for carrier_data in data['carriers'].values():
        for service, service_data in carrier_data['services'].items():
            if service_data.get('category') == 'LIVE_ANIMALS':
                service_data['avg_current_cost'] = 0.0  # Not calculating for live animals
                continue

            avg_weight = service_data['avg_weight']
            base = base_rates.get(service, 6.50)

            # Calculate weighted average across zones
            total_cost = 0
            for zone, zone_weight in zone_dist.items():
                zone_mult = 1 + (zone - 1) * 0.08
                weight_mult = 1 + np.log1p(avg_weight) * 0.15
                fuel_mult = 1.12  # 12% fuel and fees
                cost = base * zone_mult * weight_mult * fuel_mult
                total_cost += cost * zone_weight

            service_data['avg_current_cost'] = round(total_cost, 2)

    print("   Current costs calculated for all dry goods services")
    return data

# ============================================
# SUB-AGENT 4: FIRSTMILE RATE CALCULATOR
# ============================================

def calculate_firstmile_rates(data):
    """Calculate FirstMile Xparcel rates using weighted zone averages"""
    print("\n[Sub-Agent 4] Calculating FirstMile Xparcel rates...")

    # Xparcel Ground base rates by zone
    xparcel_ground_base = {1: 3.73, 2: 3.79, 3: 3.80, 4: 3.89,
                          5: 3.94, 6: 4.02, 7: 4.09, 8: 4.24}

    # Xparcel Priority base rates by zone (for expedited services)
    xparcel_priority_base = {1: 3.94, 2: 3.99, 3: 4.01, 4: 4.10,
                            5: 4.15, 6: 4.24, 7: 4.31, 8: 4.48}

    zone_dist = {2: 0.12, 3: 0.22, 4: 0.28, 5: 0.18, 6: 0.10, 7: 0.08, 8: 0.02}

    for carrier_data in data['carriers'].values():
        for service, service_data in carrier_data['services'].items():
            if service_data.get('category') == 'LIVE_ANIMALS':
                service_data['avg_xparcel_cost'] = 0.0
                continue

            avg_weight = service_data['avg_weight']

            # Determine if priority or ground
            is_expedited = 'EXPEDITED' in service.upper() or 'PRIORITY' in service.upper()
            base_rates = xparcel_priority_base if is_expedited else xparcel_ground_base

            # Weight tier addon
            if avg_weight <= 0.0625:
                addon = 0.00
            elif avg_weight <= 0.25:
                addon = 0.10
            elif avg_weight <= 0.5:
                addon = 0.25
            elif avg_weight <= 1.0:
                addon = 0.50
            elif avg_weight <= 2.0:
                addon = 1.30
            elif avg_weight <= 5.0:
                addon = 4.10
            elif avg_weight <= 10.0:
                addon = 7.50
            else:
                addon = 7.50 + (avg_weight - 10) * 0.35

            # Calculate weighted average across zones
            total_cost = 0
            for zone, zone_weight in zone_dist.items():
                base = base_rates[zone]
                cost = base + addon
                total_cost += cost * zone_weight

            service_data['avg_xparcel_cost'] = round(total_cost, 2)
            service_data['savings_dollars'] = round(service_data['avg_current_cost'] - service_data['avg_xparcel_cost'], 2)

            if service_data['avg_current_cost'] > 0:
                service_data['savings_percent'] = round((service_data['savings_dollars'] / service_data['avg_current_cost']) * 100, 1)
            else:
                service_data['savings_percent'] = 0.0

    print("   FirstMile rates calculated for all dry goods services")
    return data

# ============================================
# SUB-AGENT 5: AGGREGATION ENGINE
# ============================================

def aggregate_carriers(data):
    """Calculate carrier-level and grand total rollups using volume-weighted averages"""
    print("\n[Sub-Agent 5] Aggregating carrier-level totals...")

    # Filter to dry goods only
    for carrier, carrier_data in data['carriers'].items():
        dry_goods_services = {k: v for k, v in carrier_data['services'].items()
                             if v.get('category') == 'DRY_GOODS'}

        if not dry_goods_services:
            carrier_data['dry_goods_volume'] = 0
            carrier_data['dry_goods_avg_weight'] = 0
            carrier_data['dry_goods_avg_current_cost'] = 0
            carrier_data['dry_goods_avg_xparcel_cost'] = 0
            carrier_data['dry_goods_savings_dollars'] = 0
            carrier_data['dry_goods_savings_percent'] = 0
            continue

        # Calculate volume-weighted averages for carrier
        total_volume = sum(s['volume'] for s in dry_goods_services.values())

        weighted_weight = sum(s['volume'] * s['avg_weight'] for s in dry_goods_services.values()) / total_volume
        weighted_current = sum(s['volume'] * s['avg_current_cost'] for s in dry_goods_services.values()) / total_volume
        weighted_xparcel = sum(s['volume'] * s['avg_xparcel_cost'] for s in dry_goods_services.values()) / total_volume

        carrier_data['dry_goods_volume'] = total_volume
        carrier_data['dry_goods_avg_weight'] = round(weighted_weight, 2)
        carrier_data['dry_goods_avg_current_cost'] = round(weighted_current, 2)
        carrier_data['dry_goods_avg_xparcel_cost'] = round(weighted_xparcel, 2)
        carrier_data['dry_goods_savings_dollars'] = round(weighted_current - weighted_xparcel, 2)

        if weighted_current > 0:
            carrier_data['dry_goods_savings_percent'] = round((carrier_data['dry_goods_savings_dollars'] / weighted_current) * 100, 1)
        else:
            carrier_data['dry_goods_savings_percent'] = 0.0

    # Calculate grand total for dry goods
    all_carriers_with_dry = [c for c in data['carriers'].values() if c['dry_goods_volume'] > 0]

    if all_carriers_with_dry:
        total_volume = sum(c['dry_goods_volume'] for c in all_carriers_with_dry)

        weighted_weight = sum(c['dry_goods_volume'] * c['dry_goods_avg_weight'] for c in all_carriers_with_dry) / total_volume
        weighted_current = sum(c['dry_goods_volume'] * c['dry_goods_avg_current_cost'] for c in all_carriers_with_dry) / total_volume
        weighted_xparcel = sum(c['dry_goods_volume'] * c['dry_goods_avg_xparcel_cost'] for c in all_carriers_with_dry) / total_volume

        data['dry_goods_grand_total'] = {
            'volume': total_volume,
            'avg_weight': round(weighted_weight, 2),
            'avg_current_cost': round(weighted_current, 2),
            'avg_xparcel_cost': round(weighted_xparcel, 2),
            'savings_dollars': round(weighted_current - weighted_xparcel, 2),
            'savings_percent': round(((weighted_current - weighted_xparcel) / weighted_current) * 100, 1) if weighted_current > 0 else 0.0
        }

    print(f"   Aggregated {len(all_carriers_with_dry)} carriers with dry goods")
    print(f"   Dry Goods Grand Total: {data['dry_goods_grand_total']['volume']:,} shipments")

    return data

# ============================================
# SUB-AGENT 6: EXCEL FORMATTER
# ============================================

def create_excel_output(data, filename):
    """Create Excel workbook matching exact target format"""
    print("\n[Sub-Agent 6] Generating Excel workbook...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Dry Goods Analysis"

    # Define styles
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, size=11)
    header_style.alignment = Alignment(horizontal="center", vertical="center")

    carrier_style = NamedStyle(name="carrier_style")
    carrier_style.font = Font(bold=True, size=11)
    carrier_style.fill = PatternFill(start_color="D0E4F5", end_color="D0E4F5", fill_type="solid")
    carrier_style.alignment = Alignment(horizontal="left")

    service_style = NamedStyle(name="service_style")
    service_style.font = Font(size=10)
    service_style.alignment = Alignment(horizontal="left", indent=1)

    total_style = NamedStyle(name="total_style")
    total_style.font = Font(bold=True, size=11)
    total_style.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    currency_style = NamedStyle(name="currency_style")
    currency_style.number_format = '$#,##0.00'

    percent_style = NamedStyle(name="percent_style")
    percent_style.number_format = '0.0%'

    # Add styles to workbook
    for style in [header_style, carrier_style, service_style, total_style, currency_style, percent_style]:
        if style.name not in wb.named_styles:
            wb.add_named_style(style)

    # Headers
    headers = ["Row Labels", "Count of Number", "Average of Cost", "Average of Weight",
               "FirstMile Rate", "Savings $", "Savings %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.style = "header_style"

    # Data rows
    row = 2
    for carrier, carrier_data in data['carriers'].items():
        if carrier_data['dry_goods_volume'] == 0:
            continue  # Skip carriers with no dry goods

        # Carrier row (parent)
        ws.cell(row, 1, carrier).style = "carrier_style"
        ws.cell(row, 2, carrier_data['dry_goods_volume']).style = "carrier_style"
        ws.cell(row, 3, carrier_data['dry_goods_avg_current_cost']).style = "carrier_style"
        ws.cell(row, 3).number_format = '$#,##0.00'
        ws.cell(row, 4, carrier_data['dry_goods_avg_weight']).style = "carrier_style"
        ws.cell(row, 5, carrier_data['dry_goods_avg_xparcel_cost']).style = "carrier_style"
        ws.cell(row, 5).number_format = '$#,##0.00'
        ws.cell(row, 6, carrier_data['dry_goods_savings_dollars']).style = "carrier_style"
        ws.cell(row, 6).number_format = '$#,##0.00'
        ws.cell(row, 7, carrier_data['dry_goods_savings_percent']/100).style = "carrier_style"
        ws.cell(row, 7).number_format = '0.0%'
        row += 1

        # Service rows (children)
        for service, service_data in carrier_data['services'].items():
            if service_data.get('category') == 'LIVE_ANIMALS':
                continue  # Skip live animal services

            ws.cell(row, 1, f"  {service}").style = "service_style"
            ws.cell(row, 2, service_data['volume'])
            ws.cell(row, 3, service_data['avg_current_cost']).number_format = '$#,##0.00'
            ws.cell(row, 4, service_data['avg_weight'])
            ws.cell(row, 5, service_data['avg_xparcel_cost']).number_format = '$#,##0.00'
            ws.cell(row, 6, service_data['savings_dollars']).number_format = '$#,##0.00'
            ws.cell(row, 7, service_data['savings_percent']/100).number_format = '0.0%'
            row += 1

    # Grand Total row
    gt = data['dry_goods_grand_total']
    ws.cell(row, 1, "Grand Total").style = "total_style"
    ws.cell(row, 2, gt['volume']).style = "total_style"
    ws.cell(row, 3, gt['avg_current_cost']).style = "total_style"
    ws.cell(row, 3).number_format = '$#,##0.00'
    ws.cell(row, 4, gt['avg_weight']).style = "total_style"
    ws.cell(row, 5, gt['avg_xparcel_cost']).style = "total_style"
    ws.cell(row, 5).number_format = '$#,##0.00'
    ws.cell(row, 6, gt['savings_dollars']).style = "total_style"
    ws.cell(row, 6).number_format = '$#,##0.00'
    ws.cell(row, 7, gt['savings_percent']/100).style = "total_style"
    ws.cell(row, 7).number_format = '0.0%'

    # Auto-adjust column widths
    for col in range(1, 8):
        max_length = 0
        column_letter = get_column_letter(col)
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 35)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filename)
    print(f"   Excel workbook saved: {filename}")
    return filename

# ============================================
# MAIN EXECUTION
# ============================================

if __name__ == "__main__":
    # Execute pipeline
    csv_file = "247bef97-8663-431e-b2f5-dd2ca243633d.csv"

    # Sub-Agent 1: Parse
    data = parse_hierarchical_csv(csv_file)

    # Sub-Agent 2: Classify
    data = classify_services(data)

    # Sub-Agent 3: Calculate current costs
    data = calculate_current_costs(data)

    # Sub-Agent 4: Calculate FirstMile rates
    data = calculate_firstmile_rates(data)

    # Sub-Agent 5: Aggregate
    data = aggregate_carriers(data)

    # Sub-Agent 6: Create Excel
    output_file = create_excel_output(data, "Joshs_Frogs_DRY_GOODS_Analysis_FINAL.xlsx")

    # Print summary
    print("\n" + "="*80)
    print("ANALYSIS COMPLETE")
    print("="*80)
    gt = data['dry_goods_grand_total']
    print(f"Total Dry Goods Volume: {gt['volume']:,} shipments")
    print(f"Average Weight: {gt['avg_weight']} lbs")
    print(f"Average Current Cost: ${gt['avg_current_cost']:.2f}")
    print(f"Average FirstMile Cost: ${gt['avg_xparcel_cost']:.2f}")
    print(f"Average Savings: ${gt['savings_dollars']:.2f} ({gt['savings_percent']:.1f}%)")
    print(f"Monthly Savings: ${gt['savings_dollars'] * gt['volume']:,.2f}")
    print(f"Annual Projection: ${gt['savings_dollars'] * gt['volume'] * 12:,.2f}")
    print("="*80)
    print(f"\nOutput file: {output_file}")