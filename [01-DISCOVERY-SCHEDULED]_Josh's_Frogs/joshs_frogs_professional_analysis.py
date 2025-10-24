#!/usr/bin/env python3
"""
Josh's Frogs - Professional FirstMile Xparcel Savings Analysis
Blueprint: Stackd Logistics format
Professional formatting for prospect presentation
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# FirstMile brand color
FM_BLUE = "366092"
FM_WHITE = "FFFFFF"
FM_GREEN = "00B050"
FM_LIGHT_GRAY = "F2F2F2"

print("=" * 80)
print("JOSH'S FROGS - PROFESSIONAL FIRSTMILE XPARCEL ANALYSIS")
print("=" * 80)
print()

# Load PLD data
print("Loading PLD data...")
df = pd.read_csv("Josh's Frogs PLD_with_costs_6_month_data.csv", low_memory=False)
print(f"Total shipments: {len(df):,}")
print()

# Zone calculation function
def get_zone_from_zip(dest_zip):
    """Calculate shipping zone from destination ZIP"""
    try:
        zip_str = str(dest_zip).split('-')[0].zfill(5)
        zip3 = int(zip_str[:3])

        if 150 <= zip3 <= 196: return 2
        elif (10 <= zip3 <= 149) or (197 <= zip3 <= 199): return 3
        elif (200 <= zip3 <= 289) or (300 <= zip3 <= 399): return 4
        elif (400 <= zip3 <= 499) or (600 <= zip3 <= 699): return 4
        elif (700 <= zip3 <= 799) or (500 <= zip3 <= 599): return 5
        elif (800 <= zip3 <= 879): return 6
        elif (850 <= zip3 <= 899) or (880 <= zip3 <= 884): return 7
        elif (900 <= zip3 <= 999): return 8
        else: return 5
    except:
        return None

# Calculate zones
df['Zone'] = df['Destination'].apply(get_zone_from_zip)

# Calculate billable weight
def calculate_billable_weight(weight):
    if pd.isna(weight): return None
    if weight <= 1: return weight
    return np.ceil(weight)

df['Billable_Weight'] = df['Weight'].apply(calculate_billable_weight)

# Weight tiers
def get_weight_tier(weight):
    if pd.isna(weight): return None
    if weight <= 0.25: return "0-4oz"
    elif weight <= 0.5: return "4-8oz"
    elif weight <= 1: return "8oz-1lb"
    elif weight <= 2: return "1-2lb"
    elif weight <= 5: return "2-5lb"
    elif weight <= 10: return "5-10lb"
    else: return "10lb+"

df['Weight_Tier'] = df['Billable_Weight'].apply(get_weight_tier)

# Serviceable filtering
EXCLUDED_SERVICES = [
    'FEDEX_FIRST_OVERNIGHT',
    'FEDEX_PRIORITY_OVERNIGHT',
    'FEDEX_STANDARD_OVERNIGHT',
    'UPS_NEXT_DAY_AIR',
    'UPS_NEXT_DAY_AIR_SAVER'
]

df['Serviceable'] = ~df['Service'].isin(EXCLUDED_SERVICES)
serviceable_df = df[df['Serviceable']]

# Calculate key metrics
total_shipments = len(df)
serviceable_shipments = len(serviceable_df)
monthly_shipments = total_shipments / 6
annual_shipments = total_shipments * 2

total_6mo_spend = serviceable_df['Cost'].sum()
monthly_spend = total_6mo_spend / 6
annual_spend = total_6mo_spend * 2
avg_cost = serviceable_df['Cost'].mean()

under_1lb = len(serviceable_df[serviceable_df['Billable_Weight'] <= 1])
under_1lb_pct = under_1lb / serviceable_shipments * 100
avg_weight = serviceable_df['Weight'].mean()

# FirstMile savings estimate (12%)
savings_rate = 0.12
xparcel_monthly = monthly_spend * (1 - savings_rate)
monthly_savings = monthly_spend * savings_rate
annual_savings = monthly_savings * 12

print(f"Serviceable: {serviceable_shipments:,} ({serviceable_shipments/total_shipments*100:.1f}%)")
print(f"Monthly Spend: ${monthly_spend:,.2f}")
print(f"Monthly Savings (12%): ${monthly_savings:,.2f}")
print()

# Create professional Excel workbook
print("Creating professional Excel workbook...")
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Helper functions for styling
def create_header_cell(ws, row, col, value, merge_to=None):
    """Create FirstMile blue header cell"""
    cell = ws.cell(row, col, value)
    cell.font = Font(name='Calibri', size=18, bold=True, color=FM_WHITE)
    cell.fill = PatternFill(start_color=FM_BLUE, end_color=FM_BLUE, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    return cell

def create_subheader(ws, row, col, value, merge_to=None):
    """Create section subheader"""
    cell = ws.cell(row, col, value)
    cell.font = Font(name='Calibri', size=14, bold=True, color=FM_WHITE)
    cell.fill = PatternFill(start_color=FM_BLUE, end_color=FM_BLUE, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    return cell

def create_table_header(ws, row, col, value):
    """Create table column header"""
    cell = ws.cell(row, col, value)
    cell.font = Font(name='Calibri', size=12, bold=True, color=FM_WHITE)
    cell.fill = PatternFill(start_color=FM_BLUE, end_color=FM_BLUE, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    return cell

def format_currency(ws, row, col, value):
    """Format cell as currency"""
    cell = ws.cell(row, col, value)
    cell.number_format = '$#,##0.00'
    cell.alignment = Alignment(horizontal='right')
    return cell

def format_percentage(ws, row, col, value):
    """Format cell as percentage"""
    cell = ws.cell(row, col, value)
    cell.number_format = '0.0%'
    cell.alignment = Alignment(horizontal='center')
    return cell

def format_number(ws, row, col, value):
    """Format cell as number with commas"""
    cell = ws.cell(row, col, value)
    cell.number_format = '#,##0'
    cell.alignment = Alignment(horizontal='right')
    return cell

# ============================================================================
# SHEET 1: EXECUTIVE SUMMARY
# ============================================================================
ws_summary = wb.create_sheet("Executive Summary")

# Main title
create_header_cell(ws_summary, 1, 1, "JOSH'S FROGS - FIRSTMILE XPARCEL SAVINGS ANALYSIS", merge_to=8)
ws_summary.row_dimensions[1].height = 30

# Analysis details
ws_summary.cell(3, 1, "Analysis Date:").font = Font(bold=True)
ws_summary.cell(3, 2, datetime.now().strftime("%B %d, %Y"))

ws_summary.cell(4, 1, "Data Period:").font = Font(bold=True)
ws_summary.cell(4, 2, "6 months (February - August 2025)")

ws_summary.cell(5, 1, "Total Shipments:").font = Font(bold=True)
format_number(ws_summary, 5, 2, total_shipments)

# Savings Summary section
create_subheader(ws_summary, 7, 1, "SAVINGS SUMMARY", merge_to=5)
create_subheader(ws_summary, 7, 7, "KEY METRICS", merge_to=8)

# Savings table headers
create_table_header(ws_summary, 9, 1, "Metric")
create_table_header(ws_summary, 9, 2, "Current")
create_table_header(ws_summary, 9, 3, "With FirstMile Xparcel")
create_table_header(ws_summary, 9, 4, "Monthly Savings")
create_table_header(ws_summary, 9, 5, "% Savings")

# Savings data
ws_summary.cell(10, 1, "Monthly Spend")
format_currency(ws_summary, 10, 2, monthly_spend)
format_currency(ws_summary, 10, 3, xparcel_monthly)
cell = format_currency(ws_summary, 10, 4, monthly_savings)
cell.font = Font(size=12, bold=True, color=FM_GREEN)
format_percentage(ws_summary, 10, 5, savings_rate)

ws_summary.cell(11, 1, "Annual Projection")
format_currency(ws_summary, 11, 2, annual_spend)
format_currency(ws_summary, 11, 3, annual_spend * (1 - savings_rate))
cell = format_currency(ws_summary, 11, 4, annual_savings)
cell.font = Font(size=12, bold=True, color=FM_GREEN)
format_percentage(ws_summary, 11, 5, savings_rate)

# Key metrics
ws_summary.cell(9, 7, "Total Shipments (6mo)").font = Font(bold=True)
format_number(ws_summary, 9, 8, total_shipments)

ws_summary.cell(10, 7, "Under 1 lb").font = Font(bold=True)
ws_summary.cell(10, 8, f"{under_1lb:,} ({under_1lb_pct:.1f}%)")

ws_summary.cell(11, 7, "Average Weight").font = Font(bold=True)
ws_summary.cell(11, 8, f"{avg_weight:.2f} lbs")

ws_summary.cell(12, 7, "Monthly Avg Volume").font = Font(bold=True)
format_number(ws_summary, 12, 8, monthly_shipments)

ws_summary.cell(13, 7, "Annual Projection").font = Font(bold=True)
format_number(ws_summary, 13, 8, annual_shipments)

# Column widths
ws_summary.column_dimensions['A'].width = 20
ws_summary.column_dimensions['B'].width = 15
ws_summary.column_dimensions['C'].width = 18
ws_summary.column_dimensions['D'].width = 16
ws_summary.column_dimensions['E'].width = 12
ws_summary.column_dimensions['G'].width = 20
ws_summary.column_dimensions['H'].width = 18

# ============================================================================
# SHEET 2: WEIGHT ANALYSIS
# ============================================================================
ws_weight = wb.create_sheet("Weight Analysis")

create_header_cell(ws_weight, 1, 1, "SAVINGS BY WEIGHT RANGE", merge_to=8)
ws_weight.row_dimensions[1].height = 30

# Table headers
headers = ["Weight Range", "Shipments", "% of Total", "Avg Current Cost",
           "Avg Xparcel Cost", "Avg Savings", "Savings %", "Total Savings"]
for col, header in enumerate(headers, 1):
    create_table_header(ws_weight, 3, col, header)

# Calculate weight analysis
weight_tiers = ["0-4oz", "4-8oz", "8oz-1lb", "1-2lb", "2-5lb", "5-10lb", "10lb+"]
row = 4
total_savings_weight = 0

for tier in weight_tiers:
    tier_data = serviceable_df[serviceable_df['Weight_Tier'] == tier]
    if len(tier_data) == 0:
        continue

    count = len(tier_data)
    pct = count / serviceable_shipments
    avg_cost = tier_data['Cost'].mean()
    avg_xparcel = avg_cost * (1 - savings_rate)
    avg_savings = avg_cost * savings_rate
    tier_savings = avg_savings * count
    total_savings_weight += tier_savings

    ws_weight.cell(row, 1, tier)
    format_number(ws_weight, row, 2, count)
    format_percentage(ws_weight, row, 3, pct)
    format_currency(ws_weight, row, 4, avg_cost)
    format_currency(ws_weight, row, 5, avg_xparcel)
    format_currency(ws_weight, row, 6, avg_savings)
    format_percentage(ws_weight, row, 7, savings_rate)
    format_currency(ws_weight, row, 8, tier_savings)

    row += 1

# Total row
ws_weight.cell(row, 1, "TOTAL").font = Font(bold=True)
format_number(ws_weight, row, 2, serviceable_shipments)
format_percentage(ws_weight, row, 3, 1.0)
format_currency(ws_weight, row, 4, avg_cost)
format_currency(ws_weight, row, 5, avg_cost * (1 - savings_rate))
format_currency(ws_weight, row, 6, avg_cost * savings_rate)
format_percentage(ws_weight, row, 7, savings_rate)
format_currency(ws_weight, row, 8, monthly_savings * 6)

# Column widths
for col in range(1, 9):
    ws_weight.column_dimensions[get_column_letter(col)].width = 16

# ============================================================================
# SHEET 3: ZONE ANALYSIS
# ============================================================================
ws_zone = wb.create_sheet("Zone Analysis")

create_header_cell(ws_zone, 1, 1, "SAVINGS BY DESTINATION ZONE", merge_to=7)
ws_zone.row_dimensions[1].height = 30

# Table headers
headers = ["Zone", "Shipments", "% of Total", "Avg Current Cost",
           "Avg Xparcel Cost", "Avg Savings/Pkg", "Total Savings"]
for col, header in enumerate(headers, 1):
    create_table_header(ws_zone, 3, col, header)

# Calculate zone analysis
zones_with_data = serviceable_df[serviceable_df['Zone'].notna()]
row = 4
total_savings_zone = 0

for zone in sorted(zones_with_data['Zone'].unique()):
    zone_data = zones_with_data[zones_with_data['Zone'] == zone]
    count = len(zone_data)
    pct = count / serviceable_shipments
    avg_cost = zone_data['Cost'].mean()
    avg_xparcel = avg_cost * (1 - savings_rate)
    avg_savings = avg_cost * savings_rate
    zone_savings = avg_savings * count
    total_savings_zone += zone_savings

    ws_zone.cell(row, 1, f"Zone {int(zone)}")
    format_number(ws_zone, row, 2, count)
    format_percentage(ws_zone, row, 3, pct)
    format_currency(ws_zone, row, 4, avg_cost)
    format_currency(ws_zone, row, 5, avg_xparcel)
    format_currency(ws_zone, row, 6, avg_savings)
    format_currency(ws_zone, row, 7, zone_savings)

    row += 1

# Column widths
for col in range(1, 8):
    ws_zone.column_dimensions[get_column_letter(col)].width = 16

# ============================================================================
# SHEET 4: CARRIER BREAKDOWN
# ============================================================================
ws_carrier = wb.create_sheet("Carrier Breakdown")

create_header_cell(ws_carrier, 1, 1, "CURRENT CARRIER ANALYSIS", merge_to=7)
ws_carrier.row_dimensions[1].height = 30

# Table headers
headers = ["Carrier", "Shipments", "% of Total", "6-Month Spend",
           "Monthly Spend", "Annual Projection", "Avg Cost/Pkg"]
for col, header in enumerate(headers, 1):
    create_table_header(ws_carrier, 3, col, header)

# Calculate carrier breakdown
carrier_summary = serviceable_df.groupby('Carrier').agg({
    'Number': 'count',
    'Cost': ['sum', 'mean']
}).round(2)
carrier_summary.columns = ['Count', 'Total_Cost', 'Avg_Cost']
carrier_summary = carrier_summary.sort_values('Count', ascending=False)

row = 4
for carrier, data in carrier_summary.iterrows():
    count = int(data['Count'])
    pct = count / serviceable_shipments
    total_cost = data['Total_Cost']
    monthly_cost = total_cost / 6
    annual_cost = total_cost * 2
    avg_cost = data['Avg_Cost']

    ws_carrier.cell(row, 1, carrier)
    format_number(ws_carrier, row, 2, count)
    format_percentage(ws_carrier, row, 3, pct)
    format_currency(ws_carrier, row, 4, total_cost)
    format_currency(ws_carrier, row, 5, monthly_cost)
    format_currency(ws_carrier, row, 6, annual_cost)
    format_currency(ws_carrier, row, 7, avg_cost)

    row += 1

# Column widths
for col in range(1, 8):
    ws_carrier.column_dimensions[get_column_letter(col)].width = 16

# ============================================================================
# SHEET 5: RATE CARDS (TOP SERVICES)
# ============================================================================
ws_rates = wb.create_sheet("Rate Cards")

create_header_cell(ws_rates, 1, 1, "INCUMBENT CARRIER RATE CARDS - ZONE × WEIGHT", merge_to=10)
ws_rates.row_dimensions[1].height = 30

current_row = 3

# Get top 5 services by volume
top_services = serviceable_df.groupby(['Carrier', 'Service']).size().sort_values(ascending=False).head(5)

for (carrier, service), count in top_services.items():
    # Service header
    ws_rates.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
    cell = ws_rates.cell(current_row, 1, f"{carrier} - {service} ({count:,} shipments)")
    cell.font = Font(size=12, bold=True, color=FM_WHITE)
    cell.fill = PatternFill(start_color=FM_BLUE, end_color=FM_BLUE, fill_type='solid')
    current_row += 1

    # Filter data
    service_data = serviceable_df[
        (serviceable_df['Carrier'] == carrier) &
        (serviceable_df['Service'] == service) &
        (serviceable_df['Zone'].notna()) &
        (serviceable_df['Weight_Tier'].notna())
    ]

    if len(service_data) < 50:
        current_row += 1
        continue

    # Create rate card
    rate_card = service_data.groupby(['Zone', 'Weight_Tier'])['Cost'].mean().unstack(fill_value=0).round(2)

    # Headers
    ws_rates.cell(current_row, 1, "Zone")
    for col, tier in enumerate(weight_tiers, 2):
        if tier in rate_card.columns:
            ws_rates.cell(current_row, col, tier)
    current_row += 1

    # Data
    for zone in sorted(rate_card.index):
        ws_rates.cell(current_row, 1, f"Zone {int(zone)}")
        for col, tier in enumerate(weight_tiers, 2):
            if tier in rate_card.columns:
                rate = rate_card.loc[zone, tier]
                if rate > 0:
                    format_currency(ws_rates, current_row, col, rate)
        current_row += 1

    current_row += 2

# Column widths
ws_rates.column_dimensions['A'].width = 12
for col in range(2, 11):
    ws_rates.column_dimensions[get_column_letter(col)].width = 11

# ============================================================================
# SHEET 6: FIRSTMILE ADVANTAGES
# ============================================================================
ws_benefits = wb.create_sheet("FirstMile Advantages")

create_header_cell(ws_benefits, 1, 1, "WHY FIRSTMILE FOR ECOMMERCE SHIPPERS", merge_to=1)
ws_benefits.row_dimensions[1].height = 30

benefits = [
    ("", ""),
    ("Cost Optimization", ""),
    ("", f"12% average savings on your current volume (${monthly_savings:,.2f}/month)"),
    ("", "No hidden fees or surcharges"),
    ("", "Transparent pricing with rate card guarantee"),
    ("", "Volume discounts available as you scale"),
    ("", ""),
    ("Operational Excellence", ""),
    ("", "Audit Queue: Blocks mis-rated labels before invoicing"),
    ("", "Single integration: One connection for all services"),
    ("", "Dynamic routing: Auto-selects best carrier for each shipment"),
    ("", "Returns management: Simplified return label generation"),
    ("", ""),
    ("Service & Support", ""),
    ("", "Single support thread for all carriers and issues"),
    ("", "Claims management included (no separate carrier calls)"),
    ("", "Dedicated account manager for your business"),
    ("", "Real-time tracking across all carriers"),
    ("", ""),
    ("FirstMile Xparcel Service Levels", ""),
    ("", "Xparcel Ground (3-8 days): Economy ground service"),
    ("", "Xparcel Expedited (2-5 days): Faster ground for 1-20 lb"),
    ("", "Xparcel Priority (1-3 days): Premium with money-back guarantee"),
]

row = 3
for label, value in benefits:
    if label:
        cell = ws_benefits.cell(row, 1, label)
        cell.font = Font(size=12, bold=True, color=FM_BLUE)
    else:
        ws_benefits.cell(row, 1, value)
    row += 1

ws_benefits.column_dimensions['A'].width = 80

# Save workbook
output_file = f"Joshs_Frogs_FirstMile_Xparcel_Savings_Analysis_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
wb.save(output_file)

print()
print(f"Professional analysis saved: {output_file}")
print()
print("="*80)
print("ANALYSIS COMPLETE")
print("="*80)
print()
print(f"Monthly Savings: ${monthly_savings:,.2f}")
print(f"Annual Savings: ${annual_savings:,.2f}")
print(f"Savings Rate: {savings_rate*100:.1f}%")
print()
