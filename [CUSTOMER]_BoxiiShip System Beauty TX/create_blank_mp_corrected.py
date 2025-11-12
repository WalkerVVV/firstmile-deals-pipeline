import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

print("=" * 70)
print("BoxiiShip BLANK -MP Tracking Number Excel Generator (CORRECTED)")
print("=" * 70)

# Read the BLANK -MP tracking list
print("\nReading BLANK -MP tracking number list...")
with open('blank_mp_tracking_list.txt', 'r') as f:
    tracking_list = [line.strip() for line in f.readlines() if line.strip()]

print(f"Total tracking numbers in list: {len(tracking_list)}")

# Remove duplicates while preserving order
tracking_set = []
seen = set()
for tracking in tracking_list:
    if tracking not in seen:
        tracking_set.append(tracking)
        seen.add(tracking)

print(f"Unique tracking numbers: {len(tracking_set)}")

# Read the Labels _ Boxi sheet
print("\nReading tracking data from BoxiiShip Excel file...")
df = pd.read_excel('BoxiiShip-System Beauty _11_3_25.xlsx', sheet_name='Labels _ Boxi')

print(f"Total rows in Excel: {len(df)}")

# Filter for only the tracking numbers in our list
df['TrackingNumber_str'] = df['TrackingNumber'].astype(str)
df_filtered = df[df['TrackingNumber_str'].isin(tracking_set)].copy()

print(f"Matched tracking numbers: {len(df_filtered)}")

# Create clean dataframe with relevant columns
df_clean = pd.DataFrame({
    'Tracking Number': df_filtered['TrackingNumber'].astype(str),
    'Status': df_filtered['FM_Tracking Status_BOXI ISSUES _11_3_2025'],
    'Date': df_filtered['DATE_BOXI ISSUES _11_3_2025'],
    'Location': ''  # Location not provided in this source
})

# Show status distribution
print(f"\n--- Status Distribution ---")
status_counts = df_clean['Status'].value_counts()
print(status_counts.head(20))

# Count specific statuses
shipment_info = len(df_clean[df_clean['Status'].str.contains('Shipment info received', case=False, na=False)])
delivered = len(df_clean[df_clean['Status'].str.contains('Delivered', case=False, na=False)])
blank_status = len(df_clean[df_clean['Status'].isna()])

print(f"\n--- Summary ---")
print(f"Total Tracking Numbers: {len(df_clean)}")
print(f"Blank/No Status: {blank_status} ({blank_status/len(df_clean)*100:.1f}%)")
print(f"Shipment info received: {shipment_info} ({shipment_info/len(df_clean)*100:.1f}%)")
print(f"Delivered: {delivered} ({delivered/len(df_clean)*100:.1f}%)")

# Save to Excel with professional formatting
output_file = 'BLANK_MP_Tracking_Nov4_2025.xlsx'

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_clean.to_excel(writer, index=False, sheet_name='BLANK -MP Tracking')

    # Get the worksheet
    worksheet = writer.sheets['BLANK -MP Tracking']

    # Set column widths
    worksheet.column_dimensions['A'].width = 28  # Tracking Number
    worksheet.column_dimensions['B'].width = 50  # Status
    worksheet.column_dimensions['C'].width = 25  # Date
    worksheet.column_dimensions['D'].width = 35  # Location

    # Format header row (FirstMile blue #366092)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Center align tracking numbers and dates
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    # Auto-filter
    worksheet.auto_filter.ref = worksheet.dimensions

print(f"\n[SUCCESS] Excel file created: {output_file}")
print(f"[SUCCESS] Contains {len(df_clean)} BLANK -MP tracking numbers")
print("=" * 70)
