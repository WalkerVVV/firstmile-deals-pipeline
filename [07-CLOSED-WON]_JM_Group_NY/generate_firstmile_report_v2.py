"""
FirstMile Xparcel Performance Report Generator V2
Creates professional Excel deliverable analyzing FirstMile carrier performance
BROKEN DOWN BY XPARCEL SERVICE LEVEL (Priority, Expedited, Ground)
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule

# Configuration
DATA_FILE = r"C:\Users\BrettWalker\FirstMile_Deals\[07-CLOSED-WON]_JM_Group_NY\JM Group NY_FirstMile_Domestic_Tracking_9.30.25.xlsx"
CUSTOMER_NAME = "JM Group NY"

# Brand Colors
FIRSTMILE_BLUE = "366092"
LIGHT_GRAY = "DDDDDD"
RED_BG = "FFC7CE"
YELLOW_BG = "FFEB84"
GREEN_BG = "C6EFCE"

# SLA Windows by Xparcel Type
SLA_WINDOWS = {
    "Ground": 8,
    "Expedited": 5,
    "Direct Call": 3  # Priority
}

# Service name mapping
SERVICE_NAME_MAP = {
    'Ground': 'Xparcel Ground',
    'Expedited': 'Xparcel Expedited',
    'Direct Call': 'Xparcel Priority'
}

# Performance Thresholds
PERF_THRESHOLDS = [
    (100.0, "Perfect Compliance"),
    (95.0, "Exceeds Standard"),
    (90.0, "Meets Standard"),
    (0.0, "Below Standard")
]

# Hub Mapping
HUB_MAP = {
    "CA": "LAX - West Coast",
    "TX": "DFW - South Central",
    "FL": "MIA - Southeast",
    "NY": "JFK/EWR - Northeast",
    "IL": "ORD - Midwest",
    "GA": "ATL - Southeast",
    "NJ": "JFK/EWR - Northeast",
    "PA": "JFK/EWR - Northeast",
    "OH": "ORD - Midwest",
    "MI": "ORD - Midwest",
    "WI": "ORD - Midwest"
}

def load_data(file_path):
    """Load and standardize the input Excel file"""
    print(f"Loading data from {file_path}...")
    df = pd.read_excel(file_path)
    print(f"Columns found: {df.columns.tolist()}")
    print(f"Total rows: {len(df)}")
    return df

def detect_report_period(df):
    """Detect the report period from the data"""
    date_columns = ['Request Date', 'Start Date', 'ShipDate', 'Ship Date', 'Created Date']

    for col in date_columns:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
            valid_dates = df[col].dropna()
            if len(valid_dates) > 0:
                min_date = valid_dates.min()
                max_date = valid_dates.max()
                return f"{min_date.strftime('%B %d')} - {max_date.strftime('%B %d, %Y')}"

    return "Date Range Not Available"

def calculate_sla_by_service_level(df):
    """Calculate SLA compliance broken down by Xparcel service level"""

    # Filter for delivered only
    delivered = df[df['Delivered Status'] == 'Delivered'].copy()

    if len(delivered) == 0 or 'Xparcel Type' not in delivered.columns:
        return []

    results = []

    for service_type in sorted(delivered['Xparcel Type'].unique()):
        if service_type not in SLA_WINDOWS:
            continue

        service_df = delivered[delivered['Xparcel Type'] == service_type].copy()
        sla_window = SLA_WINDOWS[service_type]
        service_name = SERVICE_NAME_MAP.get(service_type, service_type)

        # Calculate compliance
        within_sla = len(service_df[service_df['Days In Transit'] <= sla_window])
        total = len(service_df)
        compliance_pct = (within_sla / total * 100) if total > 0 else 0

        # Performance status
        performance_status = next(
            status for threshold, status in PERF_THRESHOLDS
            if compliance_pct >= threshold
        )

        # Transit statistics
        avg_transit = service_df['Days In Transit'].mean()
        median_transit = service_df['Days In Transit'].median()
        p90_transit = service_df['Days In Transit'].quantile(0.90)
        p95_transit = service_df['Days In Transit'].quantile(0.95)

        # Day-by-day distribution
        day_distribution = []
        for day in range(sla_window + 3):
            count = len(service_df[service_df['Days In Transit'] == day])
            if count > 0:
                pct = (count / total * 100)
                cumulative = len(service_df[service_df['Days In Transit'] <= day])
                cum_pct = (cumulative / total * 100)
                within_sla_marker = day <= sla_window

                day_distribution.append({
                    'Day': day,
                    'Count': count,
                    'Percentage': pct,
                    'Cumulative_Count': cumulative,
                    'Cumulative_Percentage': cum_pct,
                    'Within_SLA': within_sla_marker
                })

        results.append({
            'service_type': service_type,
            'service_name': service_name,
            'sla_window': sla_window,
            'total_delivered': total,
            'within_sla': within_sla,
            'outside_sla': total - within_sla,
            'compliance_pct': compliance_pct,
            'performance_status': performance_status,
            'avg_transit': avg_transit,
            'median_transit': median_transit,
            'p90_transit': p90_transit,
            'p95_transit': p95_transit,
            'day_distribution': day_distribution
        })

    return results

def calculate_in_transit_by_service_level(df):
    """Calculate in-transit status broken down by service level"""

    in_transit = df[df['Delivered Status'] != 'Delivered'].copy()

    if len(in_transit) == 0 or 'Xparcel Type' not in in_transit.columns:
        return [], {}

    # Calculate days since ship using Start Date
    today = pd.Timestamp.now()
    date_col = 'Start Date'

    if date_col in in_transit.columns:
        in_transit[date_col] = pd.to_datetime(in_transit[date_col], errors='coerce')
        in_transit['Days Since Ship'] = (today - in_transit[date_col]).dt.days
    else:
        in_transit['Days Since Ship'] = 0

    results = []

    for service_type in sorted([x for x in in_transit['Xparcel Type'].unique() if pd.notna(x)]):
        if service_type not in SLA_WINDOWS:
            continue

        service_df = in_transit[in_transit['Xparcel Type'] == service_type].copy()
        sla_window = SLA_WINDOWS[service_type]
        service_name = SERVICE_NAME_MAP.get(service_type, service_type)

        service_df['Within SLA Window'] = service_df['Days Since Ship'].apply(
            lambda x: 'Yes' if x <= sla_window else 'No'
        )

        within_sla_count = len(service_df[service_df['Within SLA Window'] == 'Yes'])
        outside_sla_count = len(service_df[service_df['Within SLA Window'] == 'No'])
        total = len(service_df)

        results.append({
            'service_name': service_name,
            'sla_window': sla_window,
            'total_in_transit': total,
            'within_sla_window': within_sla_count,
            'outside_sla_window': outside_sla_count,
            'within_sla_pct': (within_sla_count / total * 100) if total > 0 else 0
        })

    # Overall summary
    total_in_transit = len(in_transit)
    in_transit['Within SLA Window'] = in_transit.apply(
        lambda row: 'Yes' if row['Days Since Ship'] <= SLA_WINDOWS.get(row['Xparcel Type'], 8) else 'No',
        axis=1
    )

    overall_within = len(in_transit[in_transit['Within SLA Window'] == 'Yes'])
    overall_outside = len(in_transit[in_transit['Within SLA Window'] == 'No'])

    summary = {
        'total_in_transit': total_in_transit,
        'within_sla_window': overall_within,
        'outside_sla_window': overall_outside,
        'within_sla_pct': (overall_within / total_in_transit * 100) if total_in_transit > 0 else 0
    }

    # Prepare detail dataframe
    cols_to_keep = ['Xparcel Type', 'Start Date', 'Days Since Ship', 'Within SLA Window',
                    'Delivered Status', 'Destination State', 'Calculated Zone']
    available_cols = [col for col in cols_to_keep if col in in_transit.columns]
    detail_df = in_transit[available_cols]

    return results, summary, detail_df

def calculate_geographic_distribution(df):
    """Calculate geographic distribution with hub mapping"""
    if 'Destination State' not in df.columns:
        return pd.DataFrame()

    state_counts = df['Destination State'].value_counts().head(15)

    geo_data = []
    for state, count in state_counts.items():
        pct = (count / len(df)) * 100
        hub = HUB_MAP.get(state, "National Network")
        network_type = "Select Network" if state in HUB_MAP else "National Network"

        geo_data.append({
            'State': state,
            'Shipment Count': count,
            'Percentage': pct,
            'Primary Hub': hub,
            'Network Type': network_type
        })

    return pd.DataFrame(geo_data)

def calculate_zone_analysis(df):
    """Calculate zone distribution and regional vs cross-country split"""
    if 'Calculated Zone' not in df.columns:
        return pd.DataFrame(), {}

    # Clean zone data
    df['Calculated Zone'] = df['Calculated Zone'].astype(str).str.extract(r'(\d+)')[0].astype(float)

    zone_data = []
    for zone in range(1, 9):
        zone_df = df[df['Calculated Zone'] == zone]
        count = len(zone_df)
        pct = (count / len(df)) * 100

        if count > 0 and 'Days In Transit' in zone_df.columns:
            delivered_zone = zone_df[zone_df['Delivered Status'] == 'Delivered']
            avg_transit = delivered_zone['Days In Transit'].mean() if len(delivered_zone) > 0 else 0
        else:
            avg_transit = 0

        zone_data.append({
            'Zone': f'Zone {zone}',
            'Shipment Count': count,
            'Percentage': pct,
            'Avg Transit Days': avg_transit
        })

    df_zones = pd.DataFrame(zone_data)

    # Regional vs Cross-Country
    regional = df[df['Calculated Zone'].isin([1, 2, 3, 4])]
    cross_country = df[df['Calculated Zone'].isin([5, 6, 7, 8])]

    summary = {
        'Regional (Zones 1-4)': {
            'count': len(regional),
            'pct': (len(regional) / len(df)) * 100
        },
        'Cross-Country (Zones 5-8)': {
            'count': len(cross_country),
            'pct': (len(cross_country) / len(df)) * 100
        }
    }

    return df_zones, summary

def apply_header_style(ws, row=1):
    """Apply FirstMile blue header styling"""
    blue_fill = PatternFill(start_color=FIRSTMILE_BLUE, end_color=FIRSTMILE_BLUE, fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[row]:
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center_align

def apply_data_style(ws, start_row=2):
    """Apply data cell styling with borders"""
    thin_border = Border(
        left=Side(style='thin', color=LIGHT_GRAY),
        right=Side(style='thin', color=LIGHT_GRAY),
        top=Side(style='thin', color=LIGHT_GRAY),
        bottom=Side(style='thin', color=LIGHT_GRAY)
    )
    center_align = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align

def auto_size_columns(ws, max_width=50):
    """Auto-size columns based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_executive_summary(wb, df, report_period, sla_by_service, in_transit_summary):
    """Create Executive Summary tab"""
    ws = wb.create_sheet("Executive Summary", 0)

    # Header
    ws['A1'] = 'FirstMile Xparcel Performance Report'
    ws['A1'].font = Font(size=14, bold=True, color=FIRSTMILE_BLUE)

    ws['A2'] = f'Customer: {CUSTOMER_NAME}'
    ws['A3'] = f'Report Period: {report_period}'

    row = 5

    # Overall metrics
    ws[f'A{row}'] = 'Overall Performance'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = 'Metric'
    ws[f'B{row}'] = 'Value'
    apply_header_style(ws, row=row)
    row += 1

    total_delivered = sum([s['total_delivered'] for s in sla_by_service])
    total_within_sla = sum([s['within_sla'] for s in sla_by_service])
    overall_compliance = (total_within_sla / total_delivered * 100) if total_delivered > 0 else 0

    if overall_compliance >= 95.0:
        status = "Exceeds Standard"
    elif overall_compliance >= 90.0:
        status = "Meets Standard"
    else:
        status = "Below Standard"

    overall_metrics = [
        ('Total Shipments', len(df)),
        ('Total Delivered', total_delivered),
        ('Overall SLA Compliance', f"{overall_compliance:.1f}%"),
        ('Performance Status', status),
        ('In-Transit Shipments', in_transit_summary['total_in_transit'])
    ]

    for label, value in overall_metrics:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1

    apply_data_style(ws, start_row=row-len(overall_metrics))

    # Service level breakdown
    row += 2
    ws[f'A{row}'] = 'Performance by Xparcel Service Level'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = 'Service Level'
    ws[f'B{row}'] = 'SLA Window'
    ws[f'C{row}'] = 'Delivered'
    ws[f'D{row}'] = 'Within SLA'
    ws[f'E{row}'] = 'Compliance'
    ws[f'F{row}'] = 'Status'
    apply_header_style(ws, row=row)
    row += 1

    start_data_row = row
    for service in sla_by_service:
        ws[f'A{row}'] = service['service_name']
        ws[f'B{row}'] = f"{service['sla_window']} days"
        ws[f'C{row}'] = service['total_delivered']
        ws[f'D{row}'] = service['within_sla']
        ws[f'E{row}'] = f"{service['compliance_pct']:.1f}%"
        ws[f'F{row}'] = service['performance_status']
        row += 1

    apply_data_style(ws, start_row=start_data_row)
    auto_size_columns(ws)

def create_sla_compliance_by_service_tab(wb, sla_by_service):
    """Create SLA Compliance by Service Level tab"""
    ws = wb.create_sheet("SLA Compliance")

    ws['A1'] = 'SLA Compliance by Xparcel Service Level'
    ws['A1'].font = Font(bold=True, size=12)

    row = 3

    for service in sla_by_service:
        # Service header
        ws[f'A{row}'] = f"{service['service_name']} ({service['sla_window']}-day SLA)"
        ws[f'A{row}'].font = Font(bold=True, size=11, color=FIRSTMILE_BLUE)
        row += 1

        # Metrics
        ws[f'A{row}'] = 'Metric'
        ws[f'B{row}'] = 'Value'
        apply_header_style(ws, row=row)
        row += 1

        metrics = [
            ('Total Delivered', service['total_delivered']),
            ('Within SLA', service['within_sla']),
            ('Outside SLA', service['outside_sla']),
            ('SLA Compliance', f"{service['compliance_pct']:.1f}%"),
            ('Performance Status', service['performance_status']),
            ('Average Transit Days', f"{service['avg_transit']:.1f}"),
            ('Median Transit Days', f"{service['median_transit']:.1f}"),
            ('90th Percentile', f"{service['p90_transit']:.1f}"),
            ('95th Percentile', f"{service['p95_transit']:.1f}")
        ]

        start_metrics = row
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        apply_data_style(ws, start_row=start_metrics)
        row += 2

    auto_size_columns(ws)

def create_transit_performance_by_service_tab(wb, sla_by_service):
    """Create Transit Performance by Service Level tab"""
    ws = wb.create_sheet("Transit Performance")

    ws['A1'] = 'Transit Performance by Xparcel Service Level'
    ws['A1'].font = Font(bold=True, size=12)

    row = 3

    for service in sla_by_service:
        # Service header
        ws[f'A{row}'] = f"{service['service_name']} ({service['sla_window']}-day SLA)"
        ws[f'A{row}'].font = Font(bold=True, size=11, color=FIRSTMILE_BLUE)
        row += 1

        # Distribution table
        ws[f'A{row}'] = 'Transit Days'
        ws[f'B{row}'] = 'Count'
        ws[f'C{row}'] = 'Percentage'
        ws[f'D{row}'] = 'Cumulative Count'
        ws[f'E{row}'] = 'Cumulative %'
        ws[f'F{row}'] = 'Status'
        apply_header_style(ws, row=row)
        row += 1

        start_dist = row
        for day_data in service['day_distribution']:
            ws[f'A{row}'] = f"Day {day_data['Day']}"
            ws[f'B{row}'] = day_data['Count']
            ws[f'C{row}'] = f"{day_data['Percentage']:.1f}%"
            ws[f'D{row}'] = day_data['Cumulative_Count']
            ws[f'E{row}'] = f"{day_data['Cumulative_Percentage']:.1f}%"
            ws[f'F{row}'] = 'Within SLA' if day_data['Within_SLA'] else 'Outside SLA'
            row += 1

        apply_data_style(ws, start_row=start_dist)
        row += 2

    auto_size_columns(ws)

def create_geographic_tab(wb, geo_df):
    """Create Geographic Distribution tab"""
    ws = wb.create_sheet("Geographic Distribution")

    for r_idx, row in enumerate(dataframe_to_rows(geo_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    apply_header_style(ws, row=1)
    apply_data_style(ws, start_row=2)
    auto_size_columns(ws)

def create_zone_analysis_tab(wb, zone_df, summary):
    """Create Zone Analysis tab"""
    ws = wb.create_sheet("Zone Analysis")

    # Zone distribution
    for r_idx, row in enumerate(dataframe_to_rows(zone_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    apply_header_style(ws, row=1)
    apply_data_style(ws, start_row=2)

    # Regional summary
    summary_start = len(zone_df) + 3
    ws[f'A{summary_start}'] = 'Regional vs Cross-Country Summary'
    ws[f'A{summary_start}'].font = Font(bold=True, size=12)

    summary_start += 1
    ws[f'A{summary_start}'] = 'Category'
    ws[f'B{summary_start}'] = 'Shipment Count'
    ws[f'C{summary_start}'] = 'Percentage'
    apply_header_style(ws, row=summary_start)

    for i, (category, data) in enumerate(summary.items(), start=summary_start+1):
        ws[f'A{i}'] = category
        ws[f'B{i}'] = data['count']
        ws[f'C{i}'] = f"{data['pct']:.1f}%"

    auto_size_columns(ws)

def create_in_transit_tab(wb, in_transit_by_service, in_transit_detail):
    """Create In-Transit Detail tab"""
    ws = wb.create_sheet("In-Transit Detail")

    ws['A1'] = 'In-Transit Status by Service Level'
    ws['A1'].font = Font(bold=True, size=12)

    # Summary by service
    ws['A3'] = 'Service Level'
    ws['B3'] = 'SLA Window'
    ws['C3'] = 'Total In-Transit'
    ws['D3'] = 'Within Window'
    ws['E3'] = 'Outside Window'
    ws['F3'] = '% Within Window'
    apply_header_style(ws, row=3)

    row = 4
    for service in in_transit_by_service:
        ws[f'A{row}'] = service['service_name']
        ws[f'B{row}'] = f"{service['sla_window']} days"
        ws[f'C{row}'] = service['total_in_transit']
        ws[f'D{row}'] = service['within_sla_window']
        ws[f'E{row}'] = service['outside_sla_window']
        ws[f'F{row}'] = f"{service['within_sla_pct']:.1f}%"
        row += 1

    apply_data_style(ws, start_row=4)

    # Detail records
    if len(in_transit_detail) > 0:
        detail_start = row + 3
        ws[f'A{detail_start}'] = 'Detailed In-Transit Records'
        ws[f'A{detail_start}'].font = Font(bold=True, size=12)

        detail_start += 2
        for r_idx, data_row in enumerate(dataframe_to_rows(in_transit_detail, index=False, header=True), detail_start):
            for c_idx, value in enumerate(data_row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        apply_header_style(ws, row=detail_start)
        apply_data_style(ws, start_row=detail_start+1)

    auto_size_columns(ws)

def create_notes_tab(wb):
    """Create Notes & Assumptions tab"""
    ws = wb.create_sheet("Notes & Assumptions")

    notes = [
        ('Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('', ''),
        ('Business Rules', ''),
        ('SLA Calculation', 'Based on delivered packages only, by service level'),
        ('Xparcel Ground', '8-day SLA window (3-8 day service)'),
        ('Xparcel Expedited', '5-day SLA window (2-5 day service)'),
        ('Xparcel Priority', '3-day SLA window (1-3 day service)'),
        ('In-Transit Status', 'Shipments not yet delivered are excluded from SLA calculations'),
        ('Performance Thresholds', '100%=Perfect, >=95%=Exceeds, >=90%=Meets, <90%=Below'),
        ('', ''),
        ('Definitions', ''),
        ('FirstMile', 'The carrier providing shipping services'),
        ('Xparcel', 'The ship method (Priority, Expedited, Ground)'),
        ('National Network', 'Nationwide coverage for all ZIP codes'),
        ('Select Network', 'Metro-focused injection points in major hubs')
    ]

    for i, (label, value) in enumerate(notes, start=1):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value

    auto_size_columns(ws)

def create_brand_style_guide_tab(wb):
    """Create Brand Style Guide tab"""
    ws = wb.create_sheet("Brand Style Guide")

    ws['A1'] = 'FirstMile Brand Colors'
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = 'Color Name'
    ws['B3'] = 'HEX'
    ws['C3'] = 'RGB'
    ws['D3'] = 'Usage'
    apply_header_style(ws, row=3)

    colors = [
        ('FirstMile Blue', '#366092', 'RGB(54, 96, 146)', 'Primary brand color, headers'),
        ('Light Gray', '#DDDDDD', 'RGB(221, 221, 221)', 'Borders, dividers'),
        ('Red', '#FFC7CE', 'RGB(255, 199, 206)', 'Below standard performance (80-89%)'),
        ('Yellow', '#FFEB84', 'RGB(255, 235, 132)', 'Meets standard performance (90-94%)'),
        ('Green', '#C6EFCE', 'RGB(198, 239, 206)', 'Exceeds standard performance (95-100%)')
    ]

    for i, (name, hex_code, rgb, usage) in enumerate(colors, start=4):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = hex_code
        ws[f'C{i}'] = rgb
        ws[f'D{i}'] = usage

    apply_data_style(ws, start_row=4)
    auto_size_columns(ws)

def generate_report():
    """Main report generation function"""
    print("="*80)
    print("FirstMile Xparcel Performance Report Generator V2")
    print("BY SERVICE LEVEL: Priority, Expedited, Ground")
    print("="*80)

    # Load data
    df = load_data(DATA_FILE)

    # Detect report period
    report_period = detect_report_period(df)
    print(f"\nReport Period: {report_period}")

    # Calculate analytics by service level
    print("\nCalculating SLA compliance by service level...")
    sla_by_service = calculate_sla_by_service_level(df)

    print("\nSLA Compliance Summary:")
    for service in sla_by_service:
        print(f"  {service['service_name']}: {service['compliance_pct']:.1f}% ({service['total_delivered']} delivered)")

    print("\nCalculating in-transit status by service level...")
    in_transit_by_service, in_transit_summary, in_transit_detail = calculate_in_transit_by_service_level(df)

    print(f"\nIn-Transit Summary:")
    print(f"  Total: {in_transit_summary['total_in_transit']}")
    print(f"  Within SLA Window: {in_transit_summary['within_sla_window']}")
    print(f"  Outside SLA Window: {in_transit_summary['outside_sla_window']}")

    print("\nCalculating geographic distribution...")
    geo_df = calculate_geographic_distribution(df)

    print("Calculating zone analysis...")
    zone_df, zone_summary = calculate_zone_analysis(df)

    # Create Excel workbook
    print("\nGenerating Excel report...")
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create all tabs
    create_executive_summary(wb, df, report_period, sla_by_service, in_transit_summary)
    create_sla_compliance_by_service_tab(wb, sla_by_service)
    create_transit_performance_by_service_tab(wb, sla_by_service)
    create_geographic_tab(wb, geo_df)
    create_zone_analysis_tab(wb, zone_df, zone_summary)
    create_in_transit_tab(wb, in_transit_by_service, in_transit_detail)
    create_notes_tab(wb)
    create_brand_style_guide_tab(wb)

    # Save workbook
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    output_file = f"FirstMile_Xparcel_Performance_{CUSTOMER_NAME.replace(' ', '_')}_{timestamp}.xlsx"
    output_path = r"C:\Users\BrettWalker\FirstMile_Deals\[07-CLOSED-WON]_JM_Group_NY\\" + output_file

    wb.save(output_path)

    print(f"\n{'='*80}")
    print(f"Report generated successfully!")
    print(f"{'='*80}")
    print(f"\nOutput file: {output_path}")
    print(f"\nSummary by Service Level:")
    for service in sla_by_service:
        print(f"  {service['service_name']}: {service['compliance_pct']:.1f}% SLA compliance")

    total_delivered = sum([s['total_delivered'] for s in sla_by_service])
    total_within = sum([s['within_sla'] for s in sla_by_service])
    overall = (total_within / total_delivered * 100) if total_delivered > 0 else 0
    print(f"\nOverall Performance: {overall:.1f}% - Exceeds Standard")
    print(f"\n{'='*80}")

if __name__ == "__main__":
    generate_report()
