"""
Create Volume Distribution Matrix by Weight and Service Level for Upstate Prep
Direct Excel output with formatting and multiple sheets
Service Levels: Xparcel Priority, Xparcel Expedited, Xparcel Ground
"""

import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the data
df = pd.read_csv('T30 PLD Upstate Prep.csv')

# Clean and prepare data
df['Weight (lb)'] = pd.to_numeric(df['Weight (lb)'], errors='coerce')
df['Weight (oz)'] = df['Weight (lb)'] * 16

# Map current services to Xparcel service levels
service_mapping = {
    # Ground Services → Xparcel Ground
    'UPS Ground': 'Xparcel Ground',
    'UPS Standard': 'Xparcel Ground',
    'UPS SurePost': 'Xparcel Ground',
    'Amazon Shipping Ground': 'Xparcel Ground',
    'usps_ground_advantage': 'Xparcel Ground',
    'Fedex Ground Services': 'Xparcel Ground',
    'FEDEX_GROUND': 'Xparcel Ground',
    'Generic': 'Xparcel Ground',
    'Generic Label': 'Xparcel Ground',
    'genericlabel': 'Xparcel Ground',
    'Wholesale Generic': 'Xparcel Ground',
    
    # Expedited Services → Xparcel Expedited  
    'UPS 2nd Day Air': 'Xparcel Expedited',
    'UPS 3 Day Select': 'Xparcel Expedited',
    'UPS Expedited': 'Xparcel Expedited',
    'usps_priority': 'Xparcel Expedited',
    'FedEx International Connect Plus': 'Xparcel Expedited',
    
    # Priority/Express Services → Xparcel Priority
    'UPS Next Day Air': 'Xparcel Priority',
    'UPS Next Day Air Saver': 'Xparcel Priority',
    'UPS Next Day Air Early': 'Xparcel Priority',
    'UPS Express': 'Xparcel Priority',
    'UPS Saver': 'Xparcel Priority',
    'Domestic Express Doc': 'Xparcel Priority',
    'Express Worldwide NonDoc': 'Xparcel Priority',
    
    # International and Other
    'usps_first_class_package_international_service': 'Xparcel Ground',
    'usps_priority_mail_international': 'Xparcel Expedited',
    
    # LTL stays separate for now
    'LTL': 'Xparcel Ground'
}

# Apply service mapping
df['Xparcel Service'] = df['Shipping Method'].map(service_mapping).fillna('Xparcel Ground')

# Create billable weight based on carrier rules
def calculate_billable_weight(weight_oz):
    if pd.isna(weight_oz):
        return np.nan
    if weight_oz < 16:
        # Under 1 lb: round up to next oz, but 15.99 is special tier
        if weight_oz > 15 and weight_oz < 16:
            return 15.99
        else:
            return np.ceil(weight_oz)
    else:
        # Over 1 lb: round up to next pound, convert to oz
        return np.ceil(weight_oz / 16) * 16

df['Billable Weight (oz)'] = df['Weight (oz)'].apply(calculate_billable_weight)

# Create complete weight list
weight_list = []
# Under 1 lb (1-15 oz)
for oz in range(1, 16):
    weight_list.append((f"{oz}oz", oz))
# Special 15.99oz tier
weight_list.append(("15.99oz", 15.99))
# 16oz (1 lb) up to 400oz (25 lbs) in 1 lb increments
for lb in range(1, 26):  # 1-25 lbs
    weight_list.append((f"{lb} lb", lb*16))

# Initialize the matrix
services = ['Xparcel Priority', 'Xparcel Expedited', 'Xparcel Ground']
matrix_data = []

# Create matrix
for weight_label, weight_val in weight_list:
    row = {'Weight': weight_label}
    
    for service in services:
        # Count packages in this weight and service level
        if weight_val == 15.99:
            count = len(df[(df['Billable Weight (oz)'] == 15.99) & (df['Xparcel Service'] == service)])
        else:
            count = len(df[(df['Billable Weight (oz)'] == weight_val) & (df['Xparcel Service'] == service)])
        
        row[service] = count
    
    # Add total for this weight
    if weight_val == 15.99:
        row['Total'] = len(df[df['Billable Weight (oz)'] == 15.99])
    else:
        row['Total'] = len(df[df['Billable Weight (oz)'] == weight_val])
    
    matrix_data.append(row)

# Create DataFrame
matrix_df = pd.DataFrame(matrix_data)

# Calculate service totals
service_totals = {'Weight': 'TOTAL'}
for service in services:
    service_totals[service] = matrix_df[service].sum()
service_totals['Total'] = matrix_df['Total'].sum()

# Add totals row to matrix
matrix_df_with_totals = pd.concat([matrix_df, pd.DataFrame([service_totals])], ignore_index=True)

# Create summary statistics DataFrame
summary_data = []
for service in services:
    pct = (service_totals[service] / service_totals['Total']) * 100 if service_totals['Total'] > 0 else 0
    summary_data.append({
        'Service Level': service,
        'Volume': int(service_totals[service]),
        'Percentage': f"{pct:.1f}%"
    })
summary_df = pd.DataFrame(summary_data)

# Find top weights with mix
top_weights_data = []
top_weights = matrix_df.nlargest(10, 'Total')
for _, row in top_weights.iterrows():
    if row['Total'] > 0:
        weight_row = {
            'Weight': row['Weight'],
            'Total Volume': int(row['Total'])
        }
        for service in services:
            if row[service] > 0:
                pct = (row[service] / row['Total']) * 100
                weight_row[f"{service}"] = f"{int(row[service])} ({pct:.1f}%)"
            else:
                weight_row[f"{service}"] = "0 (0.0%)"
        top_weights_data.append(weight_row)
top_weights_df = pd.DataFrame(top_weights_data)

# Create Excel file with multiple sheets
excel_filename = 'upstate_prep_volume_matrix.xlsx'
writer = pd.ExcelWriter(excel_filename, engine='openpyxl')

# Sheet 1: Volume Matrix
matrix_df_with_totals.to_excel(writer, sheet_name='Volume Matrix', index=False)

# Sheet 2: Summary Statistics
summary_df.to_excel(writer, sheet_name='Summary', index=False)

# Sheet 3: Top Weight Categories
top_weights_df.to_excel(writer, sheet_name='Top Weights', index=False)

# Sheet 4: Pivot Analysis - Service by Weight Range
weight_ranges = []
for _, row in df.iterrows():
    weight = row['Weight (lb)']
    if pd.isna(weight):
        weight_ranges.append('Unknown')
    elif weight < 1:
        weight_ranges.append('Under 1 lb')
    elif weight <= 5:
        weight_ranges.append('1-5 lbs')
    elif weight <= 10:
        weight_ranges.append('6-10 lbs')
    elif weight <= 20:
        weight_ranges.append('11-20 lbs')
    else:
        weight_ranges.append('Over 20 lbs')

df['Weight Range'] = weight_ranges
pivot_df = df.pivot_table(
    values='Shipping Label ID',
    index='Weight Range',
    columns='Xparcel Service',
    aggfunc='count',
    fill_value=0,
    margins=True,
    margins_name='Total'
)
pivot_df.to_excel(writer, sheet_name='Weight Range Analysis')

# Save the Excel file
writer.close()

# Now apply formatting with openpyxl
workbook = openpyxl.load_workbook(excel_filename)

# Define styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
number_alignment = Alignment(horizontal="center")

# Format Sheet 1: Volume Matrix
ws1 = workbook['Volume Matrix']
for cell in ws1[1]:  # Header row
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Format data rows
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
    for cell in row:
        cell.border = thin_border
        if cell.column > 1:  # Number columns
            cell.alignment = number_alignment
    
    # Highlight total row
    if row[0].value == "TOTAL":
        for cell in row:
            cell.fill = total_fill
            cell.font = Font(bold=True)

# Auto-adjust column widths
for column in ws1.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 25)
    ws1.column_dimensions[column_letter].width = adjusted_width

# Format Sheet 2: Summary
ws2 = workbook['Summary']
for cell in ws2[1]:  # Header row
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
    for cell in row:
        cell.border = thin_border
        if cell.column > 1:
            cell.alignment = number_alignment

# Auto-adjust column widths
for column in ws2.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 30)
    ws2.column_dimensions[column_letter].width = adjusted_width

# Format Sheet 3: Top Weights
ws3 = workbook['Top Weights']
for cell in ws3[1]:  # Header row
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=ws3.max_column):
    for cell in row:
        cell.border = thin_border
        if cell.column > 1:
            cell.alignment = number_alignment

# Auto-adjust column widths
for column in ws3.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 25)
    ws3.column_dimensions[column_letter].width = adjusted_width

# Format Sheet 4: Weight Range Analysis
ws4 = workbook['Weight Range Analysis']
for cell in ws4[1]:  # Header row
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row, min_col=1, max_col=ws4.max_column):
    for cell in row:
        cell.border = thin_border
        if cell.column > 1:
            cell.alignment = number_alignment
    
    # Highlight total row and column
    if row[0].value == "Total":
        for cell in row:
            cell.fill = total_fill
            cell.font = Font(bold=True)

# Auto-adjust column widths
for column in ws4.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 25)
    ws4.column_dimensions[column_letter].width = adjusted_width

# Save the formatted workbook
workbook.save(excel_filename)
workbook.close()

print("=" * 100)
print("UPSTATE PREP - VOLUME DISTRIBUTION EXCEL FILE CREATED")
print("=" * 100)
print()
print(f"Excel file saved: {excel_filename}")
print()
print("Contents:")
print("  Sheet 1: Volume Matrix - Complete volume by weight and service level")
print("  Sheet 2: Summary - Service level distribution summary")
print("  Sheet 3: Top Weights - Top 10 weight categories with service mix")
print("  Sheet 4: Weight Range Analysis - Pivot table by weight ranges")
print()
print(f"Total Packages Analyzed: {int(service_totals['Total']):,}")
print()
print("Service Level Distribution:")
for service in services:
    pct = (service_totals[service] / service_totals['Total']) * 100 if service_totals['Total'] > 0 else 0
    print(f"  {service}: {int(service_totals[service]):,} packages ({pct:.1f}%)")